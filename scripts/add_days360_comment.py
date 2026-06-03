"""
Add an Excel cell comment explaining days360(nextduedate, dataasof) to the Field Spec sheet
cells that reference it: days_past_due (格式/计算规则, col8) and next_payment_due_date (业务含义, col7).

Run AFTER add_glossary.py. Excel must be closed.
Run: python scripts/add_days360_comment.py
"""

import sys, io
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.comments import Comment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")

COMMENT_TEXT = (
    "days360 = 30/360 日历法天数差函数（每月30天/年360天）\n"
    "公式: (末.年-始.年)*360 + (末.月-始.月)*30 + (末.日-始.日)\n"
    "days360(nextduedate, dataasof) = 从应还款日到数据快照日的天数 = 逾期天数 DPD（逾期为正）\n"
    "ETL 分档: <30=C · <60=D30 · <90=D60 · <120=D90 · >=120=D120P（永不产生 FCL）\n"
    "示例: nextduedate=2024-01-01, dataasof=2024-04-01 → days360=90 → D90\n"
    "源码: PrefectFlow/flow/remit_validation/utils.py:14-21 ；详见「📚 术语说明 Glossary」sheet"
)

# field -> column index to attach the comment (where days360 appears)
TARGETS = {
    "days_past_due": 8,          # 格式/计算规则
    "next_payment_due_date": 7,  # 业务含义
}


def main():
    wb = load_workbook(XLSX)
    ws = None
    for s in wb.worksheets:
        if "Field Spec" in s.title:
            ws = s
            break

    added = 0
    for r in range(2, ws.max_row + 1):
        f = ws.cell(r, 3).value
        if f in TARGETS:
            col = TARGETS[f]
            c = Comment(COMMENT_TEXT, "ForeclosureRule2")
            c.width = 460
            c.height = 180
            ws.cell(r, col).comment = c
            print(f"[{f}] comment added on row {r} col {col}")
            added += 1

    wb.save(XLSX)
    print(f"\nSaved: {XLSX} — {added} comments added")


if __name__ == "__main__":
    main()
