# -*- coding: utf-8 -*-
"""
格式规则：「标准接口取值范围」(col9) 枚举值 ≥5 个时，一个取值占一行（Excel 用换行 + wrap）。
<5 个保持单行 ` | `。zh 卡片侧由 sync_fieldspec_excel_to_md.py 的 esc_ml() 把换行渲染为 <br>。

幂等：先把已有换行归一回 ` | ` 再判断/重排。前缀「…：」（如 MBA标准文本枚举…：/BPS 解码文本：）单独成行。
按表头定位列，人工列受保护。Excel 须关闭；改后运行 sync_fieldspec_excel_to_md.py。
Run: python scripts/fix_value_range_multiline.py
"""
import sys, io
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent))
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from _excel_guard import col_by_header, assert_safe

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")
THRESHOLD = 5


def one_per_line(v):
    """≥THRESHOLD 个枚举值 → 一值一行（\\n）；否则单行 ` | `。幂等。"""
    s = str(v).replace("\r", " ").replace("\n", " | ")
    # 折叠可能产生的多重分隔
    while "  | " in s or " |  " in s:
        s = s.replace("  | ", " | ").replace(" |  ", " | ")
    head, body = "", s
    if "：" in s:
        h, _, rest = s.partition("：")
        if " | " in rest:
            head, body = h + "：", rest
    parts = [p for p in body.split(" | ")]
    if len(parts) < THRESHOLD:
        return s  # 保持单行
    return (head + "\n" if head else "") + "\n".join(parts)


def main():
    wb = load_workbook(XLSX)
    ws = next(s for s in wb.worksheets if "Field Spec" in s.title)
    fcol = col_by_header(ws, "标准接口字段") or 3
    rcol = col_by_header(ws, "标准接口取值范围")
    if rcol is None:
        raise RuntimeError("找不到「标准接口取值范围」列")
    assert_safe(ws, rcol)

    changed = 0
    for r in range(2, ws.max_row + 1):
        f = ws.cell(r, fcol).value
        v = ws.cell(r, rcol).value
        if not isinstance(f, str) or not isinstance(v, str) or not v.strip():
            continue
        new = one_per_line(v)
        if new != v:
            cell = ws.cell(r, rcol)
            cell.value = new
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal or "left",
                                       vertical="top", wrap_text=True)
            changed += 1
            print(f"  {f:24} -> {new.count(chr(10)) + 1} 行")
    wb.save(XLSX)
    print(f"Saved: {XLSX}（{changed} 个字段改为一值一行）")


if __name__ == "__main__":
    main()
