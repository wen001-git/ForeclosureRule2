"""
End-to-end test of the 人工-column protection:
1. inject a test column header '人工备注' with a sentinel value,
2. run add_field_spec_verify_sql.py (header-resolved generator),
3. assert the sentinel survived AND the 验证SQL column was located by header (not shifted),
4. remove the test column and save (clean up).

Run: python scripts/_test_manual_guard.py   (Excel must be closed)
"""
import sys, io, subprocess
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent))
from openpyxl import load_workbook
from _excel_guard import next_free_col, col_by_header, manual_cols

XLSX = "docs/14_servicer_fcl_field_spec.xlsx"
SENTINEL = "USER-NOTE-SENTINEL"


def find_fieldspec(wb):
    for s in wb.worksheets:
        if "Field Spec" in s.title:
            return s


def inject():
    wb = load_workbook(XLSX); ws = find_fieldspec(wb)
    tc = next_free_col(ws)
    ws.cell(1, tc, "人工备注")
    ws.cell(3, tc, SENTINEL)
    wb.save(XLSX)
    print(f"injected 人工备注 at col {tc}, cols now {ws.max_column}")
    return tc


def check():
    wb = load_workbook(XLSX); ws = find_fieldspec(wb)
    mcols = manual_cols(ws)
    sql_col = col_by_header(ws, "验证SQL")
    # find the 人工 col + its sentinel
    tc = next(iter(mcols))
    sentinel_ok = ws.cell(3, tc).value == SENTINEL
    sql_ok = sql_col is not None and isinstance(ws.cell(3, sql_col).value, str) and "SELECT" in ws.cell(3, sql_col).value
    print(f"  manual_cols={mcols}  验证SQL col={sql_col}")
    print(f"  sentinel preserved: {sentinel_ok}")
    print(f"  验证SQL still populated (header-resolved): {sql_ok}")
    return sentinel_ok and sql_ok


def cleanup():
    wb = load_workbook(XLSX); ws = find_fieldspec(wb)
    for c in sorted(manual_cols(ws), reverse=True):
        ws.delete_cols(c, 1)
    wb.save(XLSX)
    print(f"cleaned up; cols now {ws.max_column}")


def main():
    inject()
    print("running add_field_spec_verify_sql.py ...")
    r = subprocess.run([sys.executable, "scripts/add_field_spec_verify_sql.py"],
                       capture_output=True, text=True, encoding="utf-8", errors="replace")
    print(r.stdout.strip());
    if r.returncode != 0:
        print("GENERATOR ERROR:\n", r.stderr[-800:])
    ok = check()
    cleanup()
    print("\nRESULT:", "PASS ✅" if ok else "FAIL ❌")


if __name__ == "__main__":
    main()
