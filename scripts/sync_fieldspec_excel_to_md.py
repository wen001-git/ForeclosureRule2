"""
Sync the Excel「📋 字段规范 Field Spec」(14 columns) into doc 14 zh as per-field cards.

Replaces the Section 2.0–4.1 horizontal field tables with one card per field (vertical
属性|值 table + fenced ```sql``` verify block + 实测结果 line), preserving all surrounding prose.
Generated blocks are wrapped in <!-- FS-CARDS:<id> --> markers so re-runs are idempotent.

Single source of truth = the DB-verified Excel. Run: python scripts/sync_fieldspec_excel_to_md.py
"""
import sys, io, re
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent))
from openpyxl import load_workbook
from _excel_guard import col_by_header

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")
MD = Path("docs/zh/14_bps_driven_servicer_fcl_interface.md")
RESULT_HDR_SUBSTR = "验证结果"

# section id -> the doc-14 zh heading substring that marks it
SECTION_HEADINGS = ["2.0", "2.1", "2.2", "2.3", "2.4", "2.5", "2.6", "3.1", "4.1"]


def esc(v):
    if v is None:
        return ""
    s = str(v).replace("\r", " ").replace("\n", " ").strip()
    return s.replace("|", "\\|")


def esc_ml(v):
    """多行版：保留换行（markdown 表格单元格内用 <br>），用于「标准接口取值范围」一值一行。"""
    if v is None:
        return ""
    s = str(v).replace("\r", "\n").strip()
    return s.replace("|", "\\|").replace("\n", "<br>")


def section_id_from_group(grp):
    """col2 字段组 like '2.0 四维状态基础字段 — 每个Se…' -> ('2.0', '四维状态基础字段')."""
    if not isinstance(grp, str):
        return None, ""
    m = re.match(r"\s*(\d+\.\d+)\s*(.*)", grp)
    if not m:
        return None, grp.strip()
    sid = m.group(1)
    label = m.group(2).split("—")[0].split("-")[0].strip() if "—" in m.group(2) or "-" in m.group(2) else m.group(2).strip()
    return sid, label


def read_fields():
    wb = load_workbook(XLSX)
    ws = None
    for s in wb.worksheets:
        if "Field Spec" in s.title:
            ws = s
            break
    C = {
        "grp": col_by_header(ws, "字段组"),
        "field": col_by_header(ws, "标准接口字段"),
        "src": col_by_header(ws, "Newrez原始字段"),
        "type": col_by_header(ws, "数据类型"),
        "prio": col_by_header(ws, "优先级"),
        "biz": col_by_header(ws, "业务含义"),
        "fmt": col_by_header(ws, "格式/计算规则"),
        "range": col_by_header(ws, "标准接口取值范围"),
        "typ": col_by_header(ws, "典型取值"),
        "panel": col_by_header(ws, "BPS面板/功能"),
        "status": col_by_header(ws, "Newrez状态"),
        "sql": col_by_header(ws, "验证SQL"),
        "result": col_by_header(ws, RESULT_HDR_SUBSTR),
    }
    result_hdr = ws.cell(1, C["result"]).value if C["result"] else "验证结果（实测）"
    sections = {}  # sid -> (label, [field dicts])
    for r in range(2, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        field = ws.cell(r, C["field"]).value if C["field"] else None
        if c1 is None or (isinstance(c1, str) and c1.startswith("Sect")) or not isinstance(field, str):
            continue
        sid, label = section_id_from_group(ws.cell(r, C["grp"]).value)
        if sid is None:
            continue
        d = {k: ws.cell(r, v).value for k, v in C.items() if v}
        sections.setdefault(sid, (label, []))[1].append(d)
    return sections, result_hdr


def make_cards(label, fields, result_hdr):
    out = []
    for d in fields:
        prio = esc(d.get("prio"))
        out.append(f"##### `{esc(d.get('field'))}` · {prio} · {label}")
        out.append("")
        out.append("| 属性 | 值 |")
        out.append("|---|---|")
        out.append(f"| Newrez 原始字段 | {esc(d.get('src'))} |")
        out.append(f"| 数据类型 | {esc(d.get('type'))} |")
        out.append(f"| 业务含义 | {esc(d.get('biz'))} |")
        out.append(f"| 格式/计算规则 | {esc(d.get('fmt'))} |")
        out.append(f"| 标准接口取值范围 | {esc_ml(d.get('range'))} |")
        out.append(f"| 典型取值（标准） | {esc(d.get('typ'))} |")
        out.append(f"| BPS 面板/功能 | {esc(d.get('panel'))} |")
        out.append(f"| Newrez 现状 | {esc(d.get('status'))} |")
        out.append(f"| {esc(result_hdr)} | {esc(d.get('result'))} |")
        out.append("")
        sql = d.get("sql")
        if isinstance(sql, str) and sql.strip():
            out.append("验证 SQL：")
            out.append("```sql")
            out.extend(sql.rstrip().splitlines())
            out.append("```")
            out.append("")
    return out


def main():
    sections, result_hdr = read_fields()
    print("Excel sections:", {k: len(v[1]) for k, v in sections.items()})

    lines = MD.read_text(encoding="utf-8").splitlines()
    out = []
    i = 0
    cur = None  # current target section id
    replaced = []

    heading_re = re.compile(r"^###\s+(\d+\.\d+)\b")

    while i < len(lines):
        ln = lines[i]

        # If we're at an existing FS-CARDS block, drop it (re-run idempotency)
        m_start = re.match(r"<!-- FS-CARDS:(\S+) START -->", ln.strip())
        if m_start:
            sid = m_start.group(1)
            # find END
            j = i + 1
            while j < len(lines) and "FS-CARDS:" not in lines[j]:
                j += 1
            # regenerate from Excel
            if sid in sections:
                label, fields = sections[sid]
                out.append(f"<!-- FS-CARDS:{sid} START -->")
                out.extend(make_cards(label, fields, result_hdr))
                out.append(f"<!-- FS-CARDS:{sid} END -->")
                replaced.append(sid + "(marker)")
            i = j + 1
            cur = None
            continue

        hm = heading_re.match(ln)
        if hm:
            cur = hm.group(1) if hm.group(1) in sections else None
            out.append(ln)
            i += 1
            continue

        # inside a target section: detect spec-table header -> replace both tables with cards
        if cur and ln.startswith("| 标准接口字段名 |"):
            j = i
            while j < len(lines) and lines[j].lstrip().startswith("|"):
                j += 1
            # skip blanks + business-meaning label + blanks
            while j < len(lines) and (lines[j].strip() == "" or "业务含义/计算逻辑" in lines[j]):
                j += 1
            # skip business-meaning table
            while j < len(lines) and lines[j].lstrip().startswith("|"):
                j += 1
            label, fields = sections[cur]
            out.append(f"<!-- FS-CARDS:{cur} START -->")
            out.extend(make_cards(label, fields, result_hdr))
            out.append(f"<!-- FS-CARDS:{cur} END -->")
            replaced.append(cur)
            i = j
            cur = None
            continue

        out.append(ln)
        i += 1

    MD.write_text("\n".join(out) + "\n", encoding="utf-8")
    print("Replaced sections:", replaced)
    print(f"Saved: {MD}")


if __name__ == "__main__":
    main()
