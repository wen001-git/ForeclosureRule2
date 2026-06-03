"""
(Re)generate the 验证SQL（取值范围）column (col13) of the "📋 字段规范 Field Spec" sheet.

Authoritative generator: bakes in (a) DB-verified CORRECTED source tables and (b) the
snapshot-table dataasof rule:
  - distribution/count queries (enum/binary/bk_code/idtext + special substitutes)
    -> WHERE dataasof = (SELECT MAX(dataasof) FROM <same table>)  [real latest-snapshot loans]
  - range queries (date / numeric MIN/MAX) -> full history (no dataasof filter; AVG dropped)
  - LM-code queries -> latest-snapshot-per-cycle CTE (unchanged)

newrez.* are daily snapshot tables (~891 snapshot days, ~7.6K loans), so un-filtered
GROUP BY COUNT(*) inflates counts ~200x. Verified via information_schema 2026-06-02.

Run: python scripts/add_field_spec_verify_sql.py
"""

import sys, io
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")

# ── field -> type ─────────────────────────────────────────────────────────────
TYPE = {
    "delinquency_status": "enum", "next_payment_due_date": "date", "days_past_due": "special",
    "foreclosure_flag": "special", "lm_flag": "binary", "lm_type": "lm_code",
    "lm_start_date": "date", "lm_end_date": "date", "hold_flag": "special",
    "hold_reason": "enum", "reo_flag": "special", "reo_acquisition_date": "special",
    "loan_id": "idtext", "servicer_loan_id": "idtext", "data_as_of_date": "date",
    "state": "enum", "judicial_flag": "binary", "active_fcl_flag": "binary",
    "fcl_referral_date": "date",
    "fcl_stage": "enum", "current_milestone": "enum", "last_completed_step": "enum",
    "last_completed_step_date": "date", "fcl_results": "enum", "attorney_firm": "idtext",
    "contested_flag": "binary", "servicer_days_in_fcl": "numeric", "days_in_fcl": "numeric",
    "noi_date": "special", "demand_sent_date": "date", "demand_expiration_date": "date",
    "fcl_setup_date": "date", "first_legal_date": "date", "service_complete_date": "date",
    "publication_date": "special", "title_received_date": "date", "title_clear_date": "date",
    "judgement_hearing_scheduled": "date", "judgement_entered_date": "date",
    "scheduled_sale_date": "date", "sale_held_date": "date", "deed_recorded_date": "date",
    "fcl_removal_date": "date", "third_party_proceeds_date": "date",
    "hold_1_description": "enum", "hold_1_start_date": "date", "hold_1_end_date": "date",
    "hold_1_projected_end_date": "date", "hold_2_description": "enum", "hold_2_start_date": "date",
    "hold_2_end_date": "date", "hold_2_projected_end_date": "date", "hold_3_description": "enum",
    "hold_3_start_date": "date", "hold_3_end_date": "date", "hold_3_projected_end_date": "date",
    "hold_1_modified_date": "date", "hold_2_modified_date": "date", "hold_3_modified_date": "date",
    "bid_amount": "numeric", "approved_bid_price": "numeric", "sale_amount": "numeric",
    "investor_loan_id": "idtext", "lien_position": "enum", "interest_paid_through_date": "date",
    "in_auction_flag": "binary", "borrower_deceased_flag": "binary", "reason_for_default": "enum",
    "hold_1_comment": "idtext", "hold_2_comment": "idtext", "hold_3_comment": "idtext",
    "lm_deal": "lm_code", "lm_program": "lm_code", "lm_status": "lm_code",
    "lm_cycle_open_date": "date", "lm_cycle_close_date": "date",
    "lm_final_disposition": "lm_code", "lm_denial_reason": "lm_code",
    "borrower_intentions": "lm_code", "imminent_default": "special",
    "single_point_of_contact": "special",
    "active_bk_flag": "binary", "bk_status": "bk_code", "bk_legal_status": "bk_code",
    "bk_status_date": "date", "bk_chapter": "bk_code", "bk_filed_date": "date",
    "bk_removal_date": "date", "mfr_filed_date": "date", "mfr_hearing_results": "bk_code",
    "proof_of_claim_date": "date", "post_petition_due_date": "date",
}

# ── field -> (schema.table, column)  [DB-verified CORRECTED sources] ──────────
SOURCE = {
    "delinquency_status": ("newrez.portnewrezgeneral", "delinquency_status_mba"),
    "next_payment_due_date": ("newrez.portnewrezpmt", "nextduedate"),
    "lm_start_date": ("newrez.portnewrezlm", "dealstartdate"),
    "lm_end_date": ("newrez.portnewrezlm", "lmremovaldate"),
    "hold_reason": ("newrez.portnewrezfc", "fchold1description"),
    "loan_id": ("newrez.portnewrezfc", "loanid"),
    "servicer_loan_id": ("newrez.portnewrezfc", "shellpointloanid"),
    "data_as_of_date": ("newrez.portnewrezfc", "dataasof"),
    "state": ("newrez.portnewrezprop", "propertystate"),
    "judicial_flag": ("newrez.portnewrezfc", "judicial"),
    "active_fcl_flag": ("newrez.portnewrezfc", "activefcflag"),
    "fcl_referral_date": ("newrez.portnewrezfc", "fcreferraldate"),
    "fcl_stage": ("newrez.portnewrezfc", "fcstage"),
    "current_milestone": ("newrez.portnewrezfc", "currentmilestone"),
    "last_completed_step": ("newrez.portnewrezfc", "lastfcstepcompleted"),
    "last_completed_step_date": ("newrez.portnewrezfc", "lastfcstepcompleteddate"),
    "fcl_results": ("newrez.portnewrezfc", "fcresults"),
    "attorney_firm": ("newrez.portnewrezfc", "fcfirm"),
    "contested_flag": ("newrez.portnewrezfc", "fccontestedflag"),
    "servicer_days_in_fcl": ("newrez.portnewrezfc", "smsdaysinfc"),
    "days_in_fcl": ("newrez.portnewrezfc", "daysinfc"),
    "lm_flag": ("newrez.portnewrezlm", "activelmflag"),
    "demand_sent_date": ("newrez.portnewrezfc", "demandsentdate"),
    "demand_expiration_date": ("newrez.portnewrezfc", "demandexpirationdate"),
    "fcl_setup_date": ("newrez.portnewrezfc", "fcsetupdate"),
    "first_legal_date": ("newrez.portnewrezfc", "firstlegaldate"),
    "service_complete_date": ("newrez.portnewrezfc", "servicecompletedate"),
    "title_received_date": ("newrez.portnewrezfc", "titlereceiveddate"),
    "title_clear_date": ("newrez.portnewrezfc", "titlecleardate"),
    "judgement_hearing_scheduled": ("newrez.portnewrezfc", "fcjudgmenthearingscheduled"),
    "judgement_entered_date": ("newrez.portnewrezfc", "fcjudgmententered"),
    "scheduled_sale_date": ("newrez.portnewrezfc", "fcscheduledsaledate"),
    "sale_held_date": ("newrez.portnewrezfc", "fcsalehelddate"),
    "deed_recorded_date": ("newrez.portnewrezfc", "dtdeedrecorded"),
    "fcl_removal_date": ("newrez.portnewrezfc", "fcremovaldate"),
    "third_party_proceeds_date": ("newrez.portnewrezfc", "fcl3rdpartyproceedsreceiveddate"),
    "hold_1_description": ("newrez.portnewrezfc", "fchold1description"),
    "hold_1_start_date": ("newrez.portnewrezfc", "fchold1startdate"),
    "hold_1_end_date": ("newrez.portnewrezfc", "fchold1enddate"),
    "hold_1_projected_end_date": ("newrez.portnewrezfc", "fchold1projectedenddate"),
    "hold_2_description": ("newrez.portnewrezfc", "fchold2description"),
    "hold_2_start_date": ("newrez.portnewrezfc", "fchold2startdate"),
    "hold_2_end_date": ("newrez.portnewrezfc", "fchold2enddate"),
    "hold_2_projected_end_date": ("newrez.portnewrezfc", "fchold2projectedenddate"),
    "hold_3_description": ("newrez.portnewrezfc", "fchold3description"),
    "hold_3_start_date": ("newrez.portnewrezfc", "fchold3startdate"),
    "hold_3_end_date": ("newrez.portnewrezfc", "fchold3enddate"),
    "hold_3_projected_end_date": ("newrez.portnewrezfc", "fchold3projectedenddate"),
    "hold_1_modified_date": ("newrez.portnewrezfc", "holdmodified"),
    "hold_2_modified_date": ("newrez.portnewrezfc", "holdmodified2"),
    "hold_3_modified_date": ("newrez.portnewrezfc", "holdmodified3"),
    "bid_amount": ("newrez.portnewrezfc", "fcbidamount"),
    "approved_bid_price": ("newrez.portnewrezfc", "fcapprbidprice"),
    "sale_amount": ("newrez.portnewrezfc", "fcsaleamount"),
    "investor_loan_id": ("newrez.portnewrezfc", "investorloanid"),
    "lien_position": ("newrez.portnewrezgeneral", "lienposition"),
    "interest_paid_through_date": ("newrez.portnewrezgeneral", "interestpaidthroughdate"),
    "in_auction_flag": ("newrez.portnewrezgeneral", "inauctionflag"),
    "borrower_deceased_flag": ("newrez.portnewrezgeneral", "borrowerdeceasedflag"),
    "reason_for_default": ("newrez.portnewrezgeneral", "reasonfordefault"),
    "hold_1_comment": ("newrez.portnewrezfc", "fchold1comment"),
    "hold_2_comment": ("newrez.portnewrezfc", "fchold2comment"),
    "hold_3_comment": ("newrez.portnewrezfc", "fchold3comment"),
    "lm_cycle_open_date": ("newrez.portnewrezlm", "dealstartdate"),
    "lm_cycle_close_date": ("newrez.portnewrezlm", "lmremovaldate"),
    "active_bk_flag": ("newrez.portnewrezbk", "activebkflag"),
    "bk_status": ("newrez.portnewrezbk", "bkstatus"),
    "bk_legal_status": ("newrez.portnewrezbk", "bkstage"),
    "bk_status_date": ("newrez.portnewrezbk", "bkrcurrentstatusdate"),
    "bk_chapter": ("newrez.portnewrezbk", "bkchapter"),
    "bk_filed_date": ("newrez.portnewrezbk", "bkfileddate"),
    "bk_removal_date": ("newrez.portnewrezbk", "bkremovaldate"),
    "mfr_filed_date": ("newrez.portnewrezbk", "mfrfileddate"),
    "mfr_hearing_results": ("newrez.portnewrezbk", "mfrhearingresults"),
    "proof_of_claim_date": ("newrez.portnewrezbk", "pocfileddate"),
    "post_petition_due_date": ("newrez.portnewrezbk", "bkpostpetitionduedate"),
}

LM_BPS_COL = {"lm_type": "deal", "lm_deal": "deal", "lm_program": "program", "lm_status": "lmc_status",
              "lm_final_disposition": "final_disposition", "lm_denial_reason": "denialreason",
              "borrower_intentions": "borrower_intentions"}
LM_RAW_COL = {"lm_type": "lmdeal", "lm_deal": "lmdeal", "lm_program": "lmprogram", "lm_status": "lmstatus",
              "lm_final_disposition": "lmdecision", "lm_denial_reason": "denialreason",
              "borrower_intentions": "borrowerintention"}

SNAP_NOTE = "-- 快照表：已按最新 dataasof 过滤（计数=最新快照真实贷款分布，非全历史快照行数）"


def maxsnap(t):
    return f"(SELECT MAX(dataasof) FROM {t})"


def gen_special(field):
    """Special / derived / not-provided fields, with dataasof filter on distribution ones."""
    if field == "days_past_due":
        return ("-- N/A ETL衍生（days360(nextduedate,dataasof)）；替代验证源 nextduedate（最新快照取10笔样例）\n"
                "-- 源表: newrez.portnewrezpmt | 运行于: mysql_dev\n"
                "SELECT loanid, dataasof, nextduedate\n"
                "FROM   newrez.portnewrezpmt\n"
                "WHERE  dataasof = (SELECT MAX(dataasof) FROM newrez.portnewrezpmt)\n"
                "  AND  nextduedate IS NOT NULL\n"
                "LIMIT  10;")
    if field == "foreclosure_flag":
        return ("-- N/A Newrez无fcl_flag列；FCL活跃状态由 activefcflag 表达\n"
                f"{SNAP_NOTE}\n"
                "-- 源表: newrez.portnewrezfc | 运行于: mysql_dev\n"
                "SELECT dataasof AS data_date, activefcflag, COUNT(*) AS cnt\n"
                "FROM   newrez.portnewrezfc\n"
                "WHERE  dataasof = (SELECT MAX(dataasof) FROM newrez.portnewrezfc)\n"
                "GROUP  BY dataasof, activefcflag ORDER BY cnt DESC;")
    if field == "lm_type":
        return ("-- N/A Newrez用 lmdeal 整数码代替标准 lm_type；见 lm_deal 行解码 SQL\n"
                f"{SNAP_NOTE}\n"
                "-- 源表: newrez.portnewrezlm | 运行于: mysql_dev\n"
                "SELECT dataasof AS data_date, lmdeal, COUNT(*) AS cnt\n"
                "FROM   newrez.portnewrezlm\n"
                "WHERE  dataasof = (SELECT MAX(dataasof) FROM newrez.portnewrezlm)\n"
                "  AND  lmdeal IS NOT NULL AND lmdeal != ''\n"
                "GROUP  BY dataasof, lmdeal ORDER BY cnt DESC;")
    if field == "reo_flag":
        return ("-- N/A Newrez未提供独立 REO flag；可用 fcresults 间接识别\n"
                f"{SNAP_NOTE}\n"
                "-- 源表: newrez.portnewrezfc | 运行于: mysql_dev\n"
                "SELECT dataasof AS data_date, fcresults, COUNT(*) AS cnt\n"
                "FROM   newrez.portnewrezfc\n"
                "WHERE  dataasof = (SELECT MAX(dataasof) FROM newrez.portnewrezfc)\n"
                "  AND  fcresults IN ('REO','REO Sale')\n"
                "GROUP  BY dataasof, fcresults;")
    if field == "reo_acquisition_date":
        return ("-- N/A Newrez未提供独立 REO 取得日；可参考 dtdeedrecorded=止赎契据登记日（成交后产权转让契据在县登记处登记=产权过户/止赎完成点，约在拍卖 fcsalehelddate 后 2-3 周，多数→REO），近似 REO 取得日（最新快照取10笔样例）\n"
                "-- 源表: newrez.portnewrezfc | 运行于: mysql_dev\n"
                "SELECT loanid, dataasof, dtdeedrecorded\n"
                "FROM   newrez.portnewrezfc\n"
                "WHERE  dataasof = (SELECT MAX(dataasof) FROM newrez.portnewrezfc)\n"
                "  AND  dtdeedrecorded IS NOT NULL\n"
                "LIMIT  10;")
    if field == "hold_flag":
        return ("-- N/A 无独立顶层 Hold flag；可从 fchold1startdate 推导（最新快照计数）\n"
                f"{SNAP_NOTE}\n"
                "-- 源表: newrez.portnewrezfc | 运行于: mysql_dev\n"
                "SELECT MAX(dataasof) AS data_date, COUNT(*) AS hold_loans\n"
                "FROM   newrez.portnewrezfc\n"
                "WHERE  dataasof = (SELECT MAX(dataasof) FROM newrez.portnewrezfc)\n"
                "  AND  fchold1startdate IS NOT NULL;")
    if field == "noi_date":
        return ("-- N/A Newrez未单独提供 NOI；BPS 以 demandsentdate 代替（最新快照取10笔样例）\n"
                "-- 源表: newrez.portnewrezfc | 运行于: mysql_dev\n"
                "SELECT loanid, dataasof, demandsentdate\n"
                "FROM   newrez.portnewrezfc\n"
                "WHERE  dataasof = (SELECT MAX(dataasof) FROM newrez.portnewrezfc)\n"
                "  AND  demandsentdate IS NOT NULL\n"
                "LIMIT  10;")
    if field in ("publication_date", "imminent_default", "single_point_of_contact"):
        return "-- N/A — Newrez未提供（无可验证源）"
    return "-- N/A"


def gen_sql(field, ftype):
    if ftype == "special":
        return gen_special(field)

    if ftype == "lm_code":
        raw = LM_RAW_COL[field]
        bps = LM_BPS_COL[field]
        extra = " AND l.{0} != 0".format(raw) if field in ("lm_denial_reason", "borrower_intentions") else ""
        return (
            f"-- 验证 {field} 取值范围（整数码 → BPS 解码文本）\n"
            f"-- 源表: newrez.portnewrezlm JOIN bpms_dev.sync_loan_foreclosure_loss_mitigation\n"
            f"-- 运行于: mysql_bpms_dev（跨库 JOIN）| 详见 doc 数据字典 SQL-D1~D6\n"
            f"-- 快照表：CTE 取每个 LM 周期的最新快照（latest-snapshot-per-cycle）\n"
            f"WITH latest_lm AS (\n"
            f"  SELECT loanid, dealstartdate, {raw},\n"
            f"         ROW_NUMBER() OVER (PARTITION BY loanid, dealstartdate ORDER BY dataasof DESC) AS rn\n"
            f"  FROM newrez.portnewrezlm WHERE {raw} IS NOT NULL\n"
            f")\n"
            f"SELECT l.{raw}, b.{bps}, COUNT(*) AS cnt\n"
            f"FROM   latest_lm l\n"
            f"JOIN   bpms_dev.sync_loan_foreclosure_loss_mitigation b\n"
            f"  ON   l.loanid = b.loanid AND l.dealstartdate = b.cycle_opened_date\n"
            f"WHERE  l.rn = 1{extra}\n"
            f"  AND  b.{bps} IS NOT NULL AND b.{bps} != '' AND b.{bps} NOT REGEXP '^[0-9]'\n"
            f"GROUP  BY l.{raw}, b.{bps} ORDER BY l.{raw}, cnt DESC;"
        )

    t, col = SOURCE[field]
    base_hdr = f"-- 验证 {field} 取值范围\n-- 源表: {t} | 运行于: mysql_dev\n"
    sample_hdr = f"-- 验证 {field} 取值样例（最新快照取10笔，快速查看字段的样子）\n-- 源表: {t} | 运行于: mysql_dev\n"

    if ftype in ("date", "numeric"):
        # 日期/连续数值字段：不算范围，取最新快照10笔样例查看字段实际取值（含 dataasof 显示数据日期）
        return sample_hdr + (
            f"SELECT loanid, dataasof, {col}\n"
            f"FROM   {t}\n"
            f"WHERE  dataasof = {maxsnap(t)}\n"
            f"  AND  {col} IS NOT NULL\n"
            f"LIMIT  10;"
        )
    if ftype == "enum":
        return base_hdr + SNAP_NOTE + "\n" + (
            f"SELECT dataasof AS data_date, {col} AS val, COUNT(*) AS cnt\n"
            f"FROM   {t}\n"
            f"WHERE  dataasof = {maxsnap(t)}\n"
            f"  AND  {col} IS NOT NULL AND {col} != ''\n"
            f"GROUP  BY dataasof, {col} ORDER BY cnt DESC LIMIT 30;"
        )
    if ftype == "binary":
        return base_hdr + SNAP_NOTE + "\n" + (
            f"SELECT dataasof AS data_date, {col}, COUNT(*) AS cnt\n"
            f"FROM   {t}\n"
            f"WHERE  dataasof = {maxsnap(t)}\n"
            f"GROUP  BY dataasof, {col} ORDER BY cnt DESC;"
        )
    if ftype == "idtext":
        select_cols = "loanid, dataasof" if col == "loanid" else f"loanid, dataasof, {col}"
        return sample_hdr + (
            f"SELECT {select_cols}\n"
            f"FROM   {t}\n"
            f"WHERE  dataasof = {maxsnap(t)}\n"
            f"  AND  {col} IS NOT NULL AND {col} != ''\n"
            f"LIMIT  10;"
        )
    if ftype == "bk_code":
        return base_hdr + (
            f"-- BK 整数编码（Newrez 原始码；BPS 端解码情况见 col12/数据字典）\n"
            f"{SNAP_NOTE}\n"
            f"SELECT dataasof AS data_date, {col} AS code, COUNT(*) AS cnt\n"
            f"FROM   {t}\n"
            f"WHERE  dataasof = {maxsnap(t)}\n"
            f"  AND  {col} IS NOT NULL AND {col} != ''\n"
            f"GROUP  BY dataasof, {col} ORDER BY cnt DESC;"
        )
    return "-- N/A"


def thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def main():
    from copy import copy
    sys.path.insert(0, str(Path(__file__).parent))
    from _excel_guard import col_by_header, next_free_col, assert_safe  # 人工列保护

    wb = load_workbook(XLSX)
    ws = None
    for s in wb.worksheets:
        if "Field Spec" in s.title:
            ws = s
            break

    # 按表头定位列（不硬编码列号），避免用户插入「人工」列导致列号漂移、误写
    field_col = col_by_header(ws, "标准接口字段") or 3
    COL = col_by_header(ws, "验证SQL") or next_free_col(ws)
    assert_safe(ws, COL)  # 绝不写入「人工」列

    # ensure header
    h = ws.cell(1, COL, "验证SQL（取值范围/样例）")
    h.fill = PatternFill("solid", fgColor="2E75B6")
    h.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    h.border = thin()
    ws.column_dimensions[get_column_letter(COL)].width = 82

    mono = Font(name="Consolas", size=8)
    cell_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    filled = 0
    for r in range(2, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        field = ws.cell(r, field_col).value
        if c1 is None or (isinstance(c1, str) and c1.startswith("Sect")):
            continue
        if not isinstance(field, str):
            continue
        field = field.strip()
        if field not in TYPE:
            continue
        sql = gen_sql(field, TYPE[field])
        cell = ws.cell(r, COL, sql)
        cell.font = mono
        cell.alignment = cell_align
        cell.border = thin()
        # cosmetic: copy row fill from the immediate left cell (skip if that's a 人工 col)
        src = COL - 1
        if src >= 1 and (ws.cell(1, src).value is None or "人工" not in str(ws.cell(1, src).value)):
            adj = ws.cell(r, src).fill
            if adj and adj.fgColor and adj.fgColor.rgb not in ("00000000", "FFFFFFFF", None):
                cell.fill = copy(adj)
        filled += 1

    wb.save(XLSX)
    print(f"Regenerated {filled} verify-SQL cells at col {COL} (header-resolved, 人工列受保护)")


if __name__ == "__main__":
    main()
