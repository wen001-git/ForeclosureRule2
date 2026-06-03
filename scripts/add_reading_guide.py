"""
Add a 阅读指南 (Reading Guide) sheet as the first tab of 14_servicer_fcl_field_spec.xlsx.
Run once from the project root: python scripts/add_reading_guide.py
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

XLSX_PATH = Path(__file__).parent.parent / "docs" / "14_servicer_fcl_field_spec.xlsx"
SHEET_NAME = "📖 阅读指南 Guide"

# ── colour palette ──────────────────────────────────────────────────────────
C_DARK_BLUE  = "1F3864"   # section header bg
C_MID_BLUE   = "2E75B6"   # table header bg
C_LIGHT_BLUE = "D6E4F0"   # alt row / highlight bg
C_YELLOW     = "FFF2CC"   # P0 row
C_ORANGE     = "FCE4D6"   # P1 row
C_GREEN      = "E2EFDA"   # P2 row / total row
C_WHITE      = "FFFFFF"
C_TEXT_WHITE = "FFFFFF"
C_TEXT_DARK  = "1F1F1F"

def font(bold=False, size=11, color=C_TEXT_DARK, italic=False):
    return Font(name="微软雅黑", bold=bold, size=size, color=color, italic=italic)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def align(wrap=True, h="left", v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def write_section_header(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font(bold=True, size=12, color=C_TEXT_WHITE)
    cell.fill = fill(C_DARK_BLUE)
    cell.alignment = align(h="left")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    return row + 1

def write_body(ws, row, text, indent=False, bg=None):
    val = ("    " if indent else "") + text
    cell = ws.cell(row=row, column=1, value=val)
    cell.font = font(size=11)
    cell.alignment = align()
    if bg:
        cell.fill = fill(bg)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    return row + 1

def write_table_header(ws, row, cols):
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = font(bold=True, color=C_TEXT_WHITE)
        cell.fill = fill(C_MID_BLUE)
        cell.alignment = align(h="center")
        cell.border = thin_border()
    return row + 1

def write_table_row(ws, row, values, bg=None):
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = font(size=10)
        cell.alignment = align(h="left" if c == 1 else "center")
        cell.border = thin_border()
        if bg:
            cell.fill = fill(bg)
    return row + 1

def blank(ws, row, n=1):
    return row + n


def build_sheet(ws):
    ws.sheet_view.showGridLines = False
    # column widths
    widths = [38, 18, 18, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    row = 1

    # ── Title ────────────────────────────────────────────────────────────────
    title = ws.cell(row=row, column=1,
                    value="📖  阅读指南 — 14_servicer_fcl_field_spec.xlsx")
    title.font = Font(name="微软雅黑", bold=True, size=16, color=C_TEXT_WHITE)
    title.fill = fill(C_DARK_BLUE)
    title.alignment = align(h="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row = blank(ws, row + 1)

    # ── Section 1: 文档目的 ───────────────────────────────────────────────────
    row = write_section_header(ws, row, "1  文档目的")
    row = write_body(ws, row,
        "本文档从 BPS Asset Management Foreclosure 五大功能面板的实际显示需求出发，"
        "逆向还原每个面板所需的 Servicer 源字段，形成 BPS 落地层的具体数据接口规范"
        "（约 92 个字段）。以 Newrez 为合规基准，供新 Servicer 对接工程师、"
        "数据治理团队与 BPS 运营人员使用。")
    row = blank(ws, row)

    # ── Section 2: 适用范围 ───────────────────────────────────────────────────
    row = write_section_header(ws, row, "2  适用范围")
    scope_items = [
        "✅  覆盖 BPS Foreclosure 五大面板所需的全部 Servicer 字段（约 92 个）",
        "✅  含四维状态基础标志层（12 个）+ BPS 展示明细层（68 个）+ Newrez 高价值增强字段层（12 个）",
        "✅  以 Newrez 为合规基准，其他 Servicer 以此为参照执行 gap 分析",
        "❌  不涉及 ETL 中间层代码细节（见 doc 12）",
    ]
    for item in scope_items:
        row = write_body(ws, row, item, indent=True)
    row = blank(ws, row)

    # ── Section 3: P0/P1/P2 定义 ──────────────────────────────────────────────
    row = write_section_header(ws, row, "3  P0 / P1 / P2 优先级定义")
    row = write_table_header(ws, row, ["优先级", "含义", "缺失后果", "", ""])
    # merge last 3 cols for 缺失后果
    p_rows = [
        ("P0", "系统入库前提条件",
         "缺失则拒绝入库，BPS 无法处理该贷款", C_YELLOW),
        ("P1", "核心面板显示字段",
         "缺失则对应 BPS 面板数据异常或显示空白（降级可用）", C_ORANGE),
        ("P2", "增强型分析字段",
         "可选；存在时启用对应分析功能；缺失不影响主流程", C_GREEN),
    ]
    for prio, meaning, consequence, bg in p_rows:
        r = row
        ws.cell(r, 1, prio).font = font(bold=True, size=11)
        ws.cell(r, 1).fill = fill(bg)
        ws.cell(r, 1).alignment = align(h="center")
        ws.cell(r, 1).border = thin_border()
        ws.cell(r, 2, meaning).font = font(size=10)
        ws.cell(r, 2).fill = fill(bg)
        ws.cell(r, 2).alignment = align()
        ws.cell(r, 2).border = thin_border()
        # consequence spans cols 3-5
        ws.cell(r, 3, consequence).font = font(size=10)
        ws.cell(r, 3).fill = fill(bg)
        ws.cell(r, 3).alignment = align()
        ws.cell(r, 3).border = thin_border()
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        row += 1
    row = blank(ws, row)

    # ── Section 4: Newrez 现状 ────────────────────────────────────────────────
    row = write_section_header(ws, row, "4  Newrez 现状状态说明（合规率总览）")
    row = write_table_header(ws, row,
          ["字段组", "总字段数", "✅ 已提供", "⚠️ 部分", "❌ 未提供"])

    status_data = [
        ("四维状态基础标志 (2.0)", 12, 1, 5, 6, "~8%",  C_ORANGE),
        ("识别/入库字段 (2.1)",    7,  7, 0, 0, "100%", C_GREEN),
        ("FCL 状态字段 (2.2)",     9,  7, 2, 0, "88%",  C_GREEN),
        ("FCL 时间线字段 (2.3)",  16,  9, 2, 4, "~56%", C_LIGHT_BLUE),
        ("Hold 槽位字段 (2.4)",   12, 12, 0, 0, "100%", C_GREEN),
        ("Bid/Sale 字段 (2.5)",    3,  2, 1, 0, "83%",  C_GREEN),
        ("贷款属性增强字段 (2.6)", 12,  9, 0, 0, "~75%", C_LIGHT_BLUE),
        ("LM 周期字段 (3.1)",     10,  7, 0, 3, "70%",  C_LIGHT_BLUE),
        ("BK 字段 (4.1)",         11,  8, 3, 0, "91%",  C_GREEN),
    ]
    for name, total, ok, partial, missing, rate, bg in status_data:
        row = write_table_row(ws, row, [name, total, ok, partial, missing], bg)

    # total row
    r = row
    totals = ["合计", 92, 62, 13, 13]
    for c, val in enumerate(totals, 1):
        cell = ws.cell(r, c, val)
        cell.font = font(bold=True, size=11)
        cell.fill = fill(C_DARK_BLUE)
        cell.font = Font(name="微软雅黑", bold=True, size=11, color=C_TEXT_WHITE)
        cell.alignment = align(h="center" if c > 1 else "left")
        cell.border = thin_border()
    # append ~67% in col 5 (reuse last cell position col 5 = missing; add note)
    note = ws.cell(r, 5, "合规率 ~67%")
    note.font = Font(name="微软雅黑", bold=True, size=11, color=C_TEXT_WHITE)
    note.fill = fill(C_DARK_BLUE)
    note.alignment = align(h="center")
    note.border = thin_border()
    row = blank(ws, row + 1)

    # ── Section 5: 如何评审 ────────────────────────────────────────────────────
    row = write_section_header(ws, row, "5  如何评审本文档")
    review_steps = [
        ("步骤 1", "优先审 P0 字段",
         "进入「字段规范」sheet，按优先级列筛选 P0。确认 Newrez 合规状态列全为 ✅；"
         "任何 ❌ 或 ⚠️ 均需立即列入行动清单，视为上线阻断项。"),
        ("步骤 2", "核查 P1 缺失影响",
         "筛选优先级 = P1 且合规状态为 ❌ 或 ⚠️ 的行；"
         "参阅「BPS 面板索引」确认缺失将影响哪个面板功能；"
         "将影响描述填入「行动清单」的【缺失影响】列。"),
        ("步骤 3", "跟进待确认项",
         "查看「行动清单」中状态为【待确认】的行，重点关注三类：\n"
         "① BK 面板 bkstatus/bkstage 是否在 BPS 侧解码为文本\n"
         "② noi_start_date 与 demand_start_date 的 BPS 展示差异\n"
         "③ P2 字段是否应向新 Servicer 强制请求"),
    ]
    for step, title_text, desc in review_steps:
        r = row
        label = ws.cell(r, 1, f"{step}：{title_text}")
        label.font = font(bold=True, size=11, color=C_TEXT_WHITE)
        label.fill = fill(C_MID_BLUE)
        label.alignment = align(h="left")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        row += 1
        desc_cell = ws.cell(row, 1, desc)
        desc_cell.font = font(size=10)
        desc_cell.alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True)
        desc_cell.fill = fill(C_LIGHT_BLUE)
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=5)
        ws.row_dimensions[row].height = 60
        row = blank(ws, row + 1)

    # ── Footer ────────────────────────────────────────────────────────────────
    footer = ws.cell(row, 1,
        "本工作簿共 6 张 sheet：阅读指南 · 总览 Overview · 字段规范 Field Spec · "
        "行动清单 Action List · BPS面板索引 · BPS显示映射")
    footer.font = Font(name="微软雅黑", size=9, italic=True, color="808080")
    footer.alignment = align(h="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)


def main():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Not found: {XLSX_PATH}")

    wb = load_workbook(XLSX_PATH)

    if SHEET_NAME in wb.sheetnames:
        print("Sheet already exists - removing and recreating.")
        del wb[SHEET_NAME]

    ws = wb.create_sheet(SHEET_NAME, 0)
    build_sheet(ws)

    wb.save(XLSX_PATH)
    print(f"Saved: {XLSX_PATH}")
    names = [s.encode("ascii", "replace").decode() for s in wb.sheetnames]
    print(f"Sheets: {names}")


if __name__ == "__main__":
    main()
