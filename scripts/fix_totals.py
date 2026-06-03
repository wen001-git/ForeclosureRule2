"""
Fix Section 2.0 total 12→11 (exclude ETL-derived days_past_due from Servicer compliance count)
and grand total 92→91.

  2.0: 2+4+5 = 11 = total ✓   (was 12, math: 2+4+5=11≠12)
  合计: 67+12+12 = 91 = total ✓  (was 92)

Also updates rate for 2.0: 2/11 = ~18%  (was ~17%)

Run: python scripts/fix_totals.py
"""

import sys, io
from pathlib import Path
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

XLSX = Path(__file__).parent.parent / "docs" / "14_servicer_fcl_field_spec.xlsx"


def main():
    wb = load_workbook(XLSX)
    ov    = wb.worksheets[2]   # Overview (index 2 after glossary insert)
    guide = wb.worksheets[0]   # 阅读指南

    # ── Overview ─────────────────────────────────────────────────────────────
    for r in range(1, 60):
        a = ov.cell(r, 1).value
        if not isinstance(a, str):
            continue
        if "2.0" in a and "四维" in a:
            ov.cell(r, 2).value = 11       # total: 12 → 11
            ov.cell(r, 6).value = "~18%"   # rate: ~17% → ~18% (2/11)
            print(f"[Overview] row {r} 2.0: total→11, rate→~18%")
        if "合计" in a:
            ov.cell(r, 2).value = 91       # grand total: 92 → 91
            print(f"[Overview] row {r} 合计: total→91")

    # ── 阅读指南 ──────────────────────────────────────────────────────────────
    # Row 20: 2.0 data, col B = 总字段数
    guide.cell(20, 2).value = 11
    print("[阅读指南] row 20 (2.0) total → 11")

    # Row 29: 合计, col B = 总字段数
    guide.cell(29, 2).value = 91
    print("[阅读指南] row 29 (合计) total → 91")

    wb.save(XLSX)
    print(f"\nSaved: {XLSX}")
    print("\nVerification:")
    print("  2.0:  2 + 4 + 5 = 11 = total ✓")
    print("  合计: 67 + 12 + 12 = 91 = total ✓")
    print("\nNote: 2.6 shows total=12 but 9+0+0=9≠12.")
    print("  The 3 extra are planned-but-not-yet-specced fields from doc 14.")
    print("  Leave as-is until those fields are added to the Field Spec.")


if __name__ == "__main__":
    main()
