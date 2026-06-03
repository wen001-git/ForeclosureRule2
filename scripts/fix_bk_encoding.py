"""
Fix the 3 BK fields in Field Spec (14_servicer_fcl_field_spec.xlsx) after DB verification:
the earlier "BK 整数码 ETL 未解码 / BPS 直接存数字" assumption was disproven.

DB-verified (2026-06-02):
  bk_status       -> BPS decodes to text (1→Active·2→Discharged·3→Dismissed·4→Closed·5→ReliefGranted)
  bk_legal_status -> BPS legal_status comes from portnewrezgeneral.legalstatus (text), NOT bkstage
  mfr_hearing_results -> BPS mfr_status empty in dev (0/64)

Updates col9 (标准接口取值范围), col12 (Newrez状态), col13 (验证SQL). Excel must be closed.
Run: python scripts/fix_bk_encoding.py
"""

import sys, io
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")

FIX = {
    "bk_status": {
        9: "BPS 解码文本：Active·Discharged·Dismissed·Closed·ReliefGranted（Newrez 原始 bkstatus 1~5）",
        12: "✅ 已确认：BPS 解码为文本（1→Active·2→Discharged·3→Dismissed·4→Closed·5→ReliefGranted，DB 实测）",
        13: ("-- 验证 bk_status 取值范围（BPS 解码后文本）\n"
             "-- 源表: bpms_dev.sync_loan_foreclosure_bankruptcy | 运行于: mysql_bpms_dev\n"
             "-- DB 实测确认 BPS 将 Newrez bkstatus(int) 解码为文本（非原始数字码）\n"
             "-- 注：本 BPS 表为当前态/事件表，无 dataasof 快照列（不显示单一数据日期）\n"
             "SELECT bankruptcy_status AS val, COUNT(*) AS cnt\n"
             "FROM   bpms_dev.sync_loan_foreclosure_bankruptcy\n"
             "WHERE  bankruptcy_status IS NOT NULL AND bankruptcy_status != ''\n"
             "GROUP  BY bankruptcy_status ORDER BY cnt DESC;"),
    },
    "bk_legal_status": {
        9: "BPS 文本：FCBU·BK13·BK7·BK11·BK7DCH·BK11DCH·BK13DCH·BKD7LM·BKD13LM·FCSold·REO（取自 portnewrezgeneral.legalstatus）",
        12: "✅ 已确认：BPS legal_status 取自 portnewrezgeneral.legalstatus 文本（非 bkstage 数字码解码，DB 实测）",
        13: ("-- 验证 bk_legal_status 取值范围（BPS legal_status 实际源 = portnewrezgeneral.legalstatus 文本）\n"
             "-- 源表: newrez.portnewrezgeneral | 运行于: mysql_dev\n"
             "-- 快照表：已按最新 dataasof 过滤\n"
             "SELECT dataasof AS data_date, legalstatus AS val, COUNT(*) AS cnt\n"
             "FROM   newrez.portnewrezgeneral\n"
             "WHERE  dataasof = (SELECT MAX(dataasof) FROM newrez.portnewrezgeneral)\n"
             "  AND  legalstatus IS NOT NULL AND legalstatus != ''\n"
             "GROUP  BY dataasof, legalstatus ORDER BY cnt DESC LIMIT 30;"),
    },
    "mfr_hearing_results": {
        12: "⚠️ Newrez 原始数字码（0/3/4/5/6 实测）；BPS mfr_status 列 dev 实测全空（0/64），下游填充/解码待确认",
        13: ("-- 验证 mfr_hearing_results：Newrez 原始码分布（BPS mfr_status dev 实测全空 0/64）\n"
             "-- 源表: newrez.portnewrezbk | 运行于: mysql_dev | 快照表：已按最新 dataasof 过滤\n"
             "SELECT dataasof AS data_date, mfrhearingresults AS code, COUNT(*) AS cnt\n"
             "FROM   newrez.portnewrezbk\n"
             "WHERE  dataasof = (SELECT MAX(dataasof) FROM newrez.portnewrezbk)\n"
             "  AND  mfrhearingresults IS NOT NULL AND mfrhearingresults != ''\n"
             "GROUP  BY dataasof, mfrhearingresults ORDER BY cnt DESC;\n"
             "-- 对照 BPS: SELECT COUNT(mfr_status) FROM bpms_dev.sync_loan_foreclosure_bankruptcy; (dev=0)"),
    },
}

MONO = Font(name="Consolas", size=8)
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def main():
    wb = load_workbook(XLSX)
    ws = None
    for s in wb.worksheets:
        if "Field Spec" in s.title:
            ws = s
            break

    done = 0
    for r in range(2, ws.max_row + 1):
        f = ws.cell(r, 3).value
        if f in FIX:
            for col, val in FIX[f].items():
                cell = ws.cell(r, col, val)
                if col == 13:
                    cell.font = MONO
                    cell.alignment = WRAP
            print(f"[{f}] row {r}: updated cols {sorted(FIX[f].keys())}")
            done += 1

    wb.save(XLSX)
    print(f"\nSaved: {XLSX} — {done}/3 BK fields fixed")


if __name__ == "__main__":
    main()
