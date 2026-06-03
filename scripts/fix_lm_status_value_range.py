# -*- coding: utf-8 -*-
"""
lm_status 标准接口取值范围 = 实测出现的 22 个 lmc_status（一值一行，≥5 规则），按频次排；
末附注：字典完整域约 150 码，完整 code→text 见数据字典 / Redshift portnewrezdatadic（后续整表任务）。
按表头定位列，人工列受保护。Excel 须关闭；改后运行 sync_fieldspec_excel_to_md.py。
"""
import sys, io
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent))
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from _excel_guard import col_by_header, assert_safe

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")
VALUES = [
    "Workout Denial", "Pending Financials", "Document Follow-up", "Monitor Forbearance", "Book mod",
    "Deferment Agreement Ordered", "Deferment Plan In Progress", "Monitor for pmts/funds", "Liquidation Referral",
    "Follow up for 1st Trial Payment", "Solicitation Offered", "Monitor for Mod Agreement",
    "Follow up for 2nd Trial Payment", "Countered by Supervisor", "Monitor for Mod Agreement – Final Trial Payment Due",
    "Awaiting MI Approval", "DIL Sent for Recording", "Negotiate DIL liens", "DIL Title Ordered",
    "Submitted for Approval", "Not Assigned", "Awaiting investor approval",
]
NEW = "\n".join(VALUES) + "\n（实测22种；字典完整域约150码，完整 code→text 见数据字典/Redshift portnewrezdatadic，状态流转见 doc 18 §4.5）"


def main():
    wb = load_workbook(XLSX)
    ws = next(s for s in wb.worksheets if "Field Spec" in s.title)
    fcol = col_by_header(ws, "标准接口字段") or 3
    rcol = col_by_header(ws, "标准接口取值范围")
    assert_safe(ws, rcol)
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, fcol).value == "lm_status":
            ws.cell(r, rcol).value = NEW
            ws.cell(r, rcol).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            print(f"lm_status 取值范围 -> {len(VALUES)} 值（一值一行）")
            break
    else:
        raise RuntimeError("lm_status row not found")
    wb.save(XLSX)
    print(f"Saved: {XLSX}")


if __name__ == "__main__":
    main()
