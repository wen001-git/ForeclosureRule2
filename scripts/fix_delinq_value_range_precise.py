# -*- coding: utf-8 -*-
"""
把 Field Spec 中 delinquency_status 的「标准接口取值范围」改为精确的 19 个枚举值（每个 DPD 桶逐一列全、
顶层一律用 |），避免 `1-29 / 30-59 / … DPD` 压缩被误读为单个取值。

19 值经 DB 实测（newrez.portnewrezgeneral.delinquency_status_mba 最新快照 distinct）确认，与 doc 14
「delinquency_status 允许值表」一致。Excel 存裸 |（sync 脚本 esc() 会转义为 \\| 写入 md，渲染为 |）。

按表头定位列，人工列受保护。Excel 须关闭；改后运行 sync_fieldspec_excel_to_md.py。
Run: python scripts/fix_delinq_value_range_precise.py
"""
import sys, io
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent))
from openpyxl import load_workbook
from _excel_guard import col_by_header, assert_safe

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")
NEW = ("MBA标准文本枚举（Servicer传输层）：Current | 1-29 DPD | 30-59 DPD | 60-89 DPD | 90-119 DPD | "
       "120-149 DPD | 150-179 DPD | 180+ DPD | Foreclosure | Foreclosure / Perf BK | "
       "Foreclosure / Non-Perf BK | REO | Performing Bankruptcy | Non-Performing Bankruptcy | "
       "Full Payoff | REO Sale | 3rd Party Sale | Service Release | Settlement"
       "（共19种；禁止数字串如'29.0'；完整映射见doc 08 / doc 14 允许值表）")


def main():
    wb = load_workbook(XLSX)
    ws = next(s for s in wb.worksheets if "Field Spec" in s.title)
    fcol = col_by_header(ws, "标准接口字段") or 3
    rcol = col_by_header(ws, "标准接口取值范围")
    if rcol is None:
        raise RuntimeError("找不到「标准接口取值范围」列")
    assert_safe(ws, rcol)
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, fcol).value == "delinquency_status":
            old = ws.cell(r, rcol).value
            ws.cell(r, rcol).value = NEW
            n = NEW.count("|") + 1
            print(f"row {r} 标准接口取值范围(col {rcol}) 更新为 {n} 个枚举值")
            print(f"  OLD: {old}")
            print(f"  NEW: {NEW}")
            break
    else:
        raise RuntimeError("delinquency_status row not found")
    wb.save(XLSX)
    print(f"Saved: {XLSX}")


if __name__ == "__main__":
    main()
