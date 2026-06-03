# -*- coding: utf-8 -*-
"""
更正 Field Spec 中 active_fcl_flag 的语义：activefcflag=0 不是“已完结”，而是“当前不处于活跃止赎流程”。

按表头定位列（标准接口取值范围 / 业务含义），人工列受保护（assert_safe）。Excel 须关闭。
Run: python scripts/fix_active_fcl_flag_semantics.py
"""
import sys, io
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent))
from openpyxl import load_workbook
from _excel_guard import col_by_header, assert_safe

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")
NEW_RANGE = "1（活跃止赎中）/ 0（当前不处于活跃止赎流程；含已完成 REO/3rd Party/DiL 或已退出 Reinstated/Loss Mitigation/Paid in Full）；NULL 保守视为 1"
NEW_MEANING = "FCL 是否处于活跃止赎流程；1=活跃止赎中，0=当前不处于活跃止赎（BPS 统称 Closed Foreclosure，含完成/撤销/复议/付清，非都已完成）；历史 NULL 须 NULL-safe 处理（保守视为 1）"


def main():
    wb = load_workbook(XLSX)
    ws = next(s for s in wb.worksheets if "Field Spec" in s.title)
    field_col = col_by_header(ws, "标准接口字段") or 3
    range_col = col_by_header(ws, "标准接口取值范围")
    mean_col = col_by_header(ws, "业务含义")
    for c in (range_col, mean_col):
        if c:
            assert_safe(ws, c)
    done = False
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, field_col).value == "active_fcl_flag":
            if range_col:
                old = ws.cell(r, range_col).value
                ws.cell(r, range_col).value = NEW_RANGE
                print(f"row {r} 取值范围(col {range_col}):\n  OLD: {old!r}\n  NEW: {NEW_RANGE!r}")
            if mean_col:
                oldm = ws.cell(r, mean_col).value
                ws.cell(r, mean_col).value = NEW_MEANING
                print(f"row {r} 业务含义(col {mean_col}) updated (OLD: {oldm!r})")
            done = True
            break
    if not done:
        raise RuntimeError("active_fcl_flag row not found")
    wb.save(XLSX)
    print(f"Saved: {XLSX}")


if __name__ == "__main__":
    main()
