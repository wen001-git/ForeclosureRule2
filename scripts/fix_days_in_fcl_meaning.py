# -*- coding: utf-8 -*-
"""
更新 servicer_days_in_fcl / days_in_fcl 的「业务含义」，点明 referral vs setup 起算口径（代码+DB 核实）。
按表头定位列，人工列受保护。Excel 须关闭；改后运行 sync_fieldspec_excel_to_md.py。
"""
import sys, io
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent))
from openpyxl import load_workbook
from _excel_guard import col_by_header, assert_safe

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")
BIZ = {
    "servicer_days_in_fcl": ("Servicer 口径的 FCL 在途天数（Newrez/SMS=Shellpoint Mortgage Servicing 实现为原生 "
                        "smsdaysinfc / ETL 别名 svc_days_infc，直接透传），自 fcsetupdate（servicer FCL 建案日）起算。"
                        "BPS 实时修正：smsdaysinfc + DATEDIFF(今日纽约, dataasof)。因 setup≥referral，故 ≤ days_in_fcl（投资人全程口径）。"
                        "注：标准接口字段名取 servicer-中性名 servicer_days_in_fcl"),
    "days_in_fcl": ("投资人 / 全程视角的 FCL 在途天数。ETL 按 fcreferraldate（转介日）重算："
                    "datediff(referral_start_date, snapshot)+1（仅 Active；basic_data_pool_config.py:1628）。"
                    "与 servicer_days_in_fcl 的区别=起算基准不同（referral vs setup），故 days_in_fcl ≥ servicer_days_in_fcl；"
                    "BPS 同样 + DATEDIFF(今日纽约, dataasof) 实时修正"),
}


def main():
    wb = load_workbook(XLSX)
    ws = next(s for s in wb.worksheets if "Field Spec" in s.title)
    fcol = col_by_header(ws, "标准接口字段") or 3
    bcol = col_by_header(ws, "业务含义")
    assert_safe(ws, bcol)
    done = []
    for r in range(2, ws.max_row + 1):
        f = ws.cell(r, fcol).value
        if f in BIZ:
            ws.cell(r, bcol).value = BIZ[f]
            done.append(f)
    miss = set(BIZ) - set(done)
    if miss:
        raise RuntimeError(f"未找到字段行: {miss}")
    wb.save(XLSX)
    print(f"更新业务含义: {done} -> Saved {XLSX}")


if __name__ == "__main__":
    main()
