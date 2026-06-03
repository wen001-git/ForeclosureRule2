"""
Add a bilingual glossary sheet (术语说明 / Glossary) as the 2nd tab of
14_servicer_fcl_field_spec.xlsx (immediately after 阅读指南 Guide).

Run from project root: python scripts/add_glossary.py
"""

import sys
import io
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

XLSX = Path(__file__).parent.parent / "docs" / "14_servicer_fcl_field_spec.xlsx"
SHEET_NAME = "📚 术语说明 Glossary"

C_DARK_BLUE  = "1F3864"
C_MID_BLUE   = "2E75B6"
C_LIGHT_BLUE = "D6E4F0"
C_PALE_BLUE  = "EBF3FA"
C_YELLOW     = "FFF2CC"
C_GREEN      = "E2EFDA"
C_ORANGE     = "FCE4D6"
C_WHITE      = "FFFFFF"

def thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def f(bold=False, size=10, color="1F1F1F", italic=False):
    return Font(name="微软雅黑", bold=bold, size=size, color=color, italic=italic)

def al(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── Glossary data ─────────────────────────────────────────────────────────────
# (term_zh, term_en, category, definition_zh, definition_en)
GLOSSARY = [
    # ── Core System ──────────────────────────────────────────────────────────
    ("BPS", "BPS (Business Processing System)",
     "系统",
     "贷款资产管理平台，内部运营系统。Foreclosure 模块包含五大功能面板，用于跟踪和展示法拍流程状态。",
     "Internal mortgage asset management platform. The Foreclosure module contains 5 functional panels for tracking and displaying foreclosure process status."),

    ("ETL", "ETL (Extract-Transform-Load)",
     "系统",
     "数据管道：从 Servicer 原始数据文件提取字段，经清洗/转换后写入 BPS 数据库。",
     "Data pipeline: extracts fields from Servicer raw data files, cleans/transforms them, and loads them into the BPS database."),

    ("Servicer", "Servicer (贷款服务商)",
     "业务主体",
     "代表投资方管理抵押贷款的公司，负责收款、客户沟通、止赎流程执行等。本文以 Newrez 为合规基准。",
     "A company that manages mortgage loans on behalf of investors; handles collections, borrower communication, and foreclosure execution. This document uses Newrez as the compliance baseline."),

    ("Newrez / Shellpoint", "Newrez / Shellpoint",
     "业务主体",
     "本文档的合规基准 Servicer。Shellpoint Mortgage Servicing 为 Newrez 的运营品牌名。",
     "The compliance-baseline Servicer for this document. Shellpoint Mortgage Servicing is Newrez's operating brand name."),

    # ── Foreclosure Process ───────────────────────────────────────────────────
    ("FCL (Foreclosure / 法拍)",
     "FCL — Foreclosure",
     "法拍流程",
     "当借款人持续逾期时，贷款方通过法律程序收回抵押房产的过程。美国各州分为司法程序（需法院）和非司法程序（无需法院）两类。",
     "The legal process by which a lender reclaims mortgaged property after borrower default. US states use either judicial (court-involved) or non-judicial (power-of-sale) procedures."),

    ("NOI", "NOI — Notice of Intent / Initiation",
     "法拍流程",
     "法拍启动前向借款人发出的正式意向通知，标志着法拍程序的开始。",
     "Formal notification sent to the borrower before foreclosure initiation; marks the start of the formal foreclosure process."),

    ("Demand Date", "Demand Date (催款起始日)",
     "法拍流程",
     "向借款人发送正式催款函的日期，通常早于法拍启动日，是法律程序的前置步骤。",
     "Date of the formal demand letter sent to the borrower, typically prior to foreclosure filing; a prerequisite legal step."),

    ("Judicial / Non-Judicial", "Judicial vs. Non-Judicial Foreclosure",
     "法拍流程",
     "司法程序：需通过法院审理，周期较长（常见于纽约、新泽西等州）。非司法程序：依授权销售条款执行，无需法院，周期较短（加州等州）。",
     "Judicial: requires court proceedings, longer timeline (e.g., NY, NJ). Non-Judicial: executed via power-of-sale clause without court involvement, shorter timeline (e.g., CA)."),

    # ── Four-Dimension Flags ──────────────────────────────────────────────────
    ("四维状态基础标志",
     "Four-Dimension Status Flags (Section 2.0)",
     "字段分组",
     "BPS 内部的 12 个 P0 级状态字段，覆盖 FCL / Hold / LM / BK 四个维度，是系统入库的前提条件。Newrez 目前仅提供约 1/12（合规率 ~8%），是最大缺口。",
     "12 P0-priority BPS-internal state fields covering four dimensions: FCL / Hold / LM / BK. These are prerequisites for system ingestion. Newrez currently provides ~1/12 (~8% compliance) — the largest gap."),

    ("Hold 槽位 (Hold Slots)",
     "Hold Slots (Section 2.4)",
     "字段分组",
     "15 个字段，记录法拍被暂缓的原因及时间，分为法律类（BK / LM / 诉讼）、自然灾害、业务决策等多个槽位。",
     "15 fields tracking why and when a foreclosure is paused; categories include legal holds (BK/LM/litigation), disaster holds, and business decision holds."),

    ("LM (Loss Mitigation / 损失缓解)",
     "LM — Loss Mitigation (Section 3.1)",
     "字段分组",
     "为违约借款人提供的止赎替代方案，包括贷款修改、还款计划、宽限期、短售等。进行中的 LM 会触发 Hold 并暂停法拍。",
     "Alternatives to foreclosure offered to delinquent borrowers: loan modification, repayment plan, forbearance, short sale, etc. Active LM triggers a Hold and pauses foreclosure."),

    ("BK (Bankruptcy / 破产)",
     "BK — Bankruptcy (Section 4.1)",
     "字段分组",
     "借款人申请破产（Chapter 7/11/13）后，自动中止令（Automatic Stay）暂停所有催收和法拍行为，直至法院解除。",
     "After a borrower files bankruptcy (Ch.7/11/13), the Automatic Stay halts all collection and foreclosure actions until the court lifts it."),

    # ── Field Sections ────────────────────────────────────────────────────────
    ("Section 2.0", "Section 2.0 — 四维状态基础标志",
     "字段分区",
     "12 个 P0 字段；系统入库的最低前提，任一缺失将导致该贷款记录被 BPS 拒绝。",
     "12 P0 fields. The minimum ingestion prerequisite; any missing field causes BPS to reject the loan record."),

    ("Section 2.1", "Section 2.1 — 识别/入库过滤字段",
     "字段分区",
     "7 个贷款识别和入库过滤字段（如 loan_id, investor_type 等）。Newrez 合规率 100%。",
     "7 loan identification and ingestion-filter fields (e.g., loan_id, investor_type). Newrez compliance: 100%."),

    ("Section 2.2", "Section 2.2 — FCL 状态字段",
     "字段分区",
     "9 个字段，描述当前法拍流程所处阶段（如 fcl_status, fcl_stage）。Newrez 合规率 88%。",
     "9 fields describing the current foreclosure process stage (e.g., fcl_status, fcl_stage). Newrez compliance: 88%."),

    ("Section 2.3", "Section 2.3 — FCL 时间轴字段",
     "字段分区",
     "16 个字段，记录法拍各关键节点的日期（如首次违约日、起诉日、判决日、拍卖日等）。Newrez 合规率 ~56%，缺失 5 个字段。",
     "16 fields recording dates of key foreclosure milestones (first default date, filing date, judgment date, auction date, etc.). Newrez compliance: ~56%, 5 fields missing."),

    ("Section 2.4", "Section 2.4 — Hold 槽位字段",
     "字段分区",
     "15 个字段，覆盖各类 Hold 状态的开始/结束时间（BK Hold, LM Hold, Disaster Hold 等）。Newrez 合规率 100%（在已请求字段范围内）。",
     "15 fields covering start/end dates for each Hold type (BK Hold, LM Hold, Disaster Hold, etc.). Newrez compliance: 100% for formally requested fields."),

    ("Section 2.5", "Section 2.5 — 竞标/销售字段",
     "字段分区",
     "3 个字段，记录拍卖出价和销售结果（如 bid_amount, sale_date）。Newrez 合规率 83%。",
     "3 fields for auction bid and sale outcome data (e.g., bid_amount, sale_date). Newrez compliance: 83%."),

    ("Section 2.6", "Section 2.6 — 贷款属性增强字段",
     "字段分区",
     "12 个字段，提供贷款基本属性（余额、利率、产权类型等），供 BPS 面板增强显示用。Newrez 合规率 ~75%。",
     "12 fields providing basic loan attributes (balance, interest rate, property type, etc.) for enhanced BPS panel display. Newrez compliance: ~75%."),

    ("Section 3.1", "Section 3.1 — LM 周期字段",
     "字段分区",
     "10 个字段，记录损失缓解活动的起止日期和结果。Newrez 合规率 70%，缺失 3 个字段。",
     "10 fields for loss mitigation activity dates and outcomes. Newrez compliance: 70%, 3 fields missing."),

    ("Section 4.1", "Section 4.1 — 破产字段",
     "字段分区",
     "11 个字段，记录破产案件信息（申请章节、申请日期、自动中止解除日等）。Newrez 合规率 91%。",
     "11 fields for bankruptcy case information (chapter filed, filing date, stay lift date, etc.). Newrez compliance: 91%."),

    # ── Priority & Compliance ─────────────────────────────────────────────────
    ("P0", "P0 — System Ingestion Prerequisite",
     "优先级",
     "入库前提条件。缺失任一 P0 字段，BPS 将拒绝整条贷款记录入库。",
     "System ingestion prerequisite. Missing any P0 field causes BPS to reject the entire loan record."),

    ("P1", "P1 — Core Panel Display Field",
     "优先级",
     "核心面板显示字段。缺失则对应 BPS 面板功能异常或显示空白（降级可用，不阻断入库）。",
     "Core panel display field. Missing P1 fields cause BPS panel display gaps or blank data (system still ingests the loan)."),

    ("P2", "P2 — Enhanced Analytics Field",
     "优先级",
     "增强型分析字段，可选。存在时启用高级分析功能；缺失不影响主流程。",
     "Optional enhanced analytics field. When present, enables advanced analysis features; missing P2 fields don't affect the main workflow."),

    ("合规率", "Compliance Rate (合规率)",
     "度量指标",
     "已提供字段数 ÷ BPS 所需字段总数，用于衡量 Servicer 当前数据交付的完整性。",
     "Provided fields ÷ total BPS-required fields. Measures the completeness of a Servicer's current data delivery."),

    ("已提供 / 部分提供 / 未提供",
     "Provided / Partial / Not Provided",
     "度量指标",
     "✅已提供：Newrez 已交付且 BPS 可直接读取。⚠️部分提供：已交付但有覆盖率或质量问题。❌未提供：Newrez 未交付，BPS 显示空白。",
     "✅ Provided: Newrez delivers the field and BPS can read it directly. ⚠️ Partial: delivered but with coverage or quality issues. ❌ Not Provided: Newrez does not deliver the field; BPS shows blank."),

    ("days360",
     "days360 (30/360 day-count)",
     "度量指标",
     "按 30/360 惯例计算两日期天数差的函数（每月30天/年360天）；公式 = (末.年−始.年)×360 + (末.月−始.月)×30 + (末.日−始.日)。"
     "days360(nextduedate, dataasof) = 从下次应还款日到数据快照日的天数 = 逾期天数 DPD（逾期为正）。"
     "ETL 分档：<30→C · <60→D30 · <90→D60 · <120→D90 · ≥120→D120P（永不产生 FCL）。源码 PrefectFlow/flow/remit_validation/utils.py:14-21。",
     "30/360 day-count function (each month=30 days, year=360); formula = (end.yr−start.yr)×360 + (end.mo−start.mo)×30 + (end.day−start.day). "
     "days360(nextduedate, dataasof) = days from next-payment-due date to snapshot date = DPD (positive when overdue). "
     "ETL buckets: <30→C, <60→D30, <90→D60, <120→D90, ≥120→D120P (never yields FCL). Source: PrefectFlow/flow/remit_validation/utils.py:14-21."),
]

CATEGORIES = [
    ("系统",     "D6E4F0"),
    ("业务主体", "E2EFDA"),
    ("法拍流程", "FFF2CC"),
    ("字段分组", "FCE4D6"),
    ("字段分区", "EBF3FA"),
    ("优先级",   "F4E6FF"),
    ("度量指标", "E8F5E9"),
]
CAT_COLOR = {k: v for k, v in CATEGORIES}


def build_glossary(ws):
    ws.sheet_view.showGridLines = False

    # column widths: A(term_zh) B(term_en) C(category) D(def_zh) E(def_en)
    col_widths = [22, 28, 10, 38, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # Title
    ws.row_dimensions[row].height = 28
    t = ws.cell(row, 1, "📚  术语说明 / Glossary — 14_servicer_fcl_field_spec.xlsx")
    t.font = Font(name="微软雅黑", bold=True, size=14, color=C_WHITE)
    t.fill = PatternFill("solid", fgColor=C_DARK_BLUE)
    t.alignment = al(h="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

    # Subtitle
    ws.row_dimensions[row].height = 16
    sub = ws.cell(row, 1,
        "本表收录文档中使用的所有领域术语，按类别分组，含中英双语说明。"
        "  |  This table lists all domain terms used in the document, grouped by category, with bilingual definitions.")
    sub.font = Font(name="微软雅黑", size=9, italic=True, color="595959")
    sub.alignment = al(h="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 2

    # Column headers
    ws.row_dimensions[row].height = 22
    headers = ["术语（中文）", "Term (English)", "类别", "中文说明", "English Definition"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.font = Font(name="微软雅黑", bold=True, size=10, color=C_WHITE)
        cell.fill = PatternFill("solid", fgColor=C_MID_BLUE)
        cell.alignment = al(h="center", wrap=False)
        cell.border = thin()
    row += 1

    # Group by category and write rows
    seen_cats = []
    for term_zh, term_en, cat, def_zh, def_en in GLOSSARY:
        if cat not in seen_cats:
            seen_cats.append(cat)
            # Category sub-header
            ws.row_dimensions[row].height = 18
            bg = CAT_COLOR.get(cat, C_LIGHT_BLUE)
            sec = ws.cell(row, 1, f"▌ {cat}")
            sec.font = Font(name="微软雅黑", bold=True, size=10, color=C_DARK_BLUE)
            sec.fill = PatternFill("solid", fgColor=bg)
            sec.alignment = al(h="left", wrap=False)
            sec.border = thin()
            for c in range(2, 6):
                cell = ws.cell(row, c)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.border = thin()
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            row += 1

        bg = CAT_COLOR.get(cat, C_LIGHT_BLUE)
        # alternate shade for readability
        shade = "F7FBFF" if seen_cats.index(cat) % 2 == 0 else C_PALE_BLUE

        ws.row_dimensions[row].height = 52
        data = [term_zh, term_en, cat, def_zh, def_en]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row, c, val)
            cell.font = f(bold=(c <= 2), size=10)
            cell.fill = PatternFill("solid", fgColor=shade)
            cell.alignment = al(h="center" if c == 3 else "left")
            cell.border = thin()
        row += 1

    # Footer
    row += 1
    ftr = ws.cell(row, 1,
        "来源：doc 14 BPS-driven Servicer FCL Interface Spec · doc 09 Servicer Data Interface Standard · doc 10 Glossary")
    ftr.font = Font(name="微软雅黑", size=8, italic=True, color="A0A0A0")
    ftr.alignment = al(h="left", wrap=False)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)


def main():
    if not XLSX.exists():
        raise FileNotFoundError(f"Not found: {XLSX}")

    wb = load_workbook(XLSX)

    if SHEET_NAME in wb.sheetnames:
        print("Sheet already exists — removing and recreating.")
        del wb[SHEET_NAME]

    # Insert as 2nd tab (after 阅读指南)
    ws = wb.create_sheet(SHEET_NAME, 1)
    build_glossary(ws)

    wb.save(XLSX)
    print(f"Saved: {XLSX}")
    print(f"Sheets ({len(wb.sheetnames)}):")
    for i, name in enumerate(wb.sheetnames):
        print(f"  {i}: {name}")


if __name__ == "__main__":
    main()
