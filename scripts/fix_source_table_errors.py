"""
Fix 5 source-table errors in the Field Spec sheet of 14_servicer_fcl_field_spec.xlsx:
col4 (Newrez原始字段 完整路径) was wrong for 5 fields, and col13 (验证SQL) inherited the error.

DB-verified corrections (2026-06-02):
  delinquency_status -> portnewrezgeneral.delinquency_status_mba
  foreclosure_flag   -> fcl_flag does NOT exist; FCL active = activefcflag
  lm_flag            -> portnewrezlm.activelmflag (not portnewrezfc)
  state              -> portnewrezprop.propertystate
  lien_position      -> portnewrezgeneral.lienposition

Run: python scripts/fix_source_table_errors.py
"""

import sys, io
from pathlib import Path
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")

# field -> (new col4, new col13 SQL)
FIXES = {
    "delinquency_status": (
        "newrez.portnewrezgeneral.delinquency_status_mba",
        "-- 验证 delinquency_status 取值范围\n"
        "-- 源表: newrez.portnewrezgeneral | 运行于: mysql_dev\n"
        "SELECT delinquency_status_mba AS val, COUNT(*) AS cnt\n"
        "FROM   newrez.portnewrezgeneral\n"
        "WHERE  delinquency_status_mba IS NOT NULL AND delinquency_status_mba != ''\n"
        "GROUP  BY delinquency_status_mba ORDER BY cnt DESC LIMIT 30;",
    ),
    "foreclosure_flag": (
        "— (Newrez无fcl_flag列；FCL活跃状态见 activefcflag / active_fcl_flag)",
        "-- N/A Newrez 无独立 fcl_flag 列（实测确认）；FCL 活跃状态由 activefcflag 表达\n"
        "-- 源表: newrez.portnewrezfc | 运行于: mysql_dev\n"
        "SELECT activefcflag, COUNT(*) AS cnt\n"
        "FROM   newrez.portnewrezfc\n"
        "GROUP  BY activefcflag ORDER BY cnt DESC;",
    ),
    "lm_flag": (
        "newrez.portnewrezlm.activelmflag",
        "-- 验证 lm_flag 取值范围\n"
        "-- 源表: newrez.portnewrezlm | 运行于: mysql_dev\n"
        "SELECT activelmflag, COUNT(*) AS cnt\n"
        "FROM   newrez.portnewrezlm\n"
        "GROUP  BY activelmflag ORDER BY cnt DESC;",
    ),
    "state": (
        "newrez.portnewrezprop.propertystate",
        "-- 验证 state 取值范围\n"
        "-- 源表: newrez.portnewrezprop | 运行于: mysql_dev\n"
        "SELECT propertystate AS val, COUNT(*) AS cnt\n"
        "FROM   newrez.portnewrezprop\n"
        "WHERE  propertystate IS NOT NULL AND propertystate != ''\n"
        "GROUP  BY propertystate ORDER BY cnt DESC LIMIT 60;",
    ),
    "lien_position": (
        "newrez.portnewrezgeneral.lienposition（另 portnewrezfc.jr_sr_lien_flag 为顺位标志）",
        "-- 验证 lien_position 取值范围\n"
        "-- 源表: newrez.portnewrezgeneral | 运行于: mysql_dev\n"
        "SELECT lienposition AS val, COUNT(*) AS cnt\n"
        "FROM   newrez.portnewrezgeneral\n"
        "WHERE  lienposition IS NOT NULL AND lienposition != ''\n"
        "GROUP  BY lienposition ORDER BY cnt DESC;",
    ),
}


def main():
    wb = load_workbook(XLSX)
    ws = None
    for s in wb.worksheets:
        if "Field Spec" in s.title:
            ws = s
            break

    fixed = 0
    for r in range(2, ws.max_row + 1):
        f = ws.cell(r, 3).value
        if f in FIXES:
            new_col4, new_sql = FIXES[f]
            old4 = ws.cell(r, 4).value
            ws.cell(r, 4).value = new_col4
            ws.cell(r, 13).value = new_sql
            print(f"[{f}] row {r}")
            print(f"  col4: {old4!r}\n     -> {new_col4!r}")
            print(f"  col13 SQL updated -> {new_sql.splitlines()[1]}")
            fixed += 1

    wb.save(XLSX)
    print(f"\nSaved: {XLSX} — {fixed}/5 fields fixed")


if __name__ == "__main__":
    main()
