# -*- coding: utf-8 -*-
"""
标准接口字段重命名：sms_days_in_fcl → servicer_days_in_fcl（去除单一 servicer 品牌名 SMS=Shellpoint）。
仅改 doc 14 Field Spec：该行 col C（标准接口字段）、col13（验证SQL 注释里的字段名 token；query 用 smsdaysinfc 不变）、
days_in_fcl 行业务含义里的 "sms_days_in_fcl" 提及、并把 sms 行业务含义首句点明 servicer 口径。
保留 Newrez 列 smsdaysinfc / BPS 列 summary_sms_days_in_fcl / UI "SMS Days" 不动。人工列受保护。Excel 须关闭。
改后运行 sync_fieldspec_excel_to_md.py。
"""
import sys, io
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent))
from openpyxl import load_workbook
from _excel_guard import col_by_header, assert_safe

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")
OLD, NEW = "sms_days_in_fcl", "servicer_days_in_fcl"
SMS_BIZ = ("Servicer 口径的 FCL 在途天数（Newrez/SMS=Shellpoint Mortgage Servicing 实现为原生 smsdaysinfc / "
           "ETL 别名 svc_days_infc，透传），自 servicer 建案日 fcsetupdate 起算。BPS 实时修正：smsdaysinfc + "
           "DATEDIFF(今日纽约, dataasof)。因 setup≥referral，故 ≤ days_in_fcl（投资人全程口径）。"
           "注：标准接口字段名取 servicer-中性名 servicer_days_in_fcl；smsdaysinfc/summary_sms_days_in_fcl 为 Newrez/BPS 真实列名")


def main():
    wb = load_workbook(XLSX)
    ws = next(s for s in wb.worksheets if "Field Spec" in s.title)
    fcol = col_by_header(ws, "标准接口字段") or 3
    bcol = col_by_header(ws, "业务含义")
    sqlcol = col_by_header(ws, "验证SQL")
    for c in (fcol, bcol, sqlcol):
        assert_safe(ws, c)
    changed = []
    for r in range(2, ws.max_row + 1):
        f = ws.cell(r, fcol).value
        if f == OLD:
            ws.cell(r, fcol).value = NEW
            ws.cell(r, bcol).value = SMS_BIZ
            sv = ws.cell(r, sqlcol).value
            if isinstance(sv, str) and OLD in sv:
                ws.cell(r, sqlcol).value = sv.replace(OLD, NEW)
            changed.append(f"col C {OLD}->{NEW} (row {r}) + 业务含义 + col13 注释")
        elif f == "days_in_fcl":
            bv = ws.cell(r, bcol).value
            if isinstance(bv, str) and OLD in bv:
                ws.cell(r, bcol).value = bv.replace(OLD, NEW)
                changed.append(f"days_in_fcl 业务含义提及 {OLD}->{NEW} (row {r})")
    if not any(NEW in str(ws.cell(r, fcol).value) for r in range(2, ws.max_row + 1)):
        raise RuntimeError("rename failed: servicer_days_in_fcl not found after edit")
    wb.save(XLSX)
    print("\n".join(changed) or "(no change)")
    print(f"Saved: {XLSX}")


if __name__ == "__main__":
    main()
