"""
Fix the wrong lm_flag Newrez状态 cell in Field Spec.

DB-verified (latest snapshot, newrez.portnewrezlm): activelmflag is 100% populated
(5052/5052, 0 nulls; 0:5018, 1:34) — so "🟡 部分提供（标志存在，无类型/日期）" is wrong:
the flag is fully provided (→ ✅), and LM type/dates are separate fields.

Header-resolved + 人工列保护 per CLAUDE.md. Excel must be closed.
Run: python scripts/fix_lm_flag_status.py
"""
import sys, io
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent))
from openpyxl import load_workbook
from _excel_guard import col_by_header, assert_safe

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")
NEW = ("✅ 已提供（activelmflag，最新快照 100% 填充 0/1、无 null；实测 0:5018·1:34）。"
       "以 0/1 表达需映射 Y/N；为 LM 周期表层标志，类型/起止日见 lm_type / lm_start_date / lm_end_date").replace("0:5018·1:34", "0:5018 | 1:34")


def main():
    wb = load_workbook(XLSX)
    ws = None
    for s in wb.worksheets:
        if "Field Spec" in s.title:
            ws = s
            break
    field_col = col_by_header(ws, "标准接口字段") or 3
    status_col = col_by_header(ws, "Newrez状态")
    if status_col is None:
        raise RuntimeError("找不到「Newrez状态」列")
    assert_safe(ws, status_col)

    for r in range(2, ws.max_row + 1):
        if ws.cell(r, field_col).value == "lm_flag":
            old = ws.cell(r, status_col).value
            ws.cell(r, status_col).value = NEW
            print(f"row {r} Newrez状态(col {status_col}):\n  OLD: {old!r}\n  NEW: {NEW!r}")
            break
    wb.save(XLSX)
    print(f"Saved: {XLSX}")


if __name__ == "__main__":
    main()
