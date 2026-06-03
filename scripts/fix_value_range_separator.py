# -*- coding: utf-8 -*-
"""
把 Field Spec 中「标准接口取值范围」(col9) 与「Newrez状态」(col12) 作为值分隔符的中点 `·` 改为 ` | `
（中点密集时人眼难辨；与「验证结果」列已采用的 `|` 统一）。

按表头定位列，人工列受保护（assert_safe）。`/`（DPD 分组 / 值内字符）保留不动。Excel 须关闭。
改后须运行 scripts/sync_fieldspec_excel_to_md.py 重新生成 zh 卡片。
Run: python scripts/fix_value_range_separator.py
"""
import sys, io
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent))
from openpyxl import load_workbook
from _excel_guard import col_by_header, assert_safe

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")
TARGET_HEADERS = ["标准接口取值范围", "Newrez状态"]


def norm(v):
    # 先处理带空格的 ' · '，再兜底无空格的 '·'，统一成 ' | '
    return v.replace(" · ", " | ").replace("·", " | ")


def main():
    wb = load_workbook(XLSX)
    ws = next(s for s in wb.worksheets if "Field Spec" in s.title)
    cols = []
    for h in TARGET_HEADERS:
        c = col_by_header(ws, h)
        if c is None:
            raise RuntimeError(f"找不到列「{h}」")
        assert_safe(ws, c)  # 绝不写入人工列
        cols.append((h, c))

    total = 0
    for h, c in cols:
        n = 0
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "·" in v:
                ws.cell(r, c).value = norm(v)
                n += 1
        print(f"  列「{h}」(col {c})：{n} 格的 · 已改为 |")
        total += n
    wb.save(XLSX)
    print(f"Saved: {XLSX}（共 {total} 格）")


if __name__ == "__main__":
    main()
