"""
Run each col13 verify SQL in the Field Spec sheet and write its result into a new
col14「验证结果（实测 2026-06-02）」.

One pymysql connection to bridg004-db-test (same as scripts/extract_table_stats.py) queries
both newrez.* and bpms_dev.* cross-schema. Read-only. Excel must be closed.

Run: python scripts/run_verify_sql_results.py
"""

import sys, io, os, re, datetime
from pathlib import Path
import pymysql
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# DB 凭据从环境变量读取（勿硬编码）：必填 FCL_DB_PASSWORD；host/port/user 有默认值，可用同名环境变量覆盖
HOST = os.environ.get("FCL_DB_HOST", "bridg004-db-test.mysql.database.azure.com")
PORT = int(os.environ.get("FCL_DB_PORT", "3306"))
USER = os.environ.get("FCL_DB_USER", "brgdev")
PASSWORD = os.environ.get("FCL_DB_PASSWORD")

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")
RUN_DATE = "2026-06-02"
COL = 14
MAXLEN = 900


def extract_sql(cell_text):
    """Strip -- comment lines; return the SQL statement (up to first ;). None if no SQL."""
    if not isinstance(cell_text, str):
        return None
    lines = [ln for ln in cell_text.splitlines() if not ln.strip().startswith("--")]
    sql = "\n".join(lines).strip()
    if not sql:
        return None
    # take up to first semicolon (one statement; CTE WITH...SELECT stays intact)
    stmt = sql.split(";")[0].strip()
    return stmt or None


def fmt_val(v):
    if v is None:
        return "NULL"
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def format_result(cols, rows):
    """Format compactly based on the column shape."""
    if not rows:
        return "（无数据 / 空结果）"
    lc = [c.lower() for c in cols]

    # data_date prefix if present
    date_prefix = ""
    if "data_date" in lc:
        di = lc.index("data_date")
        dval = fmt_val(rows[0][di])
        date_prefix = f"[data_date {dval}] "

    # distribution / binary / bk_code: has 'cnt'
    if "cnt" in lc:
        ci = lc.index("cnt")
        # value columns = everything except data_date and cnt
        val_idx = [i for i in range(len(cols)) if i != ci and lc[i] != "data_date"]
        pairs = []
        for r in rows:
            label = " ".join(fmt_val(r[i]) for i in val_idx) if val_idx else "(总)"
            pairs.append(f"{label}:{fmt_val(r[ci])}")
        return date_prefix + " | ".join(pairs)

    # single-row count-only (e.g. hold_loans)
    if len(rows) == 1 and any(k in lc for k in ("hold_loans",)) :
        parts = [f"{cols[i]}={fmt_val(rows[0][i])}" for i in range(len(cols)) if lc[i] != "data_date"]
        return date_prefix + " ".join(parts)

    # sample rows (loanid, dataasof, <value>) — list the value column
    val_idx = [i for i in range(len(cols)) if lc[i] not in ("loanid", "data_date", "dataasof")]
    if val_idx:
        vi = val_idx[-1]
        vals = [fmt_val(r[vi]) for r in rows]
        # prefix with dataasof if present (sample queries select dataasof, not data_date)
        pre = date_prefix
        if not pre and "dataasof" in lc:
            pre = f"[data_date {fmt_val(rows[0][lc.index('dataasof')])}] "
        return pre + " | ".join(vals) + f"（{len(rows)}笔）"

    # fallback: first column of each row
    return date_prefix + " | ".join(fmt_val(r[0]) for r in rows) + f"（{len(rows)}行）"


def main():
    conn = pymysql.connect(host=HOST, port=PORT, user=USER, password=PASSWORD, autocommit=True,
                           connect_timeout=20, read_timeout=180)
    sys.path.insert(0, str(Path(__file__).parent))
    from _excel_guard import col_by_header, next_free_col, assert_safe  # 人工列保护

    wb = load_workbook(XLSX)
    ws = None
    for s in wb.worksheets:
        if "Field Spec" in s.title:
            ws = s
            break

    # 按表头定位列（不硬编码列号），保护用户「人工」列
    field_col = col_by_header(ws, "标准接口字段") or 3
    sql_col = col_by_header(ws, "验证SQL")
    if sql_col is None:
        raise RuntimeError("找不到「验证SQL」列；请先运行 add_field_spec_verify_sql.py")
    COL = col_by_header(ws, "验证结果") or next_free_col(ws)
    assert_safe(ws, COL)  # 绝不写入「人工」列

    # header
    h = ws.cell(1, COL, f"验证结果（实测 {RUN_DATE}）")
    h.fill = PatternFill("solid", fgColor="2E75B6")
    h.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    side = Side(style="thin", color="BFBFBF")
    h.border = Border(left=side, right=side, top=side, bottom=side)
    ws.column_dimensions[get_column_letter(COL)].width = 62

    mono = Font(name="Consolas", size=8)
    align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ok = na = err = 0
    errors = []
    cur = conn.cursor()
    for r in range(2, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        field = ws.cell(r, field_col).value
        if c1 is None or (isinstance(c1, str) and c1.startswith("Sect")) or not isinstance(field, str):
            continue
        sql = extract_sql(ws.cell(r, sql_col).value)
        if sql is None:
            result = "N/A（无可验证源）"
            na += 1
        else:
            try:
                cur.execute(sql)
                rows = cur.fetchall()
                cols = [d[0] for d in cur.description]
                result = format_result(cols, rows)
                ok += 1
            except Exception as e:
                msg = str(e).splitlines()[0][:120]
                result = f"⚠️ 执行出错: {msg}"
                errors.append((field, msg))
                err += 1
        if len(result) > MAXLEN:
            result = result[:MAXLEN] + " …"
        cell = ws.cell(r, COL, result)
        cell.font = mono
        cell.alignment = align
        cell.border = Border(left=side, right=side, top=side, bottom=side)
        print(f"  {field:32} -> {'OK ' if sql else 'N/A'} {result[:60]}")

    cur.close()
    conn.close()
    wb.save(XLSX)
    print(f"\nSaved: {XLSX}")
    print(f"ok={ok}  na={na}  err={err}")
    if errors:
        print("ERRORS:")
        for f, m in errors:
            print(f"  [{f}] {m}")


if __name__ == "__main__":
    main()
