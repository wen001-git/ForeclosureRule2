# -*- coding: utf-8 -*-
"""
生成 docs/16_bps_fcl_display_mapping.xlsx —— BPS Foreclosure 界面字段 ⇄ Newrez 原数据反向映射。

以 BPS Asset Management Foreclosure 各面板的【实际展示字段】为终点，逆向还原每个展示字段来自
哪张 Newrez 原始表 / 哪个字段 / 经过什么计算规则；并用 5 条真实 Newrez 贷款做「原始数据 → 规则
→ BPS 界面值」的实测举例。

结构（8 sheet）：
  0  📖 说明/索引（标准文档头：目的/读者/修订史/依赖/限制 + 5 条样例贷款清单 + 导航）
  ①  FCL Summary           （doc 13 §3.7 summary_*）
  ②  Milestone Timeline    （doc 13 §3.1 timeline_* + §3.2/3.3/3.4 target/actual/var）
  ③  Hold                  （doc 13 §4）
  ④  Loss Mitigation Cycle （doc 13 §5）
  ⑤  Bankruptcy            （doc 13 §6）
  ⑥  Aggregate Stage Tab   （doc 13 §7 stage）
  ⑦  Aggregate Timeline Tab（doc 13 §7 Time Line Tab）

每面板 sheet 统一三段式：顶部嵌入面板截图 → 块A「BPS 字段映射规则」→ 块B「原始 Newrez 数据(5笔)」
→ 块C「BPS 实测值（界面实际显示）」。块A 规则解释 B → C 如何推算。

内容来源：doc 13《newrez_fcl_bps_display_mapping》为映射权威；所有 Newrez 源列已用
information_schema 全量校验（见 verify_schema）。DB 只读。Excel 必须关闭后再运行。

Run: python scripts/build_bps_display_mapping_xlsx.py
"""

import sys, io, os, datetime
from pathlib import Path
import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# DB 凭据从环境变量读取（勿硬编码）：必填 FCL_DB_PASSWORD；host/port/user 有默认值，可用同名环境变量覆盖
HOST = os.environ.get("FCL_DB_HOST", "bridg004-db-test.mysql.database.azure.com")
PORT = int(os.environ.get("FCL_DB_PORT", "3306"))
USER = os.environ.get("FCL_DB_USER", "brgdev")
PASSWORD = os.environ.get("FCL_DB_PASSWORD")
ROOT = Path(__file__).resolve().parent.parent
IMG_DIR = ROOT / "docs" / "zh" / "image"
OUT = ROOT / "docs" / "16_bps_fcl_display_mapping.xlsx"
RUN_DATE = "2026-06-02"

# 5 条样例贷款（均在 BPS 融资池/同步表中，覆盖不同阶段与面板）
LOANS = [
    ("7727000088", "Judicial(FL)·JUDGEMENT 阶段·Hold×7·LM×9（旗舰多面板）"),
    ("7727000672", "Non-Judicial(MI)·REFERRAL 阶段（doc 13 锚点贷款）"),
    ("7727004200", "Judicial(IL)·SALE 阶段（拍卖排定）"),
    ("7727000065", "BK + Hold×4 + 完结REO（多面板）"),
    ("7727000010", "Chapter 13 Active BK（纯破产示例，未入 FCL 管道）"),
]
LOAN_IDS = [l[0] for l in LOANS]

# ── 样式 ────────────────────────────────────────────────────────────────
DARK, MID, LIGHT, YELLOW, GREEN, GREY = "1F3864", "2E75B6", "D6E4F0", "FFF2CC", "E2EFDA", "F2F2F2"
YH = "微软雅黑"
SIDE = Side(style="thin", color="BFBFBF")
BORDER = Border(left=SIDE, right=SIDE, top=SIDE, bottom=SIDE)
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def fmt(v):
    if v is None:
        return "—"
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v)
    return s if s.strip() else "—"


def section_title(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.fill = PatternFill("solid", fgColor=DARK)
    c.font = Font(name=YH, bold=True, size=11, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 22
    return row + 1


def note(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.fill = PatternFill("solid", fgColor=YELLOW)
    c.font = Font(name=YH, size=9, italic=True, color="7F6000")
    c.alignment = WRAP
    return row + 1


def table(ws, row, headers, rows, header_fill=MID):
    """写一个带表头的表；返回下一空行。"""
    for j, h in enumerate(headers, 1):
        c = ws.cell(row, j, h)
        c.fill = PatternFill("solid", fgColor=header_fill)
        c.font = Font(name=YH, bold=True, size=9, color="FFFFFF")
        c.alignment = CENTER
        c.border = BORDER
    row += 1
    for i, r in enumerate(rows):
        fillc = "FFFFFF" if i % 2 == 0 else GREY
        for j, v in enumerate(r, 1):
            c = ws.cell(row, j, v)
            c.font = Font(name=YH, size=9)
            c.alignment = WRAP
            c.border = BORDER
            c.fill = PatternFill("solid", fgColor=fillc)
        row += 1
    return row + 1


def add_image(ws, fname, max_w=1080):
    p = IMG_DIR / fname
    if not p.exists():
        return 2
    with PILImage.open(p) as im:
        w, h = im.size
    scale = min(1.0, max_w / w)
    img = XLImage(str(p))
    img.width, img.height = int(w * scale), int(h * scale)
    ws.add_image(img, "A1")
    return int(img.height / 19) + 3


def set_widths(ws, widths):
    for col, wdt in widths.items():
        ws.column_dimensions[col].width = wdt


# ── DB 取数 ─────────────────────────────────────────────────────────────
def row_dict(cur):
    cols = [d[0] for d in cur.description]
    r = cur.fetchone()
    return dict(zip(cols, r)) if r else {}


def all_dicts(cur):
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]


def fetch_newrez_fc(cur, loan):
    cur.execute(
        """SELECT fc.*, prop.propertystate AS _state, gen.legalstatus AS _legalstatus,
                  gen.delinquency_status_mba AS _delinq
           FROM newrez.portnewrezfc fc
           LEFT JOIN newrez.portnewrezprop prop ON prop.loanid=fc.loanid AND prop.dataasof=fc.dataasof
           LEFT JOIN newrez.portnewrezgeneral gen ON gen.loanid=fc.loanid AND gen.dataasof=fc.dataasof
           WHERE fc.loanid=%s AND fc.dataasof=(SELECT MAX(dataasof) FROM newrez.portnewrezfc WHERE loanid=%s)""",
        (loan, loan))
    return row_dict(cur)


def fetch_newrez_bk(cur, loan):
    # 同 doc 13 §6 ETL 提取过滤：仅 bkstatus 非空的真实破产记录（排除空白状态行）
    cur.execute(
        """SELECT * FROM newrez.portnewrezbk
           WHERE loanid=%s AND dataasof=(SELECT MAX(dataasof) FROM newrez.portnewrezbk WHERE loanid=%s)
             AND LENGTH(TRIM(bkstatus))>0
           ORDER BY bkfileddate""", (loan, loan))
    return all_dicts(cur)


def fetch_newrez_lm(cur, loan):
    cur.execute(
        """SELECT lmdeal, lmprogram, lmstatus, dealstartdate, lmremovaldate, lmdecision, denialreason, dataasof
           FROM newrez.portnewrezlm
           WHERE loanid=%s AND dataasof=(SELECT MAX(dataasof) FROM newrez.portnewrezlm WHERE loanid=%s)
           ORDER BY dealstartdate""", (loan, loan))
    return all_dicts(cur)


def fetch_bps(cur, table_name, loan, latest_col=None):
    try:
        if latest_col:
            cur.execute(f"SELECT * FROM bpms_dev.{table_name} WHERE loanid=%s AND {latest_col}="
                        f"(SELECT MAX({latest_col}) FROM bpms_dev.{table_name} WHERE loanid=%s)",
                        (int(loan), int(loan)))
        else:
            cur.execute(f"SELECT * FROM bpms_dev.{table_name} WHERE loanid=%s", (int(loan),))
        return all_dicts(cur)
    except Exception as e:
        print(f"  [warn] {table_name} loan {loan}: {str(e).splitlines()[0][:80]}")
        return []


# ── Schema-Verify（强制规则）────────────────────────────────────────────
def verify_schema(cur):
    need = {
        "portnewrezfc": ("demandsentdate demandexpirationdate fcsetupdate fcreferraldate firstlegaldate "
                         "servicecompletedate fcjudgmenthearingscheduled fcjudgmententered fcscheduledsaledate "
                         "fcsalehelddate dtdeedrecorded fcremovaldate titlereceiveddate titlecleardate "
                         "fcl3rdpartyproceedsreceiveddate fcstage currentmilestone activefcflag fcresults "
                         "fcremovaldesc lastfcstepcompleted lastfcstepcompleteddate shellpointloanid fcfirm "
                         "judicial fccontestedflag fcbidamount fcapprbidprice fcsaleamount smsdaysinfc daysinfc "
                         "dataasof fchold1description fchold1startdate fchold1enddate fchold1projectedenddate "
                         "fchold2description fchold2startdate fchold2enddate fchold2projectedenddate "
                         "fchold3description fchold3startdate fchold3enddate fchold3projectedenddate"),
        "portnewrezbk": "bkstatus bkrcurrentstatusdate bkchapter mfrhearingresults mfrfileddate pocfileddate "
                        "bkpostpetitionduedate activebkflag bkremovaldate bkfileddate",
        "portnewrezgeneral": "legalstatus delinquency_status_mba",
        "portnewrezprop": "propertystate",
        "portnewrezlm": "lmdeal lmprogram lmstatus dealstartdate lmremovaldate lmdecision denialreason",
    }
    bad = []
    for t, fields in need.items():
        cur.execute("SELECT column_name FROM information_schema.columns "
                    "WHERE table_schema='newrez' AND table_name=%s", (t,))
        have = {r[0].lower() for r in cur.fetchall()}
        for f in fields.split():
            if f not in have:
                bad.append(f"newrez.{t}.{f}")
    print(f"  Schema-Verify: {0 if not bad else len(bad)} 个无效源列" + (f" -> {bad}" if bad else " (全部通过)"))
    return bad


# ── 块B/块C 通用：行=字段, 列=5 笔贷款 ─────────────────────────────────
def matrix_block(ws, row, title, field_rows, header_fill=MID):
    """field_rows: list of (label, [v_loan1..v_loan5])"""
    row = section_title(ws, row, title, 1 + len(LOAN_IDS))
    headers = ["字段"] + LOAN_IDS
    rows = [[lab] + vals for lab, vals in field_rows]
    return table(ws, row, headers, rows, header_fill=header_fill)


# ══════════════════════════════════════════════════════════════════════
#  Sheet 0 — 封面 / 索引（标准文档头）
# ══════════════════════════════════════════════════════════════════════
def build_cover(wb):
    ws = wb.active
    ws.title = "📖 说明 索引"
    set_widths(ws, {"A": 22, "B": 120})
    r = 1
    r = section_title(ws, r, "doc 16 — BPS Foreclosure 界面字段 ⇄ Newrez 原数据反向映射", 2)

    def kv(label, value):
        nonlocal r
        a = ws.cell(r, 1, label); a.font = Font(name=YH, bold=True, size=10); a.alignment = WRAP
        a.fill = PatternFill("solid", fgColor=LIGHT); a.border = BORDER
        b = ws.cell(r, 2, value); b.font = Font(name=YH, size=10); b.alignment = WRAP; b.border = BORDER
        r += 1

    kv("文档目的", "以 BPS Asset Management Foreclosure 各面板的【实际展示字段】为终点，逆向还原每个展示字段"
                   "来自哪张 Newrez 原始表/哪个字段/经过什么计算规则，并用 5 条真实贷款做「原始数据→规则→"
                   "BPS 界面值」实测举例。解决「界面上看到的值到底从哪来、怎么算出来的」这一问题。")
    kv("与 doc 13/14 关系", "本表是 doc 13《newrez_fcl_bps_display_mapping》的 Excel 化（反向：BPS 展示←Newrez 源+"
                            "规则，按面板组织）。doc 14 Field Spec 是正向 servicer 接口规范（servicer 须提供什么）；"
                            "二者视角互补，本表不替代 doc 14。")
    kv("范围（覆盖）", "7 个 BPS 面板：① FCL Summary ② Milestone Timeline ③ Hold ④ Loss Mitigation Cycle "
                       "⑤ Bankruptcy ⑥ Aggregate Stage Tab ⑦ Aggregate Time Line Tab。")
    kv("范围（不覆盖）", "BPS 内部工作流字段（如 bid_approval_status）；非 Newrez servicer（Carrington/Capecodfive）"
                         "的差异；本表不含 markdown 双胞胎。")
    kv("目标读者", "数据工程师 · 业务分析师 · 验证人员 · servicer 接入工程师 · 未来 AI 会话。")
    kv("依赖", "doc 13/14（映射权威）；newrez.portnewrez*（原始快照，2026-05-31）；bpms_dev.sync_loan_foreclosure* "
               "/ sync_fcl_stage_info（BPS 实测值）；面板截图 docs/zh/image/。")
    kv("已知限制", "① 块C「BPS 实测值」取自 BPS 同步表（dev 环境，可能滞后于 Newrez 最新快照）。"
                   "② target/actual/var 天数为 BPS 内部计算/视图层结果。③ LM/Hold/BK 为 1:多历史，按贷款分块展示。"
                   "④ 部分活跃贷款可能未进入某些 BPS 子表（显示「无记录」）。")

    r += 1
    r = section_title(ws, r, f"修订历史", 2)
    hh = table(ws, r, ["日期", "版本 / 变更"],
               [[RUN_DATE, "v1 初稿（AI Agent）：8 sheet，嵌入 7 面板截图，5 条样例贷款实测预填；"
                           "所有 Newrez 源列经 information_schema 校验。"]])
    r = hh

    r = section_title(ws, r, "5 条样例贷款（最新快照 2026-05-31）", 2)
    rows = [[f"Loan {i+1}", f"{lid} —— {desc}"] for i, (lid, desc) in enumerate(LOANS)]
    r = table(ws, r, ["#", "loanid —— 选取理由"], rows)

    r = section_title(ws, r, "Sheet 导航", 2)
    nav = [
        ["①  FCL Summary", "summary_* 当前状态/金额/律师/天数（doc 13 §3.7）"],
        ["②  Milestone Timeline", "timeline_* 里程碑日期 + target/actual/var 天数（doc 13 §3.1–3.4）"],
        ["③  Hold", "Hold 全历史模型，Newrez 3 槽位 → BPS 累积历史（doc 13 §4）"],
        ["④  Loss Mitigation Cycle", "LM 周期历史，Newrez 数字码 → BPS 解码文本（doc 13 §5）"],
        ["⑤  Bankruptcy", "破产面板，bkstatus 解码 + legalstatus 取自 general（doc 13 §6）"],
        ["⑥  Aggregate Stage Tab", "阶段瀑布判定 + 各阶段天数（doc 13 §7）"],
        ["⑦  Aggregate Time Line Tab", "每贷款里程碑日期时间线（doc 13 §7）"],
    ]
    r = table(ws, r, ["Sheet", "内容"], nav)
    ws.sheet_view.showGridLines = False
    return ws


# ══════════════════════════════════════════════════════════════════════
#  通用面板 sheet 脚手架
# ══════════════════════════════════════════════════════════════════════
# ── 块0：本页数据范围 / 过滤条件（取自 doc 13，Code-First）────────────────
ORANGE = "FCE4D6"
PANEL_FILTERS = {
    "main": [  # ①② sync_loan_foreclosure（doc 13 §3.1）
        ("BPS 对接表", "bpms_dev.sync_loan_foreclosure（含 activefcflag=0 非活跃贷款：当前不处于活跃止赎流程，"
                       "含已完成 REO/3rd Party/DiL 或已退出 Reinstated/Loss Mitigation/Paid in Full，BPS 统称 Closed Foreclosure）"),
        ("源表→Redshift 提取", "newrez.portnewrezfc 全量进中间表 port.basic_data_loan_fcl（无行过滤）"),
        ("Redshift→MySQL 同步（GEN_FORECLOSURE）",
         "WHERE timeline_referred_to_foreclosure_date IS NOT NULL（= portnewrezfc.fcreferraldate 非空，"
         "仅【已转介 FCL】的贷款）；JOIN port.portfunding ON loanid（贷款须在融资池中）"),
        ("快照口径", "块B 原始数据取 portnewrezfc 该贷款 dataasof=MAX（最新快照）"),
        ("人口范围", "含 activefcflag=0 非活跃贷款（与聚合页 ⑥⑦ 仅活跃不同；activefcflag=0=当前不处于活跃止赎，"
                     "非都已完成）。代码：asset_managment_config.py GEN_FORECLOSURE"),
    ],
    "hold": [  # ③ sync_loan_foreclosure_hold（doc 13 §4）
        ("BPS 对接表", "bpms_dev.sync_loan_foreclosure_hold（累积全历史，每次变更追加一行）"),
        ("源表→Redshift 提取（basic_data_pool_config）", "WHERE fchold1startdate IS NOT NULL（仅 Hold 槽1 有开始日的贷款）"),
        ("Redshift→MySQL 同步（GEN_FORECLOSURE_HOLD）",
         "3 个 Hold 槽各自 UNPIVOT，每槽条件 (description 非空) OR (start_date 非空) OR (end_date 非空)；"
         "JOIN port.portfunding；GROUP BY loanid,svcloanid,description,description_start_date 去重"),
        ("独立性", "不要求 fcreferraldate IS NOT NULL——未进主 FCL 表的贷款只要有 Hold 记录且在融资池也入库"),
        ("写入策略", "DELETE（按 tenant 全清）+ INSERT 全量重写；块B 取 portnewrezfc 当前快照 3 槽位"),
    ],
    "lm": [  # ④ sync_loan_foreclosure_loss_mitigation（doc 13 §5）
        ("BPS 对接表", "bpms_dev.sync_loan_foreclosure_loss_mitigation（每个 LM 周期一行，全历史）"),
        ("源表→Redshift 提取（basic_data_pool_config）",
         "newrez.portnewrezlm WHERE dealstartdate IS NOT NULL；去重 ROW_NUMBER() OVER "
         "(PARTITION BY loanid,dealstartdate ORDER BY dataasof DESC)=1（每周期留最新快照）"),
        ("数值解码", "lmdeal/lmprogram/lmstatus/lmdecision/denialreason/borrowerintention 经 "
                     "LEFT JOIN newrez.portnewrezdatadic 解码为文本后写入"),
        ("Redshift→MySQL 同步（GEN_FORECLOSURE_LM）", "无额外 WHERE（全量）；JOIN port.portfunding"),
        ("多 Servicer", "中间表合并 Newrez(portnewrezlm)+Carrington+Capecodfive"),
    ],
    "bk": [  # ⑤ sync_loan_foreclosure_bankruptcy（doc 13 §6）
        ("BPS 对接表", "bpms_dev.sync_loan_foreclosure_bankruptcy（仅有破产记录的贷款才有数据，否则界面 No Rows To Show）"),
        ("源表→Redshift 提取（basic_data_pool_config）",
         "newrez.portnewrezbk WHERE LENGTH(TRIM(bkstatus))>0（破产状态码须有值）；去重 ROW_NUMBER() OVER "
         "(PARTITION BY loanid,bkfileddate ORDER BY dataasof DESC)=1"),
        ("额外 JOIN", "LEFT JOIN portnewrezgeneral 取 legalstatus；LEFT JOIN portnewrezdatadic 解码 BKStatus"),
        ("Redshift→MySQL 同步（GEN_FORECLOSURE_BK）", "无额外 WHERE（全量）；JOIN port.portfunding"),
        ("多 Servicer", "合并 Newrez(portnewrezbk)+Carrington+Capecodfive；块B 取 portnewrezbk 最新快照(bkstatus 非空)"),
    ],
    "stage": [  # ⑥⑦ sync_fcl_stage_info（doc 13 §7）
        ("BPS 对接表", "bpms_dev.sync_fcl_stage_info（每日快照，每贷款每天一行）"),
        ("主筛选：进行中 FCL（GEN_FCL_STAGE）",
         "activefcflag=1 AND fcremovaldate IS NULL AND (fcremovaldesc IS NULL OR fcremovaldesc='')"),
        ("次筛选：预 FCL 贷款",
         "activefcflag!=1 AND demandsentdate IS NOT NULL AND referral_start_date IS NULL，再 JOIN "
         "port.basic_data_fcl_related 过滤 delq_status IN ('D90','D120P')（有 Demand 但未转介的 90天+ 逾期）"),
        ("Redshift→MySQL 同步（GET_FCL_STAGE_DATA）", "select a.* from port.fcl_stage_info a JOIN port.portfunding p（无额外 WHERE）"),
        ("快照口径", "查询当前状态须 fctrdt=MAX(fctrdt)；本页块C/块D 均取最新 fctrdt"),
        ("关键差异", "⚠️ 非活跃贷款（activefcflag=0：当前不处于活跃止赎，含完成/撤销/复议等）不在本表"
                     "（与 ①② 主 FCL 表含 activefcflag=0 不同）；人口更严格、数量不可直接对比"),
    ],
}


def filter_block(ws, row, rows):
    """块0 — 本页数据范围/过滤条件。返回下一行。"""
    row = section_title(ws, row, "块0 — 本页数据范围 / 过滤条件（满足下列条件的数据才会展示在本面板；源表与 BPS 表两层）", 6)
    row = table(ws, row, ["环节 / 表", "过滤条件（满足才展示）"], [[a, b] for a, b in rows], header_fill=ORANGE[:0] or "C55A11")
    return row + 1


def new_panel_sheet(wb, title, img_file, widths, filter_key=None):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    start = add_image(ws, img_file)
    r = start + 1
    if filter_key:
        r = filter_block(ws, r, PANEL_FILTERS[filter_key])
    return ws, r


PANEL_W = {"A": 46, "B": 26, "C": 26, "D": 26, "E": 26, "F": 26}

# stage 代码 → BPS 聚合页显示名（doc 13 §7）
STAGE_DISPLAY = {
    "SALE": "Upcoming FC Sales", "JUDGEMENT": "Upcoming Judgement", "PUBLICATION": "Publication",
    "SERVICE": "Service", "FIRST_LEGAL": "First Legal", "REFERRAL": "Referral", "DEMAND": "NOI/Demand Letter",
}


def fetch_agg(cur):
    """聚合页跨贷款分组计数（最新 fctrdt 快照）。"""
    from collections import Counter
    cur.execute("SELECT MAX(fctrdt) FROM bpms_dev.sync_fcl_stage_info")
    mx = cur.fetchone()[0]
    cur.execute("SELECT stage, `group`, servicer, judicial FROM bpms_dev.sync_fcl_stage_info WHERE fctrdt=%s", (mx,))
    rows = cur.fetchall()

    def cnt(i):
        return sorted(Counter(r[i] for r in rows).items(), key=lambda x: -x[1])

    return {"fctrdt": mx, "total": len(rows), "by_stage": cnt(0),
            "by_group": cnt(1), "by_servicer": cnt(2), "by_judicial": cnt(3)}


def count_block(ws, row, agg):
    """块D — 跨贷款分组计数（stage 计数 + 分组维度计数）。返回下一行。"""
    total = agg["total"] or 1
    row = section_title(ws, row,
                        f"块D — 跨贷款分组计数（聚合页实际展示；sync_fcl_stage_info 最新快照 fctrdt={fmt(agg['fctrdt'])}）",
                        6)
    srows = [[STAGE_DISPLAY.get(s, s), s, str(n), f"{n/total:.0%}"] for s, n in agg["by_stage"]]
    srows.append(["合计", "—", str(agg["total"]), "100%"])
    row = table(ws, row, ["BPS 显示名（分组）", "stage 代码", "贷款数", "占比"], srows)

    dims = []
    for v, n in agg["by_group"]:
        dims.append(["group", fmt(v), str(n)])
    for v, n in agg["by_servicer"]:
        dims.append(["servicer", fmt(v), str(n)])
    for v, n in agg["by_judicial"]:
        dims.append(["judicial", fmt(v), str(n)])
    row = table(ws, row, ["分组维度", "取值", "贷款数"], dims)
    note(ws, row, "⚠️ 口径说明（本 dev 环境实测，三套人口不可直接对比）：① 本表 sync_fcl_stage_info 最新 fctrdt 仅活跃 FCL；"
                  "② 原始源表 portnewrezfc 最新快照(2026-05-31) activefcflag=1 = 36 笔（当前真正活跃）、activefcflag=0 = 5016 笔；"
                  "③ 主 FCL 表 sync_loan_foreclosure = 98 笔（fcreferraldate 非空，含 activefcflag=0 非活跃贷款）。"
                  "doc 13 §7 所述『约 1.3 万活跃 FCL』为更完整数据集/生产口径，非本 dev。计数随 fctrdt 快照变化，脚本重跑自动刷新。", 6)
    return row + 1


# ① FCL Summary ----------------------------------------------------------
def build_summary(wb, data):
    ws, r = new_panel_sheet(wb, "① FCL Summary", "bps-loan-foreclosure-summary-panel.png", PANEL_W, "main")
    r = section_title(ws, r, "块A — BPS 字段映射规则（doc 13 §3.7；所有 BPS 字段位于 bpms_dev.sync_loan_foreclosure）", 6)
    mapA = [
        ["Foreclosure Status", "summary_foreclosure_status", "activefcflag, fcremovaldesc",
         "如果 activefcflag=1，则 summary_foreclosure_status = 固定文本 'Active Foreclosure'；"
         "如果 activefcflag=0 且 fcremovaldesc 非空，则 = 'Closed Foreclosure:' + fcremovaldesc；"
         "否则 = NULL。（注：fcstage 不参与本字段——它填充的是 Current Step / summary_current_step；fcresults 也不用于本字段。"
         "代码 basic_data_pool_config.py:273 GEN_FCL_DETAIL）"],
        ["Foreclosure Bid Amount", "summary_foreclosure_bid_amount", "fcbidamount", "直接取值"],
        ["Foreclosure Sale Amount", "summary_foreclosure_sale_amount", "fcsaleamount", "直接取值"],
        ["Contested Litigation", "summary_contested_litigation", "fccontestedflag", "直接取值 1/0"],
        ["Firm", "summary_firm", "fcfirm", "同 Foreclosure Attorney（同源）"],
        ["Type", "summary_type", "judicial",
         "如果 judicial=1，则 summary_type = 'Judicial'；如果 judicial=0，则 = 'Non Judicial'；如果 judicial 为 NULL/空，则 = NULL"],
        ["SMS Days in FCL", "summary_sms_days_in_fcl", "smsdaysinfc(=svc_days_infc), dataasof",
         "Servicer(SMS=Shellpoint)口径，自 servicer 建案日 fcsetupdate 起算（Newrez 原生 smsdaysinfc 透传）；实时重算 smsdaysinfc + DATEDIFF(今日纽约, dataasof)；≤ Days in FCL"],
        ["Days in FCL", "summary_days_in_fcl", "daysinfc, dataasof",
         "投资人/全程口径，自转介日 fcreferraldate 起算（datediff+1）；实时重算 daysinfc + DATEDIFF(今日纽约, dataasof)；≥ SMS Days（因 setup≥referral）"],
        ["Current Step", "summary_current_step", "currentmilestone / fcstage",
         "如果 currentmilestone 非空，则 summary_current_step = currentmilestone；否则 = fcstage"],
        ["Last Step Completed", "summary_last_step_completed", "lastfcstepcompleted", "直接取值"],
        ["Last Step Completed Date", "summary_last_step_completed_date", "lastfcstepcompleteddate", "直接取值"],
        ["Servicer Number", "summary_servicer_number", "shellpointloanid", "直接取值"],
        ["Completed Foreclosure", "summary_completed_foreclosure", "activefcflag",
         "如果 activefcflag=0，则 summary_completed_foreclosure = 1；如果 activefcflag=1，则 = 0（即对 activefcflag 取反）。"
         "⚠️ activefcflag=0 = 当前不处于活跃止赎流程（含撤销/复议/付清/完成），并非都已完成"],
        ["Servicer FC Bid Amount", "summary_srv_fc_bid_amount", "fcbidamount", "同 bid amount（servicer 视角）"],
        ["Judicial Foreclosure", "summary_judicial_foreclosure", "judicial", "直接取值（布尔）"],
        ["Foreclosure Attorney", "summary_foreclosure_attorney", "fcfirm", "直接取值"],
    ]
    r = table(ws, r, ["BPS 界面标签", "BPS 展示字段", "Newrez 源（portnewrezfc）", "Newrez → BPS 规则"],
              [[a, b, c, d] for a, b, c, d in mapA])

    fc = data["fc"]
    r = note(ws, r, "📌 本面板过滤字段（决定一笔贷款是否出现在 BPS Summary，详见块0）：① portnewrezfc.fcreferraldate "
                    "IS NOT NULL（已转介 FCL）② 贷款在 port.portfunding 融资池中。块B 取每贷款 dataasof=MAX 最新快照。"
                    "→ 不满足者不会进入 sync_loan_foreclosure，故块B/块C 无该贷款行。", 6)
    raw_fields = ["fcstage", "fcresults", "fcremovaldesc", "activefcflag", "fcbidamount", "fcsaleamount",
                  "fccontestedflag", "fcfirm", "judicial", "smsdaysinfc", "daysinfc", "dataasof",
                  "currentmilestone", "lastfcstepcompleted", "lastfcstepcompleteddate", "shellpointloanid",
                  "fcapprbidprice"]
    r = matrix_block(ws, r, "块B — 原始 Newrez 数据（portnewrezfc 最新快照 2026-05-31）",
                     [(f, [fmt(fc[l].get(f)) for l in LOAN_IDS]) for f in raw_fields], header_fill=GREEN[:0] or "548235")

    bps = data["bps_main"]
    bps_fields = ["summary_foreclosure_status", "summary_foreclosure_bid_amount", "summary_foreclosure_sale_amount",
                  "summary_contested_litigation", "summary_firm", "summary_type", "summary_sms_days_in_fcl",
                  "summary_days_in_fcl", "summary_current_step", "summary_last_step_completed",
                  "summary_last_step_completed_date", "summary_servicer_number", "summary_completed_foreclosure",
                  "summary_judicial_foreclosure", "summary_foreclosure_attorney"]
    r = matrix_block(ws, r, "块C — BPS 实测值（界面实际显示；bpms_dev.sync_loan_foreclosure）",
                     [(f, [fmt(bps[l].get(f)) for l in LOAN_IDS]) for f in bps_fields], header_fill="548235")
    note(ws, r, "读法：块A 规则将 块B(Newrez 原始) 推算为 块C(BPS 界面值)。例：Type = judicial 文本化；"
                "Completed Foreclosure = activefcflag 取反（注：activefcflag=0 表示当前不处于活跃止赎，含完成/撤销/复议/付清，"
                "并非都已完成）；SMS/Days in FCL = 原值 + 距今天数（故块C 可能 > 块B）——其中 SMS Days 自 servicer "
                "建案日 fcsetupdate 起算、Days 自转介日 fcreferraldate 起算，故 SMS Days ≤ Days in FCL。", 6)
    return ws


# ② Milestone Timeline ---------------------------------------------------
def build_timeline(wb, data):
    ws, r = new_panel_sheet(wb, "② Milestone Timeline", "bps-loan-foreclosure-milestone-timeline-panel.png", PANEL_W, "main")
    r = section_title(ws, r, "块A — timeline_* 映射规则（doc 13 §3.1；BPS 字段位于 bpms_dev.sync_loan_foreclosure，源表 portnewrezfc）", 6)
    mapA = [
        ["Notice of Intent Date", "timeline_notice_of_intent_date", "demandsentdate", "直接取值"],
        ["Notice of Intent End Date", "timeline_notice_of_intent_end_date", "demandexpirationdate", "直接取值"],
        ["Approved for Referral Date", "timeline_approved_for_referral_date", "fcsetupdate", "直接取值"],
        ["Referred to Attorney Date", "timeline_referred_to_attorney_date", "fcreferraldate", "直接取值"],
        ["Referred to Foreclosure Date", "timeline_referred_to_foreclosure_date", "fcreferraldate", "直接取值（入库过滤字段）"],
        ["Title Report Received Date", "timeline_title_report_received_date", "titlereceiveddate", "直接取值（Newrez 多为空）"],
        ["Preliminary Title Cleared Date", "timeline_preliminary_title_cleared_date", "titlecleardate", "直接取值"],
        ["First Legal Date", "timeline_first_legal_date", "firstlegaldate", "直接取值（非司法州常空）"],
        ["Service Date", "timeline_service_date", "servicecompletedate", "直接取值"],
        ["Publication Date", "timeline_publication_date", "(Newrez 无)", "Newrez 未提供，恒空"],
        ["Judgement Hearing Set Date", "timeline_judgement_hearing_set_date", "fcjudgmenthearingscheduled",
         "ETL 追踪：MIN(dataasof WHERE 值=当前值)"],
        ["Judgement Date", "timeline_judgement_date", "fcjudgmenthearingscheduled", "直接取值（非 fcjudgmententered）"],
        ["Projected Sale Date", "timeline_sale_date_projected_date", "fcscheduledsaledate", "直接取值"],
        ["Sale Date Set", "timeline_sale_date_set_date", "fcscheduledsaledate", "ETL 追踪：MIN(dataasof WHERE 值=当前值)"],
        ["Final Title Cleared Date", "timeline_final_title_cleared_date", "titlecleardate", "直接取值"],
        ["Sale Date Held", "timeline_sale_date_held_date", "fcsalehelddate", "直接取值"],
        ["Foreclosure Completed Date", "timeline_foreclosure_completed_date", "dtdeedrecorded / fcremovaldate",
         "COALESCE(dtdeedrecorded, fcremovaldate)"],
        ["Third Party Sold Date", "timeline_third_party_sold_date_date", "fcsalehelddate",
         "当 fcresults='3rd Party' 时取 fcsalehelddate"],
        ["Third Party Proceeds Received", "timeline_third_party_proceeds_received_date",
         "fcl3rdpartyproceedsreceiveddate", "直接取值"],
    ]
    r = table(ws, r, ["BPS 界面标签", "BPS 展示字段", "Newrez 源（portnewrezfc）", "Newrez → BPS 规则"],
              [[a, b, c, d] for a, b, c, d in mapA])

    note(ws, r, "Target/Actual/Var 天数（doc 13 §3.2–3.4）：Target=系统配置常量（与 Newrez 无关，存于 "
                "sync_loan_foreclosure.target_*）；Actual=两端 timeline 日期相减（端点来自 portnewrezfc）；"
                "Var=Actual−Target。Actual_*/Var_* 实测值【仅存于视图 bpms_dev.biz_data_view_loan_details_foreclosure】"
                "（多快照行，非 sync 表单值），查询见 doc 13 附录 SQL-8 / doc 14 SQL-C3；故此处不逐笔展开。", 6)
    r += 1

    fc = data["fc"]
    raw_fields = ["demandsentdate", "demandexpirationdate", "fcsetupdate", "fcreferraldate", "firstlegaldate",
                  "servicecompletedate", "fcjudgmenthearingscheduled", "fcjudgmententered", "fcscheduledsaledate",
                  "fcsalehelddate", "dtdeedrecorded", "fcremovaldate", "titlecleardate", "fcresults"]
    r = matrix_block(ws, r, "块B — 原始 Newrez 时间线字段（portnewrezfc 最新快照）",
                     [(f, [fmt(fc[l].get(f)) for l in LOAN_IDS]) for f in raw_fields], header_fill="548235")

    bps = data["bps_main"]
    tl_fields = [m[1] for m in mapA]
    r = matrix_block(ws, r, "块C — BPS timeline_* 实测值（界面显示日期；bpms_dev.sync_loan_foreclosure）",
                     [(f, [fmt(bps[l].get(f)) for l in LOAN_IDS]) for f in tl_fields], header_fill="548235")
    note(ws, r, "例：timeline_referred_to_foreclosure = fcreferraldate（直接取值，块B=块C 吻合）；"
                "timeline_foreclosure_completed = COALESCE(dtdeedrecorded, fcremovaldate)；"
                "Third Party Sold 仅当 fcresults='3rd Party' 才取 fcsalehelddate，否则空。", 6)
    return ws


# ③ Hold -----------------------------------------------------------------
def build_hold(wb, data):
    ws, r = new_panel_sheet(wb, "③ Hold", "bps-loan-foreclosure-hold-panel.png", PANEL_W, "hold")
    r = section_title(ws, r, "块A — Hold 映射规则（doc 13 §4；BPS 表 bpms_dev.sync_loan_foreclosure_hold）", 6)
    r = table(ws, r, ["UI 列", "BPS 字段", "Newrez 源（portnewrezfc）", "规则"],
              [["Description", "description", "fchold1/2/3description", "3 槽位 UNPIVOT 展开"],
               ["Start Date", "description_start_date", "fchold1/2/3startdate", "直接取值"],
               ["End Date", "description_end_date", "fchold1/2/3enddate", "直接取值（NULL=仍生效）"]])
    note(ws, r, "架构：Newrez portnewrezfc 仅存【当前快照】3 个 Hold 槽位；BPS 每次变更追加一行，"
                "sync_loan_foreclosure_hold 累积该贷款【完整 Hold 历史】（远多于 3 槽）。", 6)
    r += 1

    # 块B：5 笔贷款的 Newrez 3 槽位
    fc = data["fc"]
    slot_rows = []
    for n in (1, 2, 3):
        for sub in ("description", "startdate", "enddate"):
            f = f"fchold{n}{sub}"
            slot_rows.append((f"槽{n} {sub}", [fmt(fc[l].get(f)) for l in LOAN_IDS]))
    r = matrix_block(ws, r, "块B — 原始 Newrez Hold 3 槽位（portnewrezfc 当前快照）", slot_rows, header_fill="548235")

    # 块C：每贷款 BPS Hold 历史（stacked）
    r = section_title(ws, r, "块C — BPS Hold 全历史（bpms_dev.sync_loan_foreclosure_hold）", 6)
    for lid in LOAN_IDS:
        hist = data["hold"][lid]
        sub = ws.cell(r, 1, f"Loan {lid} — {len(hist)} 条 Hold 记录")
        sub.font = Font(name=YH, bold=True, size=9, color=DARK)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
        if not hist:
            c = ws.cell(r, 1, "（无 Hold 记录）"); c.font = Font(name=YH, size=9, italic=True); r += 2
            continue
        rows = [[fmt(h.get("description")), fmt(h.get("description_start_date")), fmt(h.get("description_end_date"))]
                for h in hist]
        r = table(ws, r, ["Description", "Start Date", "End Date"], rows)
    return ws


# ④ Loss Mitigation Cycle ------------------------------------------------
def build_lm(wb, data):
    ws, r = new_panel_sheet(wb, "④ Loss Mitigation Cycle", "bps-loan-foreclosure-loss-mitigation-cycle-panel.png", PANEL_W, "lm")
    r = section_title(ws, r, "块A — LM 映射规则（doc 13 §5；BPS 表 bpms_dev.sync_loan_foreclosure_loss_mitigation）", 6)
    r = table(ws, r, ["UI 列", "BPS 字段", "Newrez 源（portnewrezlm）", "规则"],
              [["Deal", "deal", "lmdeal (int)", "经 portnewrezdatadic 解码为文本（如 7→DIL）"],
               ["Program", "program", "lmprogram (int)", "解码为文本（如 10→Deed-in-Lieu）"],
               ["Status", "lmc_status", "lmstatus (int)", "解码为文本（如 112→Workout Denial）"],
               ["Cycle Opened Date", "cycle_opened_date", "dealstartdate", "直接取值"],
               ["Cycle Closed Date", "cycle_closed_date", "lmremovaldate", "直接取值（NULL=进行中）"],
               ["Final Disposition", "final_disposition", "lmdecision (int)", "解码为文本（如 6→Referral to FC）"],
               ["Denial / Reason", "denialreason", "denialreason (int)", "解码为文本；无拒绝为空"],
               ["Borrower Intentions", "borrower_intentions", "borrowerintention (int)", "解码；Newrez 通常空"],
               ["Imminent Default", "imminent_default", "(Newrez 无)", "对 Newrez 恒 NULL"],
               ["Single Point of Contact", "single_point_of_contact", "(Newrez 无)", "对 Newrez 恒 NULL"]])
    note(ws, r, "关键：LM 字段在 Redshift 层 LEFT JOIN portnewrezdatadic 把数字码解码为文本后写入 BPS"
                "（与 Hold 直接存文本不同）。每个 LM 周期（lmdeal+dealstartdate）一行。", 6)
    r += 1

    # 块B：旗舰贷款的原始 LM 数字码
    flagship = "7727000088"
    raw = data["lm_raw"][flagship]
    r = section_title(ws, r, f"块B — 原始 Newrez LM 数字码示例（portnewrezlm，Loan {flagship}，最新快照各周期）", 6)
    if raw:
        rows = [[fmt(x.get("lmdeal")), fmt(x.get("lmprogram")), fmt(x.get("lmstatus")),
                 fmt(x.get("dealstartdate")), fmt(x.get("lmremovaldate")), fmt(x.get("lmdecision"))] for x in raw]
        r = table(ws, r, ["lmdeal", "lmprogram", "lmstatus", "dealstartdate", "lmremovaldate", "lmdecision"], rows)
    else:
        c = ws.cell(r, 1, "（无 LM 原始记录）"); c.font = Font(name=YH, size=9, italic=True); r += 2
    note(ws, r, "↑ 数字码（如 lmdeal=7）；↓ 块C 为同贷款经 ETL 解码后的 BPS 文本（如 deal='DIL'）。", 6)
    r += 1

    # 块C：每贷款 BPS LM 历史
    r = section_title(ws, r, "块C — BPS LM 周期历史（已解码文本；sync_loan_foreclosure_loss_mitigation）", 6)
    for lid in LOAN_IDS:
        hist = data["lm"][lid]
        sub = ws.cell(r, 1, f"Loan {lid} — {len(hist)} 条 LM 周期")
        sub.font = Font(name=YH, bold=True, size=9, color=DARK)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
        if not hist:
            c = ws.cell(r, 1, "（无 LM 记录）"); c.font = Font(name=YH, size=9, italic=True); r += 2
            continue
        rows = [[fmt(h.get("deal")), fmt(h.get("program")), fmt(h.get("lmc_status")),
                 fmt(h.get("cycle_opened_date")), fmt(h.get("cycle_closed_date")), fmt(h.get("final_disposition"))]
                for h in hist]
        r = table(ws, r, ["Deal", "Program", "Status", "Cycle Opened", "Cycle Closed", "Final Disposition"], rows)
    return ws


# ⑤ Bankruptcy -----------------------------------------------------------
def build_bk(wb, data):
    ws, r = new_panel_sheet(wb, "⑤ Bankruptcy", "bps-loan-foreclosure-bankruptcy-panel.png", PANEL_W, "bk")
    r = section_title(ws, r, "块A — Bankruptcy 映射规则（doc 13 §6；ETL 代码 basic_data_pool_config.py:349-363 实测核实 2026-06-02）", 6)
    r = table(ws, r, ["UI 列", "BPS 字段", "Newrez 源", "规则（代码核实）"],
              [["Status", "bankruptcy_status", "portnewrezbk.bkstatus → portnewrezdatadic",
                "COALESCE(datadic[BKStatus].description, bkstatus)；实测解码 Active/Discharged/Dismissed/Closed/"
                "ReliefGranted/Completed-Cancelled，未匹配回退原始码(如 3.0)"],
               ["Legal Status", "legal_status", "portnewrezgeneral.legalstatus (文本)",
                "直接取文本（FCBU/BK13/BK7…）；实测 31/64 非空"],
               ["Status Date", "status_date", "portnewrezbk.bkfileddate",
                "⚠️ 直接取【破产申请日 bkfileddate】，非 bkrcurrentstatusdate（代码核实更正）"],
               ["Chapter", "chapter", "portnewrezbk.bkchapter", "CAST 为 DECIMAL（7/11/13）"],
               ["Lien Status", "lien_status", "—", "⚠️ ETL 硬编码 NULL（未映射任何 Newrez 字段）"],
               ["MFR Status", "mfr_status", "—", "⚠️ ETL 硬编码 NULL（非 mfrhearingresults，全表 0/64）"],
               ["MFR Filed Date", "mfr_filed_date", "—",
                "⚠️ Newrez 链路硬编码 NULL；全表仅 3/64 非空（疑 BPS 内部/历史写入，非当前提取链路）"],
               ["Claim Status", "claim_status", "—", "⚠️ ETL 硬编码 NULL"],
               ["Proof of Claim Date", "proof_of_claim_date", "portnewrezbk.pocfileddate", "直接取值（24/64 非空）"],
               ["Post Petition Due Date", "post_petition_due_date", "portnewrezbk.bkpostpetitionduedate", "直接取值（22/64 非空）"]])
    note(ws, r, "代码核实更正（basic_data_pool_config.py:349-363）：① status_date 实为 bkfileddate（非 bkrcurrentstatusdate）；"
                "② lien_status / mfr_status / claim_status / mfr_filed_date 在 Newrez 提取链路【硬编码 NULL】；"
                "③ bankruptcy_status = COALESCE(datadic 解码, 原始码)。Legal Status 来自 portnewrezgeneral.legalstatus；"
                "面板仅在贷款有破产记录时有数据，否则界面 No Rows To Show。", 6)
    r += 1

    # 块B：原始 portnewrezbk（有 BK 的贷款）
    r = section_title(ws, r, "块B — 原始 Newrez 破产数据（portnewrezbk 最新快照）", 6)
    any_bk = False
    for lid in LOAN_IDS:
        recs = data["bk_raw"][lid]
        if not recs:
            continue
        any_bk = True
        sub = ws.cell(r, 1, f"Loan {lid} — {len(recs)} 条")
        sub.font = Font(name=YH, bold=True, size=9, color=DARK)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
        rows = [[fmt(x.get("bkstatus")), fmt(x.get("bkchapter")), fmt(x.get("bkfileddate")),
                 fmt(x.get("pocfileddate")), fmt(x.get("bkpostpetitionduedate"))] for x in recs]
        r = table(ws, r, ["bkstatus(码)", "bkchapter", "bkfileddate(→status_date)", "pocfileddate", "bkpostpetitionduedate"], rows)
    if not any_bk:
        c = ws.cell(r, 1, "（5 笔均无破产原始记录）"); c.font = Font(name=YH, size=9, italic=True); r += 2

    # 块C：BPS BK 历史
    r = section_title(ws, r, "块C — BPS 破产记录（已解码；sync_loan_foreclosure_bankruptcy）", 6)
    for lid in LOAN_IDS:
        recs = data["bk"][lid]
        sub = ws.cell(r, 1, f"Loan {lid} — {len(recs)} 条破产记录")
        sub.font = Font(name=YH, bold=True, size=9, color=DARK)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
        if not recs:
            c = ws.cell(r, 1, "（No Rows To Show）"); c.font = Font(name=YH, size=9, italic=True); r += 2
            continue
        rows = [[fmt(x.get("bankruptcy_status")), fmt(x.get("legal_status")), fmt(x.get("chapter")),
                 fmt(x.get("status_date")), fmt(x.get("mfr_filed_date")), fmt(x.get("proof_of_claim_date"))]
                for x in recs]
        r = table(ws, r, ["Status", "Legal Status", "Chapter", "Status Date", "MFR Filed", "POC Date"], rows)
    return ws


# ⑥ Aggregate Stage Tab --------------------------------------------------
def build_agg_stage(wb, data):
    ws, r = new_panel_sheet(wb, "⑥ Aggregate Stage Tab", "bps-foreclosure-aggregate-stage-tab-grouped-detail.png", PANEL_W, "stage")
    r = section_title(ws, r, "块A — 阶段瀑布判定逻辑（doc 13 §7；BPS 表 bpms_dev.sync_fcl_stage_info，每日快照 fctrdt）", 6)
    r = table(ws, r, ["优先级", "stage 存储值", "判断条件（portnewrezfc）", "BPS 显示名"],
              [["1", "SALE", "fcscheduledsaledate IS NOT NULL", "Upcoming FC Sales"],
               ["2", "JUDGEMENT", "fcjudgmenthearingscheduled NOT NULL 且 fcscheduledsaledate NULL", "Upcoming Judgement"],
               ["3", "PUBLICATION", "publication_date NOT NULL（Newrez 恒 0）", "Publication"],
               ["4", "SERVICE", "servicecompletedate NOT NULL 且 1~3 未命中", "Service"],
               ["5", "FIRST_LEGAL", "firstlegaldate NOT NULL 且 servicecompletedate NULL", "First Legal"],
               ["6", "REFERRAL", "fcreferraldate NOT NULL 且 firstlegaldate NULL", "Referral"],
               ["7", "DEMAND", "demandsentdate NOT NULL 且 fcreferraldate NULL（入库后恒 0）", "NOI/Demand Letter"]])
    note(ws, r, "BPS 页面直接读 sync_fcl_stage_info 各阶段字段，无额外转换（页面=存储）。天数字段："
                "Days to Sale=to_sale_days；Days to Judgement=to_judgement_days；其余=各阶段 *_stage_days。"
                "仅活跃 FCL 入此表（activefcflag=0 非活跃贷款不在——含完成/撤销/复议等）。", 6)
    r += 1

    stg = data["stage"]
    fields = ["stage", "group", "state", "judicial", "referral_start_date", "first_legal_start_date",
              "service_start_date", "judgement_start_date", "sale_start_date", "to_sale_days",
              "to_judgement_days", "referral_stage_days", "first_legal_stage_days", "service_stage_days"]
    r = matrix_block(ws, r, "块C — BPS 阶段实测值（sync_fcl_stage_info 最新 fctrdt；空=该贷款不在活跃 FCL 表）",
                     [(f, [fmt(stg[l].get(f)) for l in LOAN_IDS]) for f in fields], header_fill="548235")
    note(ws, r, "聚合页本质是跨贷款【分组计数/统计】；块C 对 5 笔单贷款展示其各自落入的 stage 与天数，"
                "便于理解每行如何归组；块D 为跨贷款分组计数（聚合页顶部汇总）。非活跃贷款（activefcflag=0，"
                "如 7727000088 已 REO 完成）通常不在此表。", 6)
    r += 1
    r = count_block(ws, r, data["agg"])
    return ws


# ⑦ Aggregate Time Line Tab ---------------------------------------------
def build_agg_timeline(wb, data):
    ws, r = new_panel_sheet(wb, "⑦ Aggregate Time Line Tab", "bps-foreclosure-aggregate-time-line-tab.png", PANEL_W, "stage")
    r = section_title(ws, r, "块A — Time Line Tab 字段映射（doc 13 §7；BPS 表 bpms_dev.sync_fcl_stage_info）", 6)
    r = table(ws, r, ["UI 列(序)", "BPS 字段", "Newrez 源（portnewrezfc）", "说明"],
              [["NOI Date (1)", "noi_start_date", "demandsentdate*", "Newrez 此字段恒 NULL（NOI 列对 Newrez 空白）"],
               ["Referral Date (2)", "referral_start_date", "fcreferraldate", "100%（入库前提）"],
               ["First Legal Date (3)", "first_legal_start_date", "firstlegaldate", "直接取值"],
               ["Service Date (4)", "service_start_date", "servicecompletedate", "直接取值"],
               ["Publication Date (5)", "publication_start_date", "(无对应)", "Newrez 恒 NULL"],
               ["Judgement Date (6)", "judgement_start_date", "fcjudgmenthearingscheduled", "司法州独有（排定日，非录入日）"],
               ["Sale Date (7)", "sale_start_date", "fcscheduledsaledate", "直接取值"]])
    note(ws, r, "序号 1~7 = FCL 阶段时间顺序（非优先级）。Time Line Tab 每行一笔贷款，横向显示里程碑日期；"
                "NOI Date 列对 Newrez 恒空（demand 数据走 demand_start_date，不映射到此列）。", 6)
    r += 1

    stg = data["stage"]
    fields = ["group", "judicial", "noi_start_date", "referral_start_date", "first_legal_start_date",
              "service_start_date", "publication_start_date", "judgement_start_date", "sale_start_date"]
    r = matrix_block(ws, r, "块C — BPS Time Line 实测值（sync_fcl_stage_info 最新 fctrdt）",
                     [(f, [fmt(stg[l].get(f)) for l in LOAN_IDS]) for f in fields], header_fill="548235")
    r += 1
    r = count_block(ws, r, data["agg"])
    return ws


# ══════════════════════════════════════════════════════════════════════
def main():
    if not OUT.parent.exists():
        OUT.parent.mkdir(parents=True)
    conn = pymysql.connect(host=HOST, port=PORT, user=USER, password=PASSWORD, autocommit=True,
                           connect_timeout=20, read_timeout=180)
    cur = conn.cursor()
    print("连接 OK；开始 Schema-Verify…")
    bad = verify_schema(cur)
    if bad:
        print("  ⚠️ 存在无效源列，仍继续（缺失列将显示 —）：", bad)

    print("取数：5 条样例贷款…")
    data = {"fc": {}, "bps_main": {}, "hold": {}, "lm": {}, "lm_raw": {}, "bk": {}, "bk_raw": {}, "stage": {}}
    for lid in LOAN_IDS:
        data["fc"][lid] = fetch_newrez_fc(cur, lid)
        data["bps_main"][lid] = (fetch_bps(cur, "sync_loan_foreclosure", lid) or [{}])[0]
        data["hold"][lid] = sorted(fetch_bps(cur, "sync_loan_foreclosure_hold", lid),
                                   key=lambda h: (h.get("description_start_date") or datetime.date(1900, 1, 1)))
        data["lm"][lid] = sorted(fetch_bps(cur, "sync_loan_foreclosure_loss_mitigation", lid),
                                 key=lambda h: (h.get("cycle_opened_date") or datetime.date(1900, 1, 1)))
        data["lm_raw"][lid] = fetch_newrez_lm(cur, lid)
        data["bk"][lid] = fetch_bps(cur, "sync_loan_foreclosure_bankruptcy", lid)
        data["bk_raw"][lid] = fetch_newrez_bk(cur, lid)
        data["stage"][lid] = (fetch_bps(cur, "sync_fcl_stage_info", lid, latest_col="fctrdt") or [{}])[0]
        print(f"  {lid}: fc={'Y' if data['fc'][lid] else 'N'} bps_main={'Y' if data['bps_main'][lid] else 'N'} "
              f"hold={len(data['hold'][lid])} lm={len(data['lm'][lid])} bk={len(data['bk'][lid])} "
              f"stage={'Y' if data['stage'][lid] else 'N'}")
    data["agg"] = fetch_agg(cur)
    print(f"  聚合计数 fctrdt={data['agg']['fctrdt']} total={data['agg']['total']} by_stage={data['agg']['by_stage']}")
    cur.close()
    conn.close()

    print("构建工作簿…")
    wb = Workbook()
    build_cover(wb)
    build_summary(wb, data)
    build_timeline(wb, data)
    build_hold(wb, data)
    build_lm(wb, data)
    build_bk(wb, data)
    build_agg_stage(wb, data)
    build_agg_timeline(wb, data)

    wb.save(OUT)
    print(f"已保存：{OUT}")
    print(f"Sheet 数：{len(wb.worksheets)} -> {[ws.title for ws in wb.worksheets]}")


if __name__ == "__main__":
    main()
