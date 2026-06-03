"""
Add 2 columns to the "📋 字段规范 Field Spec" sheet of 14_servicer_fcl_field_spec.xlsx:
  col9: 标准接口取值范围  — what BPS spec REQUIRES (format, allowed values)
  col10: 典型取值（标准） — one concrete valid standard example

Also upgrades content of the existing Newrez状态 column (col12 after shift) for key fields
with actual Newrez value ranges/decode notes, reusing data verified via MCP queries.

Run: python scripts/add_field_spec_value_columns.py
"""

import sys, io, copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")

# ── Value range + typical value mapping (keyed by col3 field name) ────────────

MAPPING = {
    # Section 2.0 — 四维状态基础标志
    "delinquency_status":       ("MBA标准文本枚举：Foreclosure · 120-149 DPD · 150+ DPD · D90（禁止数字串如'29.0'）", "Foreclosure"),
    "next_payment_due_date":    ("YYYY-MM-DD", "2024-03-01"),
    "days_past_due":            ("0~999 正整数（ETL衍生：days360(nextduedate, dataasof)）", "215"),
    "foreclosure_flag":         ("Y / N", "Y"),
    "lm_flag":                  ("Y / N", "Y"),
    "lm_type":                  ("Forbearance · Modification · Payment Plan · Short Sale · DIL", "Modification"),
    "lm_start_date":            ("YYYY-MM-DD", "2024-01-15"),
    "lm_end_date":              ("YYYY-MM-DD（空=进行中）", "2024-06-30"),
    "hold_flag":                ("Y / N", "Y"),
    "hold_reason":              ("BK · LM · HUD · Covid · Other", "BK"),
    "reo_flag":                 ("Y / N", "Y"),
    "reo_acquisition_date":     ("YYYY-MM-DD", "2024-05-01"),
    # Section 2.1 — 贷款识别与入库过滤
    "loan_id":                  ("纯数字字符串", "7727000088"),
    "servicer_loan_id":         ("Servicer内部编号（格式随Servicer）", "NR-2024-001234"),
    "data_as_of_date":          ("YYYY-MM-DD", "2026-05-27"),
    "state":                    ("US 2字母州代码（大写）", "FL"),
    "judicial_flag":            ("1（Judicial州）/ 0（Non-Judicial州）", "1"),
    "active_fcl_flag":          ("1（进行中）/ 0（已完结）；NULL视为1", "1"),
    "fcl_referral_date":        ("YYYY-MM-DD", "2024-01-20"),
    # Section 2.2 — FCL状态字段
    "fcl_stage":                ("自由文本", "Service Complete"),
    "current_milestone":        ("Closed · First Legal · Judgment Entered · Sale Held · Sold · Service Complete · Sale Scheduled", "First Legal"),
    "last_completed_step":      ("自由文本", "FC Referral"),
    "last_completed_step_date": ("YYYY-MM-DD", "2024-02-15"),
    "fcl_results":              ("REO · 3rd Party（活跃FCL为空）", "REO"),
    "attorney_firm":            ("自由文本（律师事务所名称）", "Kelley Kronenberg, P.A."),
    "contested_flag":           ("0 / 1", "0"),
    "servicer_days_in_fcl":     ("正整数", "128"),
    "days_in_fcl":              ("正整数", "215"),
    # Section 2.3 — FCL时间轴
    "noi_date":                       ("YYYY-MM-DD", "—"),
    "demand_sent_date":               ("YYYY-MM-DD", "2024-02-01"),
    "demand_expiration_date":         ("YYYY-MM-DD", "2024-03-02"),
    "fcl_setup_date":                 ("YYYY-MM-DD", "2024-01-10"),
    "first_legal_date":               ("YYYY-MM-DD", "2024-02-01"),
    "service_complete_date":          ("YYYY-MM-DD", "2024-03-10"),
    "publication_date":               ("YYYY-MM-DD", "—"),
    "title_received_date":            ("YYYY-MM-DD", "2025-04-01"),
    "title_clear_date":               ("YYYY-MM-DD", "2025-04-15"),
    "judgement_hearing_scheduled":    ("YYYY-MM-DD（Judicial州）", "2026-01-18"),
    "judgement_entered_date":         ("YYYY-MM-DD（Judicial州）", "2026-01-07"),
    "scheduled_sale_date":            ("YYYY-MM-DD", "2024-05-15"),
    "sale_held_date":                 ("YYYY-MM-DD", "2024-05-15"),
    "deed_recorded_date":             ("YYYY-MM-DD", "2024-06-01"),
    "fcl_removal_date":               ("YYYY-MM-DD", "2024-06-01"),
    "third_party_proceeds_date":      ("YYYY-MM-DD", "2026-03-04"),
    # Section 2.4 — Hold槽位
    "hold_1_description":        ("BK · LM · Court Delay · Service Delay 等（Newrez扩展枚举可接受）", "Loss Mitigation Workout"),
    "hold_1_start_date":         ("YYYY-MM-DD", "2024-01-05"),
    "hold_1_end_date":           ("YYYY-MM-DD（空=Hold仍持续）", "2024-03-15"),
    "hold_1_projected_end_date": ("YYYY-MM-DD", "2024-03-20"),
    "hold_2_description":        ("BK · LM · Court Delay · Service Delay 等（Newrez扩展枚举可接受）", "Loss Mitigation Workout"),
    "hold_2_start_date":         ("YYYY-MM-DD", "2024-01-05"),
    "hold_2_end_date":           ("YYYY-MM-DD（空=Hold仍持续）", "2024-03-15"),
    "hold_2_projected_end_date": ("YYYY-MM-DD", "2024-03-20"),
    "hold_3_description":        ("BK · LM · Court Delay · Service Delay 等（Newrez扩展枚举可接受）", "Loss Mitigation Workout"),
    "hold_3_start_date":         ("YYYY-MM-DD", "2024-01-05"),
    "hold_3_end_date":           ("YYYY-MM-DD（空=Hold仍持续）", "2024-03-15"),
    "hold_3_projected_end_date": ("YYYY-MM-DD", "2024-03-20"),
    "hold_1_modified_date":      ("YYYY-MM-DD", "2024-02-20"),
    "hold_2_modified_date":      ("YYYY-MM-DD", "2024-02-20"),
    "hold_3_modified_date":      ("YYYY-MM-DD", "2024-02-20"),
    # Section 2.5 — 竞标与销售
    "bid_amount":         ("正数，单位 USD", "160000.00"),
    "approved_bid_price": ("正数，单位 USD", "162000.00"),
    "sale_amount":        ("正数，单位 USD", "170000.00"),
    # Section 2.6 — 贷款属性增强
    "investor_loan_id":            ("投资人贷款编号", "INV-20240110-088"),
    "lien_position":               ("1（First Lien）/ 2（Second Lien）", "1"),
    "interest_paid_through_date":  ("YYYY-MM-DD", "2024-01-31"),
    "in_auction_flag":             ("0 / 1", "0"),
    "borrower_deceased_flag":      ("0 / 1", "0"),
    "reason_for_default":          ("自由文本（借款人违约原因）", "Unable to Contact Borrower"),
    "hold_1_comment":              ("自由文本", "Awaiting Bankruptcy Resolution"),
    "hold_2_comment":              ("自由文本", "Awaiting Court Scheduling"),
    "hold_3_comment":              ("自由文本", "—"),
    # Section 3.1 — LM周期
    "lm_deal":                ("Evaluation · Modification · Forbearance · Payment Plan · Deferment · Short Sale · DIL · Payoff", "Modification"),
    "lm_program":             ("Evaluation · Bridger mod · SLS Standard Mod · VA Traditional · Deferment · Repayment Plan 等20+种", "Bridger mod"),
    "lm_status":              ("Pending Financials · Workout Denial · Document Follow-up · Book mod · Monitor Forbearance 等20+种", "Pending Financials"),
    "lm_cycle_open_date":     ("YYYY-MM-DD", "2024-01-15"),
    "lm_cycle_close_date":    ("YYYY-MM-DD（空=进行中）", "2024-03-15"),
    "lm_final_disposition":   ("Pending · Modification Complete · Referral to FC · Request Incomplete · LMS Opened in Error · Forbearance Complete · Reinstated/Current · Deferment Completed · Full Pay Off · DIL Complete 等", "Referral to FC"),
    "lm_denial_reason":       ("Loan not due 3+ months · Withdrawal · Ineligible Borrower · Request Incomplete · Failed Plan 等18种", "Withdrawal of Request/Non-Acceptance"),
    "borrower_intentions":    ("Unknown · Retention · Disposition", "Retention"),
    "imminent_default":       ("CFPB Reg X标准值", "—"),
    "single_point_of_contact": ("联系人姓名或工号", "—"),
    # Section 4.1 — 破产
    "active_bk_flag":         ("0 / 1", "1"),
    "bk_status":              ("数值编码（Newrez内部，BPS直接存储）", "2"),
    "bk_legal_status":        ("数值编码（Newrez内部，BPS直接存储）", "8"),
    "bk_status_date":         ("YYYY-MM-DD", "2024-01-15"),
    "bk_chapter":             ("7（清算）/ 11（重组）/ 13（个人还款计划）", "7"),
    "bk_filed_date":          ("YYYY-MM-DD", "2023-01-15"),
    "bk_removal_date":        ("YYYY-MM-DD", "2024-06-01"),
    "mfr_filed_date":         ("YYYY-MM-DD", "2025-08-01"),
    "mfr_hearing_results":    ("数值编码（Newrez内部）", "0"),
    "proof_of_claim_date":    ("YYYY-MM-DD", "2023-03-01"),
    "post_petition_due_date": ("YYYY-MM-DD", "2024-01-01"),
}

# ── Newrez状态 content upgrades (col12 after insert) ─────────────────────────
NEWREZ_UPGRADES = {
    "demand_sent_date":           "✅ 85.9%（demandsentdate，2021-10-18~2026-04-20，见doc 13 Q5）",
    "servicer_days_in_fcl":       "✅ 100%（smsdaysinfc，1~606天）",
    "days_in_fcl":                "✅ 100%（daysinfc，1~814天）",
    "fcl_setup_date":             "✅ 100%（fcsetupdate，2024-02-07~2026-05-26）",
    "hold_reason":                "❌ 自由文本（fchold1description：Loss Mitigation Workout·Court Delay 等，非标准枚举）",
    "bid_amount":                 "✅ 9.0%（fcbidamount，$90,000~$543,305.96）",
    "approved_bid_price":         "✅ 8.9%（fcapprbidprice，$90,000~$536,008.42）",
    "sale_amount":                "⚠️ 4.7%（fcsaleamount，$90,001~$400,000；填充率异常，见doc 13 Q9）",
    "lien_position":              "✅ 100%（lienposition，1或2，portnewrezgeneral实测）",
    "interest_paid_through_date": "✅ 100%（interestpaidthroughdate，2022-09-30~2026-11-30）",
    "reason_for_default":         "✅ 46.2%（reasonfordefault：Unable to Contact Borrower·Reduction in Income·Excessive Obligations 等15+种）",
    "lm_deal":                    "✅ 整数编码（lmdeal: 1=Modification·2=Evaluation·4=Payment Plan·5=Forbearance·6=Short Sale·7=DIL·9=Payoff·11=Deferment；ETL解码为文本）",
    "lm_program":                 "✅ 整数编码（lmprogram: 21=Evaluation·73=Deferment·419=Bridger mod·396=VA Traditional 等20+种；ETL解码为文本）",
    "lm_status":                  "✅ 整数编码（lmstatus: 166=Pending Financials·112=Workout Denial·5=Document Follow-up 等20+种；ETL解码为文本）",
    "lm_final_disposition":       "✅ 整数编码（lmdecision: 99=Pending·6=Referral to FC·10=Request Incomplete·11=LMS Opened in Error 等；ETL解码）",
    "lm_denial_reason":           "✅ 整数编码（denialreason: 109=Loan not due 3+ months·4=Withdrawal·6=Ineligible 等18种；ETL解码）",
    "borrower_intentions":        "✅ 整数编码（borrowerintention: 1=Unknown·2=Retention·3=Disposition；SQL-D6实测）",
    "bk_status":                  "⚠️ 整数编码（bkstatus: 1~5，BPS直接存储数字，ETL未解码，见SQL-13）",
    "bk_legal_status":            "⚠️ 整数编码（bkstage: 0~22，常见8/21/1/4/9，BPS直接存储数字，ETL未解码，见SQL-13）",
}


def thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def copy_style(src, dst):
    """Copy fill, font (except value), alignment, border from src to dst."""
    if src.fill and src.fill.fgColor:
        dst.fill = copy.copy(src.fill)
    if src.font:
        dst.font = copy.copy(src.font)
    dst.alignment = copy.copy(src.alignment) if src.alignment else Alignment(wrap_text=True, vertical="center")
    dst.border = thin()


def main():
    wb = load_workbook(XLSX)
    ws = None
    for s in wb.worksheets:
        if "Field Spec" in s.title:
            ws = s
            break

    print(f"Sheet: {ws.title}")
    print(f"Original dimensions: {ws.max_row} rows x {ws.max_column} cols")

    # ── Step 1: Insert 2 columns at position 9 ──────────────────────────────
    ws.insert_cols(9, 2)

    # ── Step 2: Header row ──────────────────────────────────────────────────
    hdr_fill = PatternFill("solid", fgColor="2E75B6")
    hdr_font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, name in [(9, "标准接口取值范围"), (10, "典型取值（标准）")]:
        cell = ws.cell(1, col, name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = thin()

    # Set column widths
    ws.column_dimensions[get_column_letter(9)].width = 42
    ws.column_dimensions[get_column_letter(10)].width = 22

    # ── Step 3: Data rows ────────────────────────────────────────────────────
    # Find row height and fill patterns from a reference data row (row 3)
    ref_fill_even = ws.cell(3, 7).fill  # existing row fill
    ref_fill_odd = ws.cell(4, 7).fill if ws.max_row >= 4 else ref_fill_even
    ref_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    filled = 0
    for r in range(2, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        c3 = ws.cell(r, 3).value

        # Skip section header rows (c1 starts with 'Sect' or c3 is None)
        if c1 is None or (isinstance(c1, str) and c1.startswith("Sect")):
            continue
        if not isinstance(c3, str):
            continue

        field = c3.strip()
        if field not in MAPPING:
            continue

        val_range, typical = MAPPING[field]

        # Get fill color from adjacent cell to match row alternation
        adj_fill = ws.cell(r, 7).fill

        for col, val in [(9, val_range), (10, typical)]:
            cell = ws.cell(r, col, val)
            if adj_fill and adj_fill.fgColor and adj_fill.fgColor.rgb not in ("00000000", "FFFFFFFF", None):
                cell.fill = copy.copy(adj_fill)
            cell.font = Font(name="微软雅黑", size=9)
            cell.alignment = ref_align
            cell.border = thin()

        filled += 1

    print(f"Filled {filled} data rows with value range + typical value")

    # ── Step 4: Upgrade Newrez状态 (now col12) ───────────────────────────────
    upgraded = 0
    for r in range(2, ws.max_row + 1):
        c3 = ws.cell(r, 3).value
        if not isinstance(c3, str):
            continue
        field = c3.strip()
        if field in NEWREZ_UPGRADES:
            cell = ws.cell(r, 12)
            cell.value = NEWREZ_UPGRADES[field]
            upgraded += 1

    print(f"Upgraded {upgraded} Newrez状态 cells")

    wb.save(XLSX)
    print(f"Saved: {XLSX}")
    print(f"Final dimensions: {ws.max_row} rows x {ws.max_column} cols")


if __name__ == "__main__":
    main()
