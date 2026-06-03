"""
Fix all compliance count and rate errors in Overview and 阅读指南 sheets.

Errors found by comparing Field Spec (ground truth) vs Overview:

  Count errors:
    2.3  ✅:  9 → 10   ❌: 5 → 4    (one field miscounted)
    2.4  ✅: 12 → 15               (all 15 Hold fields are ✅ in FS)

  Rate errors (rate formula: ✅ / (✅+⚠️+❌)):
    2.2  88%  → ~78%   (7/9)
    2.3  ~56% → ~62%   (10/16)
    2.5  83%  → ~67%   (2/3)
    4.1  91%  → ~73%   (8/11)

  Grand totals update:
    ✅: 63 → 67   (2.3+1 + 2.4+3)
    ⚠️: 12 (unchanged)
    ❌: 13 → 12   (2.3-1)
    Overall rate: ~67% → ~74%   (67/91, excluding N/A ETL-derived field)

Run from project root: python scripts/fix_all_compliance.py
"""

import sys, io
from pathlib import Path
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

XLSX = Path(__file__).parent.parent / "docs" / "14_servicer_fcl_field_spec.xlsx"

# Corrections: section → (✅, ⚠️, ❌, rate)
CORRECTIONS = {
    "2.3": (10, 2, 4, "~62%"),
    "2.4": (15, 0, 0, "100%"),
    "2.2": (7,  2, 0, "~78%"),   # counts already correct, rate wrong
    "2.5": (2,  1, 0, "~67%"),   # counts already correct, rate wrong
    "4.1": (8,  3, 0, "~73%"),   # counts already correct, rate wrong
}

GRAND_TOTAL = (67, 12, 12, "~74%")  # ✅, ⚠️, ❌, rate


def fix_overview(ov):
    import re
    changed = []
    for r in range(1, 80):
        a = ov.cell(r, 1).value
        if not isinstance(a, str):
            continue
        # Match section rows
        m = re.search(r"(\d+\.\d+)", a)
        if m and m.group(1) in CORRECTIONS:
            sec = m.group(1)
            ok, part, miss, rate = CORRECTIONS[sec]
            old = (ov.cell(r,3).value, ov.cell(r,4).value,
                   ov.cell(r,5).value, ov.cell(r,6).value)
            ov.cell(r, 3).value = ok
            ov.cell(r, 4).value = part
            ov.cell(r, 5).value = miss
            ov.cell(r, 6).value = rate
            changed.append(f"  [{sec}] row {r}: {old} → ({ok}, {part}, {miss}, {rate!r})")
        # Grand total row
        if "合计" in a:
            ok, part, miss, rate = GRAND_TOTAL
            old = (ov.cell(r,3).value, ov.cell(r,4).value,
                   ov.cell(r,5).value, ov.cell(r,6).value)
            ov.cell(r, 3).value = ok
            ov.cell(r, 4).value = part
            ov.cell(r, 5).value = miss
            # rate cell may or may not exist in Overview totals row
            if ov.cell(r, 6).value is not None:
                ov.cell(r, 6).value = rate
            changed.append(f"  [合计] row {r}: {old} → ({ok}, {part}, {miss}, {rate!r})")
    return changed


def fix_guide(guide):
    """
    阅读指南 Newrez 现状 table:
      col A=字段组, B=总字段数, C=✅, D=⚠️, E=❌  (no rate column)
    Data rows: 20-28, total row: 29
    Only update sections whose counts changed (2.3, 2.4).
    """
    import re
    changed = []
    count_corrections = {
        "2.3": (10, 2, 4),
        "2.4": (15, 0, 0),
    }
    for r in range(18, 32):
        a = guide.cell(r, 1).value
        if not isinstance(a, str):
            continue
        m = re.search(r"(\d+\.\d+)", a)
        if m and m.group(1) in count_corrections:
            sec = m.group(1)
            ok, part, miss = count_corrections[sec]
            old = (guide.cell(r,3).value, guide.cell(r,4).value, guide.cell(r,5).value)
            guide.cell(r, 3).value = ok
            guide.cell(r, 4).value = part
            guide.cell(r, 5).value = miss
            changed.append(f"  [{sec}] row {r}: {old} → ({ok}, {part}, {miss})")
        if "合计" in a:
            ok, part, miss, _ = GRAND_TOTAL
            old = (guide.cell(r,3).value, guide.cell(r,4).value, guide.cell(r,5).value)
            guide.cell(r, 3).value = ok
            guide.cell(r, 4).value = part
            guide.cell(r, 5).value = miss
            changed.append(f"  [合计] row {r}: {old} → ({ok}, {part}, {miss})")
    return changed


def main():
    if not XLSX.exists():
        raise FileNotFoundError(f"Not found: {XLSX}")

    wb = load_workbook(XLSX)

    guide = wb.worksheets[0]   # 阅读指南
    ov    = wb.worksheets[2]   # Overview

    print("=== Overview fixes ===")
    for msg in fix_overview(ov):
        print(msg)

    print("\n=== 阅读指南 fixes ===")
    for msg in fix_guide(guide):
        print(msg)

    wb.save(XLSX)
    print(f"\nSaved: {XLSX}")
    print("\nFinal grand totals: ✅=67  ⚠️=12  ❌=12  rate=~74%")


if __name__ == "__main__":
    main()
