# -*- coding: utf-8 -*-
"""
略扩 reo_acquisition_date 的「验证SQL」头注，解释 dtdeedrecorded 含义（查询本身不变，验证结果不变）。
按表头定位列，人工列受保护。Excel 须关闭；改后运行 sync_fieldspec_excel_to_md.py。
Run: python scripts/fix_reo_acq_sql_comment.py
"""
import sys, io
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent))
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from _excel_guard import col_by_header, assert_safe

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")
NEW_SQL = (
    "-- N/A Newrez未提供独立 REO 取得日；可参考 dtdeedrecorded=止赎契据登记日（成交后产权转让契据在县登记处登记="
    "产权过户/止赎完成点，约在拍卖 fcsalehelddate 后 2-3 周，多数→REO），近似 REO 取得日（最新快照取10笔样例）\n"
    "-- 源表: newrez.portnewrezfc | 运行于: mysql_dev\n"
    "SELECT loanid, dataasof, dtdeedrecorded\n"
    "FROM   newrez.portnewrezfc\n"
    "WHERE  dataasof = (SELECT MAX(dataasof) FROM newrez.portnewrezfc)\n"
    "  AND  dtdeedrecorded IS NOT NULL\n"
    "LIMIT  10;"
)


def main():
    wb = load_workbook(XLSX)
    ws = next(s for s in wb.worksheets if "Field Spec" in s.title)
    fcol = col_by_header(ws, "标准接口字段") or 3
    sqlcol = col_by_header(ws, "验证SQL")
    assert_safe(ws, sqlcol)
    thin = Side(style="thin", color="BFBFBF")
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, fcol).value == "reo_acquisition_date":
            cell = ws.cell(r, sqlcol)
            cell.value = NEW_SQL
            cell.font = Font(name="Consolas", size=8)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            print("reo_acquisition_date 验证SQL 头注已扩充")
            break
    else:
        raise RuntimeError("reo_acquisition_date row not found")
    wb.save(XLSX)
    print(f"Saved: {XLSX}")


if __name__ == "__main__":
    main()
