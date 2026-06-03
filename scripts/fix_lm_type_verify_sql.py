# -*- coding: utf-8 -*-
"""
把 lm_type 的「验证SQL」从「仅验 Newrez 源 lmdeal 码」改为跨表 join（同时验 Newrez 源 lmdeal 与
BPS 表 deal），与 lm_deal 一致——因 lm_type ≡ lmdeal 解码后的 deal 大类。

按表头定位列，人工列受保护。Excel 须关闭；改后运行 run_verify_sql_results.py 重填验证结果，再 sync。
Run: python scripts/fix_lm_type_verify_sql.py
"""
import sys, io
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent))
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from _excel_guard import col_by_header, assert_safe

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")
NEW_SQL = (
    "-- 验证 lm_type 取值范围（lm_type ≡ lmdeal 解码后的 deal 大类；同时验证 Newrez 源 lmdeal 与 BPS 表 deal）\n"
    "-- 源表: newrez.portnewrezlm JOIN bpms_dev.sync_loan_foreclosure_loss_mitigation\n"
    "-- 运行于: mysql_bpms_dev（跨库 JOIN）| 解码源: Redshift portnewrezdatadic(LMDeal)，代码 basic_data_pool_config.py:835\n"
    "-- 快照表：CTE 取每个 LM 周期的最新快照（latest-snapshot-per-cycle）\n"
    "WITH latest_lm AS (\n"
    "  SELECT loanid, dealstartdate, lmdeal,\n"
    "         ROW_NUMBER() OVER (PARTITION BY loanid, dealstartdate ORDER BY dataasof DESC) AS rn\n"
    "  FROM newrez.portnewrezlm WHERE lmdeal IS NOT NULL\n"
    ")\n"
    "SELECT l.lmdeal, b.deal, COUNT(*) AS cnt\n"
    "FROM   latest_lm l\n"
    "JOIN   bpms_dev.sync_loan_foreclosure_loss_mitigation b\n"
    "  ON   l.loanid = b.loanid AND l.dealstartdate = b.cycle_opened_date\n"
    "WHERE  l.rn = 1\n"
    "  AND  b.deal IS NOT NULL AND b.deal != '' AND b.deal NOT REGEXP '^[0-9]'\n"
    "GROUP  BY l.lmdeal, b.deal ORDER BY l.lmdeal, cnt DESC;"
)


def main():
    wb = load_workbook(XLSX)
    ws = next(s for s in wb.worksheets if "Field Spec" in s.title)
    fcol = col_by_header(ws, "标准接口字段") or 3
    sqlcol = col_by_header(ws, "验证SQL")
    assert_safe(ws, sqlcol)
    thin = Side(style="thin", color="BFBFBF")
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, fcol).value == "lm_type":
            cell = ws.cell(r, sqlcol)
            print("OLD lm_type 验证SQL (前2行):", "\n".join((cell.value or "").splitlines()[:2]))
            cell.value = NEW_SQL
            cell.font = Font(name="Consolas", size=8)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            print("NEW lm_type 验证SQL = 跨表 join（同 lm_deal）")
            break
    else:
        raise RuntimeError("lm_type row not found")
    wb.save(XLSX)
    print(f"Saved: {XLSX}")


if __name__ == "__main__":
    main()
