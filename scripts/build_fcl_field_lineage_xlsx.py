"""
Build docs/21_fcl_field_lineage_matrix.xlsx — a filterable field-level lineage matrix
for the foreclosure (FCL) pipeline, derived from doc 21 (the single source of truth).

One row per core FCL field; columns trace L0 source file -> L1 source col -> L2/L3 ->
L4 business col -> L5 BPS col, plus the per-hop transform rule, servicer scope,
prod fill-rate, DB-verify status and code location.

Guard rules (project CLAUDE.md「Excel 人工列 / 批注保护规则」):
- generated columns are located BY HEADER (col_by_header), never by a hardcoded index;
- any column whose header contains 「人工」 is user-owned and never written (assert_safe);
- re-runnable: rows are keyed by the 字段 name (col 1); existing rows are updated in place
  and new fields appended, so user-inserted 人工 columns/comments survive a regenerate.

Run:  python scripts/build_fcl_field_lineage_xlsx.py
"""
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# --- guard helpers, inlined from scripts/_excel_guard.py (kept self-contained so this
#     generator has no cross-.py import; behaviour identical to that module) ---
MANUAL_MARK = "人工"

def thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def is_manual(ws, col):
    v = ws.cell(1, col).value
    return isinstance(v, str) and MANUAL_MARK in v

def col_by_header(ws, name_substring):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and name_substring in v and MANUAL_MARK not in v:
            return c
    return None

def assert_safe(ws, col):
    if is_manual(ws, col):
        raise RuntimeError(f"REFUSED to write column {col} — header contains 「人工」 (user-owned).")

def style_header(ws, col, text, fg, width):
    h = ws.cell(1, col, text)
    h.fill = PatternFill("solid", fgColor=fg)
    h.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    h.border = thin()
    ws.column_dimensions[get_column_letter(col)].width = width
    return h

HERE = os.path.dirname(os.path.abspath(__file__)) if "__file__" in globals() else os.path.join(os.getcwd(), "scripts")
ROOT = os.path.dirname(HERE)
OUT = os.path.normpath(os.path.join(ROOT, "docs", "21_fcl_field_lineage_matrix.xlsx"))
SHEET = "FCL字段血缘矩阵"
KEY = "字段(业务名)"  # row key column

# (header, width). Order = left-to-right layout for a fresh build.
HEADERS = [
    (KEY, 20),
    ("L0 来源文件", 18),
    ("L1 源 (servicer.表.列)", 34),
    ("L2 统一日表", 22),
    ("L3 清洗", 22),
    ("L4 业务表.列", 34),
    ("L5 BPS 列 (bpms)", 32),
    ("每跳转换规则", 50),
    ("Servicer 适用", 16),
    ("填充率 (prod 2026-06-06)", 22),
    ("DB 验证", 12),
    ("代码位置 (file:line)", 36),
]

# One dict per field. Keys MUST match HEADERS text. All facts are code-read + prod-verified
# (see doc 21 §1/§2/§6/§8). Newrez is the primary line; Carrington differences noted inline.
ROWS = [
    {KEY: "催告日 Demand", "L0 来源文件": "Newrez FC 文件 (S3)",
     "L1 源 (servicer.表.列)": "newrez.portnewrezfc.demandsentdate",
     "L2 统一日表": "—(走业务族支线)", "L3 清洗": "—",
     "L4 业务表.列": "fcl_stage_info.demand_start_date", "L5 BPS 列 (bpms)": "sync_fcl_stage_info.demand_start_date",
     "每跳转换规则": "改名直传;非 timeline 列;pre-FCL 催告支入库 (activefcflag!=1 且 demandsentdate 非空)",
     "Servicer 适用": "Newrez", "填充率 (prod 2026-06-06)": "—", "DB 验证": "✓ schema",
     "代码位置 (file:line)": "basic_data_pool_config.py L2037/L2354"},

    {KEY: "移交止赎日 Referral", "L0 来源文件": "Newrez/Carrington FC 文件",
     "L1 源 (servicer.表.列)": "newrez.portnewrezfc.fcreferraldate | carrington.portcarrington.fcl_referral_date",
     "L2 统一日表": "—", "L3 清洗": "—",
     "L4 业务表.列": "basic_data_loan_foreclosure.timeline_referred_to_foreclosure_date · fcl_stage_info.referral_start_date",
     "L5 BPS 列 (bpms)": "sync_loan_foreclosure.timeline_referred_to_foreclosure_date",
     "每跳转换规则": "改名直传 (fc[l_]referral[_]date→referral_start_date);也是 5-FORECLOSURE 入库过滤键",
     "Servicer 适用": "Newrez · Carrington", "填充率 (prod 2026-06-06)": "Newrez 100% · Carr 100%",
     "DB 验证": "✓ prod", "代码位置 (file:line)": "L1544 / Carr ~L1574-1614; GEN_FCL_DETAIL L259"},

    {KEY: "首次法律 First Legal", "L0 来源文件": "Newrez FC 文件",
     "L1 源 (servicer.表.列)": "newrez.portnewrezfc.firstlegaldate | Carrington 无源列",
     "L2 统一日表": "—", "L3 清洗": "—",
     "L4 业务表.列": "basic_data_loan_foreclosure.timeline_first_legal_date · fcl_stage_info.first_legal_start_date",
     "L5 BPS 列 (bpms)": "sync_loan_foreclosure.timeline_first_legal_date",
     "每跳转换规则": "Newrez 改名直传;⚠️ Carrington 无映射→全 NULL",
     "Servicer 适用": "Newrez (Carr=NULL)", "填充率 (prod 2026-06-06)": "Newrez 61% · Carr 0%",
     "DB 验证": "✓ prod", "代码位置 (file:line)": "L1549; GEN_FCL_DETAIL L262"},

    {KEY: "文书送达 Service", "L0 来源文件": "Newrez FC 文件",
     "L1 源 (servicer.表.列)": "newrez.portnewrezfc.servicecompletedate",
     "L2 统一日表": "—", "L3 清洗": "—",
     "L4 业务表.列": "basic_data_loan_foreclosure.timeline_service_date · fcl_stage_info.service_start_date",
     "L5 BPS 列 (bpms)": "sync_loan_foreclosure.timeline_service_date",
     "每跳转换规则": "改名直传", "Servicer 适用": "Newrez",
     "填充率 (prod 2026-06-06)": "—", "DB 验证": "✓ schema", "代码位置 (file:line)": "L1550; L263"},

    {KEY: "判决听证 Judgement", "L0 来源文件": "Newrez FC 文件",
     "L1 源 (servicer.表.列)": "newrez.portnewrezfc.fcjudgmenthearingscheduled | Carrington 无源列",
     "L2 统一日表": "—", "L3 清洗": "—",
     "L4 业务表.列": "timeline_judgement_date(当前) + timeline_judgement_hearing_set_date(min dataasof) · fcl_stage_info.judgement_start_date(max dataasof rnk1)",
     "L5 BPS 列 (bpms)": "sync_loan_foreclosure.timeline_judgement_date",
     "每跳转换规则": "同一原始列复用;set 日期=快照首见 min(dataasof);⚠️ Carrington 无→NULL",
     "Servicer 适用": "Newrez (Carr=NULL)", "填充率 (prod 2026-06-06)": "Newrez 13% · Carr 0%",
     "DB 验证": "✓ prod", "代码位置 (file:line)": "GEN_FCL_DETAIL L264-265; STAGE L1948-1958"},

    {KEY: "拍卖排期 Sale", "L0 来源文件": "Newrez/Carrington FC 文件",
     "L1 源 (servicer.表.列)": "newrez.portnewrezfc.fcscheduledsaledate | carrington.fcl_scheduled_sale_date",
     "L2 统一日表": "—", "L3 清洗": "—",
     "L4 业务表.列": "timeline_sale_date_projected_date + timeline_sale_date_set_date(min dataasof) · fcl_stage_info.sale_start_date",
     "L5 BPS 列 (bpms)": "sync_loan_foreclosure.timeline_sale_date_projected_date",
     "每跳转换规则": "改名直传;set 日期=min(dataasof) 首见", "Servicer 适用": "Newrez · Carrington",
     "填充率 (prod 2026-06-06)": "—", "DB 验证": "✓ schema", "代码位置 (file:line)": "L1553; L266-267"},

    {KEY: "拍卖举行 Sale Held", "L0 来源文件": "Newrez/Carrington FC 文件",
     "L1 源 (servicer.表.列)": "newrez.portnewrezfc.fcsalehelddate | carrington.fcl_sale_held_date",
     "L2 统一日表": "—", "L3 清洗": "—",
     "L4 业务表.列": "basic_data_loan_foreclosure.timeline_sale_date_held_date",
     "L5 BPS 列 (bpms)": "sync_loan_foreclosure.timeline_sale_date_held_date",
     "每跳转换规则": "改名直传", "Servicer 适用": "Newrez · Carrington",
     "填充率 (prod 2026-06-06)": "—", "DB 验证": "✓ schema", "代码位置 (file:line)": "L1554; L269"},

    {KEY: "止赎状态 Status", "L0 来源文件": "Newrez/Carrington FC 文件",
     "L1 源 (servicer.表.列)": "portnewrezfc.activefcflag+fcremovaldesc | carrington.fcl_flag(→activefcflag)+fcremovaldesc",
     "L2 统一日表": "—", "L3 清洗": "—",
     "L4 业务表.列": "basic_data_loan_foreclosure.summary_foreclosure_status",
     "L5 BPS 列 (bpms)": "sync_loan_foreclosure.summary_foreclosure_status",
     "每跳转换规则": "CASE: activefcflag=1→'Active Foreclosure'; =0且fcremovaldesc非空→CONCAT('Closed Foreclosure:',fcremovaldesc); 无 decode 表",
     "Servicer 适用": "Newrez · Carrington", "填充率 (prod 2026-06-06)": "Newrez 100% · Carr 83%",
     "DB 验证": "✓ prod", "代码位置 (file:line)": "GEN_FCL_DETAIL CASE L273"},

    {KEY: "司法/非司法 Judicial", "L0 来源文件": "Newrez FC 文件",
     "L1 源 (servicer.表.列)": "newrez.portnewrezfc.judicial(0/1) | Carrington 无源列",
     "L2 统一日表": "—", "L3 清洗": "—",
     "L4 业务表.列": "summary_judicial_foreclosure(decimal) · summary_type(标签) · fcl_stage_info.judicial(Y/N+州兜底)",
     "L5 BPS 列 (bpms)": "sync_loan_foreclosure.summary_judicial_foreclosure",
     "每跳转换规则": "CAST(judicial as decimal)+标签;⚠️ Carrington 无→NULL,阶段表按 basic_data_judicial_config 州兜底",
     "Servicer 适用": "Newrez (Carr=NULL)", "填充率 (prod 2026-06-06)": "Newrez 100% · Carr 0%",
     "DB 验证": "✓ prod", "代码位置 (file:line)": "L277/L279; STAGE L2351-2353"},

    {KEY: "阶段天数 Stage Days", "L0 来源文件": "(计算)",
     "L1 源 (servicer.表.列)": "(由里程碑日期算)", "L2 统一日表": "—", "L3 清洗": "—",
     "L4 业务表.列": "fcl_stage_info.{stage}_stage_days/_in_lm_days/_on_hold_days/to_*_days",
     "L5 BPS 列 (bpms)": "sync_fcl_stage_info.*",
     "每跳转换规则": "datediff(start,end|curr_date)+1;扣减与未结束 LM/Hold 区间重叠 (greatest/least);Carr daysinfc=DATEDIFF(referral,snap)+1 动态",
     "Servicer 适用": "Newrez · Carrington", "填充率 (prod 2026-06-06)": "—", "DB 验证": "✓ schema",
     "代码位置 (file:line)": "GEN_FCL_STAGE L2053-2266"},

    {KEY: "Hold 暂停", "L0 来源文件": "Newrez/Carrington FC 文件",
     "L1 源 (servicer.表.列)": "portnewrezfc.fchold1..4{desc,start,end} | carrington.fcl_active_hold/fcl_latest_hold_*",
     "L2 统一日表": "—", "L3 清洗": "—",
     "L4 业务表.列": "basic_data_loan_foreclosure_hold.description1..4{,_start_date,_end_date} (宽)",
     "L5 BPS 列 (bpms)": "sync_loan_foreclosure_hold.{description,_start_date,_end_date} (长)",
     "每跳转换规则": "Newrez 宽槽直拷;Carrington 多次 Hold 用 ROW_NUMBER rank→4 槽,hold_end=MAX(snap);L5 宽→长 unpivot",
     "Servicer 适用": "Newrez · Carrington (1:N)", "填充率 (prod 2026-06-06)": "样本 7727000088=9 条",
     "DB 验证": "✓ prod", "代码位置 (file:line)": "Hold L466-768; Carr ~L504-629"},

    {KEY: "LM 周期", "L0 来源文件": "Newrez LM 文件",
     "L1 源 (servicer.表.列)": "portnewrezlm.{lmdeal,lmprogram,lmstatus,dealstartdate,lmremovaldate,lmdecision}",
     "L2 统一日表": "—", "L3 清洗": "—",
     "L4 业务表.列": "basic_data_loan_foreclosure_loss_mitigation.{deal,program,lmc_status,cycle_opened_date,cycle_closed_date,final_disposition}",
     "L5 BPS 列 (bpms)": "sync_loan_foreclosure_loss_mitigation.*",
     "每跳转换规则": "datadic 解码 COALESCE(desc,原值) join concat(code,'.0');按 (loanid,dealstartdate) rn=1",
     "Servicer 适用": "Newrez (1:N)", "填充率 (prod 2026-06-06)": "样本 7727000088=9 周期",
     "DB 验证": "✓ prod", "代码位置 (file:line)": "LM L799-843; datadic L821-838"},

    {KEY: "BK 破产 + MFR", "L0 来源文件": "Newrez BK 文件 / Carrington",
     "L1 源 (servicer.表.列)": "portnewrezbk.{bkstatus,bkchapter,bkfileddate,pocfileddate}+general.legalstatus | carrington.bk_mfr_filed_date",
     "L2 统一日表": "—", "L3 清洗": "—",
     "L4 业务表.列": "basic_data_loan_foreclosure_bankruptcy.{bankruptcy_status,chapter,status_date,proof_of_claim_date,legal_status,mfr_filed_date}",
     "L5 BPS 列 (bpms)": "sync_loan_foreclosure_bankruptcy.*",
     "每跳转换规则": "bkstatus datadic 解码;chapter cast;⚠️ mfr_filed_date 仅 Carrington 有 (Newrez NULL)",
     "Servicer 适用": "Newrez · Carrington(MFR) (1:N)", "填充率 (prod 2026-06-06)": "样本 7727000010=2 次",
     "DB 验证": "✓ prod", "代码位置 (file:line)": "BK L331-370; Carr MFR L403"},

    {KEY: "逾期码 delinq", "L0 来源文件": "各家 general/daily 文件",
     "L1 源 (servicer.表.列)": "newrez.portnewrezgeneral.delinquency_status_mba,nextduedate | carrington.loan_status",
     "L2 统一日表": "basic_data_daily_loan_common.delq_status",
     "L3 清洗": "basic_data_daily_loan_common_clean.delinq",
     "L4 业务表.列": "basic_data_fcl_related.delq_status → fcl_stage_info.group",
     "L5 BPS 列 (bpms)": "(经阶段 group/月度表,不直进 sync_loan_foreclosure)",
     "每跳转换规则": "CASE: 文本判 FCL/REO/P/BK + days360(nextduedate,fctrdt) 分档 C/D30/D60/D90/D120P;各家文案不同(见 §7.3)",
     "Servicer 适用": "全部 (各家 CASE 不同)", "填充率 (prod 2026-06-06)": "C 121437·FCL 599 等",
     "DB 验证": "✓ prod", "代码位置 (file:line)": "clean_config Newrez L355-368 / Carr L670-683"},
]

HEADER_FILL = "305496"


def ensure_headers(ws):
    """Return {header_text: col_index}, creating any missing generated header by name.
    New headers append after the last *used* row-1 column (so a fresh, empty sheet starts
    at column 1 — openpyxl reports max_column=1 for an empty sheet, which next_free_col
    would wrongly skip)."""
    used = [c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value not in (None, "")]
    nxt = (max(used) + 1) if used else 1
    pos = {}
    for text, width in HEADERS:
        col = col_by_header(ws, text)
        if col is None:
            col = nxt
            nxt += 1
            style_header(ws, col, text, HEADER_FILL, width)
        pos[text] = col
    ws.freeze_panes = "A2"
    return pos


def field_row_index(ws, key_col):
    """Map existing 字段 name -> row (rows 2+), for in-place idempotent update."""
    idx = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, key_col).value
        if v:
            idx[str(v).strip()] = r
    return idx


def main():
    if os.path.exists(OUT):
        wb = load_workbook(OUT)
        ws = wb[SHEET] if SHEET in wb.sheetnames else wb.create_sheet(SHEET)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET

    pos = ensure_headers(ws)
    key_col = pos[KEY]
    existing = field_row_index(ws, key_col)

    wrote = 0
    for row in ROWS:
        name = row[KEY].strip()
        r = existing.get(name) or (ws.max_row + 1)  # append after last row for a new field
        if name not in existing:
            existing[name] = r
        for text, _w in HEADERS:
            col = pos[text]
            assert_safe(ws, col)  # never write a 人工 column
            cell = ws.cell(r, col, row.get(text, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="微软雅黑", size=9)
            cell.border = thin()
        wrote += 1

    wb.save(OUT)
    print(f"OK: wrote {wrote} field rows x {len(HEADERS)} generated cols -> {OUT}")
    print("    (re-runnable; manual/'人工' columns + comments preserved)".encode("ascii", "replace").decode())


if __name__ == "__main__":
    main()
