"""
Fix the stale foreclosure_flag Newrez状态 (col12) cell in Field Spec.

It still said "🟡 字段存在但ETL目前设为null（内部ETL问题，需优先修复）" — a claim about the
non-existent `fcl_flag` column. DB-verified: no fcl_flag; the real signal `activefcflag` is
populated 0/1 (no nulls). col4/col13/MD were already corrected; this cell was missed.

Run: python scripts/fix_foreclosure_flag_status.py  (Excel must be closed)
"""

import sys, io
from pathlib import Path
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")
NEW = ("🟡 Newrez 无独立 fcl_flag 列（DB 实测）；FCL 活跃状态由 activefcflag 表达——最新快照实测 "
       "0/1 正常填充（非 null）。原\"ETL 置 null\"说法系针对不存在的 fcl_flag，已更正")


def main():
    wb = load_workbook(XLSX)
    ws = None
    for s in wb.worksheets:
        if "Field Spec" in s.title:
            ws = s
            break
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 3).value == "foreclosure_flag":
            old = ws.cell(r, 12).value
            ws.cell(r, 12).value = NEW
            print(f"row {r} col12:\n  OLD: {old!r}\n  NEW: {NEW!r}")
            break
    wb.save(XLSX)
    print(f"Saved: {XLSX}")


if __name__ == "__main__":
    main()
