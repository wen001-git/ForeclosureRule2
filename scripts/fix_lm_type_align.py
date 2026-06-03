# -*- coding: utf-8 -*-
"""
统一 lm_type（标准接口字段）= Newrez 实际 8 个 deal（与 lm_deal/DB/doc 13 §5 一致），消除 doc 14 自身
col H(格式规则,6)≠col I(取值范围,5) 的矛盾。8 deal 经 DB 实测 + Redshift portnewrezdatadic(LMDeal) 确认。

按表头定位列，人工列受保护。Excel 须关闭；改后运行 sync_fieldspec_excel_to_md.py。
Run: python scripts/fix_lm_type_align.py
"""
import sys, io
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, str(Path(__file__).parent))
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from _excel_guard import col_by_header, assert_safe

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")
DEALS = ["Evaluation", "Modification", "Forbearance", "Payment Plan", "Deferment", "Short Sale", "DIL", "Payoff"]
NEW_RANGE = "\n".join(DEALS)  # 一值一行（≥5）
NEW_FMT = ("Newrez lmdeal 整数码解码为 deal 大类（8）：Evaluation/Modification/Forbearance/Payment Plan/"
           "Deferment/Short Sale/DIL/Payoff；Trial Period Plan 属 program 层（Modification 试行），非独立 deal")
NEW_STATUS = ("❌ Newrez 不直接提供标准 lm_type 文本；以 lmdeal 整数码提供，ETL 解码为上述 8 个 deal"
              "（见 lm_deal 行；解码源 Redshift portnewrezdatadic）")


def main():
    wb = load_workbook(XLSX)
    ws = next(s for s in wb.worksheets if "Field Spec" in s.title)
    fcol = col_by_header(ws, "标准接口字段") or 3
    rcol = col_by_header(ws, "标准接口取值范围")
    hcol = col_by_header(ws, "格式/计算规则")
    scol = col_by_header(ws, "Newrez状态")
    for c in (rcol, hcol, scol):
        assert_safe(ws, c)
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, fcol).value == "lm_type":
            cell = ws.cell(r, rcol)
            print("OLD 取值范围:", repr(cell.value))
            cell.value = NEW_RANGE
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.cell(r, hcol).value = NEW_FMT
            ws.cell(r, scol).value = NEW_STATUS
            print("NEW 取值范围:", repr(NEW_RANGE))
            print("NEW 格式规则:", NEW_FMT)
            print("NEW Newrez状态:", NEW_STATUS)
            break
    else:
        raise RuntimeError("lm_type row not found")
    wb.save(XLSX)
    print(f"Saved: {XLSX}")


if __name__ == "__main__":
    main()
