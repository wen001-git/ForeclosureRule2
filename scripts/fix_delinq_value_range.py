"""
Fix delinquency_status 标准接口取值范围 (col9) in Field Spec sheet.

Old (wrong - mixes internal code D90 + invented '150+ DPD'):
  MBA标准文本枚举：Foreclosure · 120-149 DPD · 150+ DPD · D90（禁止数字串如'29.0'）

New (correct - raw MBA transmission-layer text only, per doc 08 / doc 14 allowed-values table):
  MBA标准文本枚举（Servicer传输层）：Current · 1-29 / 30-59 / 60-89 / 90-119 / 120-149 / 150-179 DPD ...

Also scans all col9 cells for stray internal codes.

Run: python scripts/fix_delinq_value_range.py
"""

import sys, io, re
from pathlib import Path
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

XLSX = Path("docs/14_servicer_fcl_field_spec.xlsx")

NEW_VALUE = (
    "MBA标准文本枚举（Servicer传输层）：Current · 1-29 / 30-59 / 60-89 / 90-119 / "
    "120-149 / 150-179 DPD · 180+ DPD · Foreclosure · Foreclosure/Non-Perf BK · "
    "Performing/Non-Performing Bankruptcy · REO · REO Sale · 3rd Party Sale · "
    "Full Payoff · Service Release（共19种；禁止数字串如'29.0'；完整映射见doc 08 / doc 14 允许值表）"
)

# internal codes that should NOT appear in 标准接口取值范围 as standalone tokens
INTERNAL_CODES = ["D30", "D60", "D90", "D120P"]


def main():
    wb = load_workbook(XLSX)
    ws = None
    for s in wb.worksheets:
        if "Field Spec" in s.title:
            ws = s
            break

    # Fix delinquency_status row
    fixed = False
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 3).value == "delinquency_status":
            old = ws.cell(r, 9).value
            ws.cell(r, 9).value = NEW_VALUE
            print(f"[FIX] row {r} col9:")
            print(f"  OLD: {old!r}")
            print(f"  NEW: {NEW_VALUE!r}")
            fixed = True
            break
    if not fixed:
        print("WARNING: delinquency_status row not found")

    # Sanity scan: other col9 cells with stray internal codes
    print("\n[SCAN] col9 cells containing internal codes (excluding delinquency_status):")
    found_any = False
    for r in range(2, ws.max_row + 1):
        field = ws.cell(r, 3).value
        if field == "delinquency_status" or not isinstance(field, str):
            continue
        v = ws.cell(r, 9).value
        if not isinstance(v, str):
            continue
        # match internal codes as whole tokens
        for code in INTERNAL_CODES:
            if re.search(rf"\b{code}\b", v):
                print(f"  row {r} ({field}): {v!r}")
                found_any = True
                break
    if not found_any:
        print("  (none — all clean)")

    wb.save(XLSX)
    print(f"\nSaved: {XLSX}")


if __name__ == "__main__":
    main()
