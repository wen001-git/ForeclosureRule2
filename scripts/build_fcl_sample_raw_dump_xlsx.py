# -*- coding: utf-8 -*-
"""
生成 docs/19_fcl_sample_loan_raw_dump.xlsx —— 5 个样例贷款在 BPS foreclosure 相关【所有表的全部字段】
的原始取值转储，一个 sheet 一张表。

目的：doc 16 只展示各面板【用到的部分字段】；本表把 5 个样例贷款在每张相关表的**每个字段**逐一列出，
让读者完整观察"业务 ↔ 数据"的关系。

涉及的表（DB 实测）：
  Newrez 源表(5)：portnewrezfc / portnewrezbk / portnewrezlm / portnewrezgeneral / portnewrezprop
  BPS 直接对接表(5 基础 + 1 视图)：sync_loan_foreclosure / _hold / _loss_mitigation / _bankruptcy /
    sync_fcl_stage_info / biz_data_view_loan_details_foreclosure(视图)
  不纳入(中间层/dev无)：portnewrezdatadic(解码字典) / port.basic_data_*(Redshift中间) / port.portfunding(融资池)

排版（混合）：1 行/贷款的表 → 转置（字段为行 × 5 贷款列）；多行/贷款的表 → 平铺（记录为行 × 全字段列）。
取数口径：Newrez 源表 dataasof=MAX(每贷款)；sync_fcl_stage_info fctrdt=MAX(每贷款)；sync 基础表当前态；
视图取每贷款 fctrdt 最新一行。

Run: python scripts/build_fcl_sample_raw_dump_xlsx.py（Excel 须关闭）
"""

import sys, io, os, datetime
from pathlib import Path
import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# DB 凭据从环境变量读取（勿硬编码）：必填 FCL_DB_PASSWORD；host/port/user 有默认值，可用同名环境变量覆盖
HOST = os.environ.get("FCL_DB_HOST", "bridg004-db-test.mysql.database.azure.com")
PORT = int(os.environ.get("FCL_DB_PORT", "3306"))
USER = os.environ.get("FCL_DB_USER", "brgdev")
PASSWORD = os.environ.get("FCL_DB_PASSWORD")
ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "docs" / "19_fcl_sample_loan_raw_dump.xlsx"
RUN_DATE = "2026-06-03"

LOANS = [
    ("7727000088", "Judicial(FL)·JUDGEMENT·Hold×7·LM×9"),
    ("7727000672", "Non-Judicial(MI)·REFERRAL"),
    ("7727004200", "Judicial(IL)·SALE"),
    ("7727000065", "BK + Hold×4 + 完结REO"),
    ("7727000010", "Chapter 13 Active BK（未入 FCL 管道）"),
]
LOAN_IDS = [l[0] for l in LOANS]

# 转置组（1 行/贷款）：(sheet_title, schema, table, kind)
TRANSPOSE = [
    ("② src·portnewrezfc", "newrez", "portnewrezfc", "newrez", "FCL 主源表（时间线/状态/Hold槽/金额/律师）"),
    ("③ src·portnewrezbk", "newrez", "portnewrezbk", "newrez", "破产源表"),
    ("④ src·portnewrezlm", "newrez", "portnewrezlm", "newrez", "损失缓解(LM)源表"),
    ("⑤ src·portnewrezgeneral", "newrez", "portnewrezgeneral", "newrez", "通用源表（legalstatus / delinquency_status_mba 等）"),
    ("⑥ src·portnewrezprop", "newrez", "portnewrezprop", "newrez", "房产源表（propertystate 等）"),
    ("⑦ bps·sync_loan_foreclosure", "bpms_dev", "sync_loan_foreclosure", "current", "Summary/Timeline/target 主表"),
    ("⑧ bps·sync_fcl_stage_info", "bpms_dev", "sync_fcl_stage_info", "stage", "聚合 Stage/Timeline 表"),
    ("⑫ bps·view_loan_details", "bpms_dev", "biz_data_view_loan_details_foreclosure", "view", "详情页视图（actual/var 天数；取每贷款最新 fctrdt 一行）"),
]
# 平铺组（多行/贷款）：(sheet_title, table, sortcol, desc)
FLAT = [
    ("⑨ bps·hold", "sync_loan_foreclosure_hold", "description_start_date", "Hold 全历史（每次变更一行）"),
    ("⑩ bps·loss_mitigation", "sync_loan_foreclosure_loss_mitigation", "cycle_opened_date", "LM 周期历史"),
    ("⑪ bps·bankruptcy", "sync_loan_foreclosure_bankruptcy", "status_date", "破产记录"),
]

# ── 样式 ──
DARK, MID, GREEN, LIGHT, GREY = "1F3864", "2E75B6", "548235", "D6E4F0", "F2F2F2"
YH = "微软雅黑"
SIDE = Side(style="thin", color="BFBFBF")
BORDER = Border(left=SIDE, right=SIDE, top=SIDE, bottom=SIDE)
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def fmt(v):
    if v is None:
        return "—"
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v)
    return s if s.strip() else "—"


def section_title(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.fill = PatternFill("solid", fgColor=DARK)
    c.font = Font(name=YH, bold=True, size=11, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 20
    return row + 1


def note(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.fill = PatternFill("solid", fgColor="FFF2CC")
    c.font = Font(name=YH, size=9, italic=True, color="7F6000")
    c.alignment = WRAP
    return row + 1


def header_cell(ws, row, col, text, fill=MID):
    c = ws.cell(row, col, text)
    c.fill = PatternFill("solid", fgColor=fill)
    c.font = Font(name=YH, bold=True, size=9, color="FFFFFF")
    c.alignment = CENTER
    c.border = BORDER
    return c


def body_cell(ws, row, col, text, fill="FFFFFF", bold=False):
    c = ws.cell(row, col, text)
    c.font = Font(name=YH, size=9, bold=bold)
    c.alignment = WRAP
    c.border = BORDER
    c.fill = PatternFill("solid", fgColor=fill)
    return c


# ── DB ──
def cols_of(cur, schema, table):
    cur.execute("SELECT column_name FROM information_schema.columns "
                "WHERE table_schema=%s AND table_name=%s ORDER BY ordinal_position", (schema, table))
    return [r[0] for r in cur.fetchall()]


def one_row(cur, schema, table, loan, kind):
    """返回该贷款单行 dict（无则 None）。"""
    if kind == "newrez":
        sql = (f"SELECT * FROM {schema}.{table} WHERE loanid=%s "
               f"AND dataasof=(SELECT MAX(dataasof) FROM {schema}.{table} WHERE loanid=%s) LIMIT 1")
        args = (loan, loan)
    elif kind == "stage":
        sql = (f"SELECT * FROM {schema}.{table} WHERE loanid=%s "
               f"AND fctrdt=(SELECT MAX(fctrdt) FROM {schema}.{table} WHERE loanid=%s) LIMIT 1")
        args = (int(loan), int(loan))
    elif kind == "view":
        sql = f"SELECT * FROM {schema}.{table} WHERE loanid=%s ORDER BY fctrdt DESC, id DESC LIMIT 1"
        args = (int(loan),)
    else:  # current
        sql = f"SELECT * FROM {schema}.{table} WHERE loanid=%s LIMIT 1"
        args = (int(loan),)
    cur.execute(sql, args)
    r = cur.fetchone()
    if not r:
        return None
    return dict(zip([d[0] for d in cur.description], r))


def all_rows(cur, table, sortcol):
    ids = ",".join(str(int(x)) for x in LOAN_IDS)
    cur.execute(f"SELECT * FROM bpms_dev.{table} WHERE loanid IN ({ids}) "
                f"ORDER BY loanid, COALESCE({sortcol}, '1900-01-01')")
    cols = [d[0] for d in cur.description]
    return cols, [dict(zip(cols, r)) for r in cur.fetchall()]


# ── 写表 ──
def write_transpose(ws, schema, table, kind, desc, cur):
    cols = cols_of(cur, schema, table)
    rows = {lid: one_row(cur, schema, table, lid, kind) for lid in LOAN_IDS}
    present = sum(1 for v in rows.values() if v)
    snap = {"newrez": "dataasof=MAX(每贷款)", "stage": "fctrdt=MAX(每贷款)",
            "view": "每贷款 fctrdt 最新一行", "current": "当前态(1行/贷款)"}[kind]
    r = 1
    r = section_title(ws, r, f"{schema}.{table} — {desc}", 1 + len(LOAN_IDS))
    r = note(ws, r, f"全 {len(cols)} 字段 · 取数口径：{snap} · 命中 {present}/5 贷款（无行的贷款列显示 —）"
                    f" · 运行 {RUN_DATE}", 1 + len(LOAN_IDS))
    # header
    header_cell(ws, r, 1, "字段")
    for j, (lid, d) in enumerate(LOANS, start=2):
        header_cell(ws, r, j, lid)
    r += 1
    # field rows
    for i, col in enumerate(cols):
        fill = "FFFFFF" if i % 2 == 0 else GREY
        body_cell(ws, r, 1, col, fill=LIGHT, bold=True)
        for j, lid in enumerate(LOAN_IDS, start=2):
            d = rows[lid]
            body_cell(ws, r, j, fmt(d.get(col)) if d else "—", fill=fill)
        r += 1
    ws.column_dimensions["A"].width = 40
    for j in range(2, 2 + len(LOAN_IDS)):
        ws.column_dimensions[get_column_letter(j)].width = 24
    ws.freeze_panes = "B4"
    return present


def write_flat(ws, table, sortcol, desc, cur):
    cols, recs = all_rows(cur, table, sortcol)
    r = 1
    r = section_title(ws, r, f"bpms_dev.{table} — {desc}", max(2, len(cols)))
    per = {}
    for d in recs:
        per[str(d["loanid"])] = per.get(str(d["loanid"]), 0) + 1
    r = note(ws, r, f"全 {len(cols)} 字段 · 当前态(多行/贷款) · 共 {len(recs)} 行 · 各贷款行数：{per} · 运行 {RUN_DATE}",
             max(2, len(cols)))
    # header
    for j, col in enumerate(cols, start=1):
        header_cell(ws, r, j, col)
    r += 1
    for i, d in enumerate(recs):
        fill = "FFFFFF" if i % 2 == 0 else GREY
        for j, col in enumerate(cols, start=1):
            body_cell(ws, r, j, fmt(d.get(col)), fill=fill)
        r += 1
    for j, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(j)].width = 13 if col == "loanid" else 20
    ws.freeze_panes = "B4"
    return len(recs)


def build_cover(wb, cur, counts):
    ws = wb.active
    ws.title = "① 表清单 索引"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    r = 1
    r = section_title(ws, r, "doc 19 — 5 个样例贷款 · FCL 全表全字段原始转储", 5)

    def kv(label, value):
        nonlocal r
        a = ws.cell(r, 1, label); a.font = Font(name=YH, bold=True, size=10); a.fill = PatternFill("solid", fgColor=LIGHT)
        a.alignment = WRAP; a.border = BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        b = ws.cell(r, 2, value); b.font = Font(name=YH, size=10); b.alignment = WRAP; b.border = BORDER
        r += 1

    kv("文档目的", "把 5 个样例贷款在 BPS foreclosure 相关【所有表的全部字段】逐一列出（一个 sheet 一张表），"
                   "便于读者完整观察业务与数据的对应关系。doc 16 仅含各面板用到的部分字段，本表是其原始数据底座。")
    kv("目标读者", "数据工程师 · 业务分析师 · 验证人员 · 接入工程师")
    kv("取数口径", "Newrez 源表 dataasof=MAX(每贷款)；sync_fcl_stage_info fctrdt=MAX；sync 基础表当前态；"
                   "视图取每贷款 fctrdt 最新一行。日期统一 YYYY-MM-DD；空值显示 —。")
    kv("排版", "1 行/贷款的表 → 转置（字段为行 × 5 贷款列）；多行/贷款的表（hold/lm/bk）→ 平铺（记录为行 × 全字段列）。")
    kv("修订历史", f"{RUN_DATE}　v1 初稿（AI Agent）：12 sheet；所有取值 DB 实测。")
    r += 1

    r = section_title(ws, r, "一、Newrez 源数据表（5）", 5)
    header_cell(ws, r, 1, "Sheet"); header_cell(ws, r, 2, "表"); header_cell(ws, r, 3, "用途")
    header_cell(ws, r, 4, "列数"); header_cell(ws, r, 5, "本dump行/命中")
    r += 1
    for t in TRANSPOSE[:5]:
        title, sch, tab, kind, desc = t
        body_cell(ws, r, 1, title); body_cell(ws, r, 2, f"{sch}.{tab}"); body_cell(ws, r, 3, desc)
        body_cell(ws, r, 4, str(counts[tab][0])); body_cell(ws, r, 5, f"{counts[tab][1]}/5 贷款")
        r += 1

    r = section_title(ws, r, "二、BPS 直接对接表（5 基础表 + 1 视图）", 5)
    header_cell(ws, r, 1, "Sheet"); header_cell(ws, r, 2, "表 / 视图"); header_cell(ws, r, 3, "用途")
    header_cell(ws, r, 4, "列数"); header_cell(ws, r, 5, "本dump行")
    r += 1
    bps_order = [TRANSPOSE[5], TRANSPOSE[6]] + [(f[0], "bpms_dev", f[1], "flat", f[3]) for f in FLAT] + [TRANSPOSE[7]]
    for title, sch, tab, kind, desc in bps_order:
        body_cell(ws, r, 1, title); body_cell(ws, r, 2, f"{sch}.{tab}" + ("（视图）" if kind == "view" else ""))
        body_cell(ws, r, 3, desc); body_cell(ws, r, 4, str(counts[tab][0]))
        body_cell(ws, r, 5, f"{counts[tab][1]}" + ("/5 贷款" if kind in ("current", "stage", "view") else " 行"))
        r += 1

    r = section_title(ws, r, "三、不纳入本转储的表（中间层 / dev 不存在）", 5)
    for t, why in [("newrez.portnewrezdatadic", "数值码解码字典；dev newrez 无此表，解码在 Redshift 完成"),
                   ("port.basic_data_loan_foreclosure*", "Redshift→MySQL 间的中间表，非源表也非 BPS 直接展示表"),
                   ("port.portfunding", "融资池过滤表（JOIN 条件），非业务字段来源")]:
        body_cell(ws, r, 1, "—"); body_cell(ws, r, 2, t); ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        body_cell(ws, r, 3, why)
        r += 1

    r += 1
    r = section_title(ws, r, "样例贷款（5）", 5)
    for i, (lid, d) in enumerate(LOANS, 1):
        body_cell(ws, r, 1, f"Loan {i}"); ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        body_cell(ws, r, 2, f"{lid} —— {d}")
        r += 1


def main():
    conn = pymysql.connect(host=HOST, port=PORT, user=USER, password=PASSWORD, autocommit=True,
                           connect_timeout=20, read_timeout=180)
    cur = conn.cursor()
    wb = Workbook()
    counts = {}  # table -> (ncols, nrows_or_present)

    # 封面占位（active，已在 index 0）；按编号顺序建数据 sheet：②–⑧ 转置 → ⑨⑩⑪ 平铺 → ⑫ 视图
    transpose_main = [t for t in TRANSPOSE if t[3] != "view"]
    view_entry = [t for t in TRANSPOSE if t[3] == "view"]

    def add_transpose(entry):
        title, sch, tab, kind, desc = entry
        ws = wb.create_sheet(title)
        ws.sheet_view.showGridLines = False
        present = write_transpose(ws, sch, tab, kind, desc, cur)
        counts[tab] = (len(cols_of(cur, sch, tab)), present)
        print(f"  {title}: {counts[tab][0]} cols, 命中 {present}/5")

    for entry in transpose_main:
        add_transpose(entry)
    for title, tab, sortcol, desc in FLAT:
        ws = wb.create_sheet(title)
        ws.sheet_view.showGridLines = False
        n = write_flat(ws, tab, sortcol, desc, cur)
        counts[tab] = (len(cols_of(cur, "bpms_dev", tab)), n)
        print(f"  {title}: {counts[tab][0]} cols, {n} rows")
    for entry in view_entry:
        add_transpose(entry)

    build_cover(wb, cur, counts)  # 填充 active(index 0) 作为封面

    cur.close(); conn.close()
    wb.save(OUT)
    print(f"已保存：{OUT}")
    print(f"Sheet 数：{len(wb.worksheets)} -> {[w.title for w in wb.worksheets]}")


if __name__ == "__main__":
    main()
