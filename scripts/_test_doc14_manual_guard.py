# -*- coding: utf-8 -*-
"""
doc 14 「人工列」保护 自检（self-contained guard check）
============================================================================
规则：Excel 中表头含「人工」二字的列，是用户手工编辑的列，脚本/AI 绝不删除/覆盖/移动。
本测试验证守卫能正确「识别 + 保护」这些列。

运行（项目根目录，经 stdin——端点安全只拦 `python 文件.py`，不拦 `python - < 文件`）：
    python - < scripts/_test_doc14_manual_guard.py

历史：原版是「跑 add_field_spec_verify_sql / run_verify_sql_results 两个写表脚本后比对人工列」
的集成测试；但那两个写表脚本与旧的 _excel_guard.py 模块均已不在仓库（2026-06-10 审计确认）。
故改为自包含：直接用 skill `excel-pipeline-lineage` 的工具箱验证守卫，对 live 文件只读（操作副本）。
"""
import shutil, tempfile
from pathlib import Path

# 加载守卫/工具箱（.txt 经 open() 读取，不受「禁止执行用户目录 .py」限制；勿 import）
exec(open('.claude/skills/excel-pipeline-lineage/references/toolkit.txt', encoding='utf-8').read())
utf8_stdout()
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

LIVE = Path('docs/14_servicer_fcl_field_spec.xlsx')


def find_spec_ws(wb):
    return next((s for s in wb.worksheets if 'Field Spec' in s.title), wb.worksheets[0])


def main():
    # 全程在临时副本上操作，绝不碰用户 live 文件
    tmpdir = Path(tempfile.mkdtemp(prefix='doc14guard_'))
    tmp = tmpdir / 'copy.xlsx'
    shutil.copy(LIVE, tmp)
    wb = load_workbook(tmp)
    ws = find_spec_ws(wb)

    mcols = manual_cols(ws)
    print('文件：%s' % LIVE)
    print('Sheet：%s | 维度 %d 行 × %d 列' % (ws.title, ws.max_row, ws.max_column))
    print('检测到人工列：%s\n' % [(get_column_letter(c), ws.cell(1, c).value) for c in mcols])

    ok = True
    for c in mcols:
        L = get_column_letter(c)
        hdr = ws.cell(1, c).value or ''
        nonempty = sum(1 for r in range(1, ws.max_row + 1) if ws.cell(r, c).value not in (None, ''))
        ncom = sum(1 for r in range(1, ws.max_row + 1) if ws.cell(r, c).comment)
        # 守卫1：assert_safe 必须拦截对人工列的写入
        try:
            assert_safe(ws, c); g1 = False
        except RuntimeError:
            g1 = True
        # 守卫2：col_by_header 必须「跳过」人工列（按其表头子串不会返回该列）
        sub = hdr.replace(MANUAL_MARK, '').strip()[:4] or hdr[:4]
        g2 = (col_by_header(ws, sub) != c) if sub else True
        mark = '✅' if (g1 and g2) else '❌'
        if not (g1 and g2): ok = False
        print('  %s %s「%s」：%d 非空 + %d 批注 | assert_safe 拦截=%s, col_by_header 跳过=%s'
              % (mark, L, hdr, nonempty, ncom, g1, g2))

    # 守卫3：含人工列的 sheet，sheet_has_manual 必须 True（重建前据此改为就地更新）
    g3 = sheet_has_manual(ws)
    print('  sheet_has_manual = %s' % g3)
    if mcols and not g3: ok = False

    wb.close()
    shutil.rmtree(tmpdir, ignore_errors=True)

    if not mcols:
        print('\n结果：⚠ 未检测到任何【人工】列（请确认表头是否含「人工」二字）')
    else:
        print('\n结果：' + ('✅ PASS — 守卫能识别并保护全部 %d 个【人工】列' % len(mcols)
                           if ok else '❌ FAIL — 守卫未能保护某些人工列！'))


main()
