"""
Two targeted patches to 14_servicer_fcl_field_spec.xlsx:
  1. Fix 3 stale numbers in the 阅读指南 sheet's Newrez 现状 table.
  2. Insert a compact legend block above the compliance table in the Overview sheet.

Run from project root: python scripts/fix_numbers_and_legend.py
"""

import sys
import io
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

XLSX = Path(__file__).parent.parent / "docs" / "14_servicer_fcl_field_spec.xlsx"

C_DARK_BLUE  = "1F3864"
C_MID_BLUE   = "2E75B6"
C_LIGHT_BLUE = "D6E4F0"
C_WHITE      = "FFFFFF"
C_YELLOW     = "FFF2CC"
C_ORANGE     = "FCE4D6"
C_GREEN      = "E2EFDA"

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


# ── Patch 1: fix 阅读指南 numbers ─────────────────────────────────────────────

CORRECTIONS = {
    # key: substring of column-A cell value → {col_index: new_value}
    # col indices are 1-based; columns: A=1(name) B=2(total) C=3(ok) D=4(partial) E=5(missing)
    "2.3": {5: 5},        # ❌ 未提供: 4 → 5
    "2.4": {2: 15},       # 总字段数: 12 → 15
    "合计": {5: 14},      # ❌ 未提供: 13 → 14
}

def fix_guide_numbers(wb):
    ws = None
    for sheet in wb.worksheets:
        if "阅读指南" in sheet.title:
            ws = sheet
            break
    if ws is None:
        print("ERROR: 阅读指南 sheet not found")
        return

    fixed = []
    for row in ws.iter_rows():
        a_val = row[0].value
        if not isinstance(a_val, str):
            continue
        for key, col_map in CORRECTIONS.items():
            if key in a_val:
                for col_idx, new_val in col_map.items():
                    cell = ws.cell(row=row[0].row, column=col_idx)
                    old = cell.value
                    cell.value = new_val
                    fixed.append(f"  [阅读指南] row {row[0].row} col {col_idx}: {old!r} → {new_val!r}  (matched '{key}')")
    for msg in fixed:
        print(msg)
    if not fixed:
        print("  [阅读指南] No corrections applied (values may already be correct)")


# ── Patch 2: insert legend above Overview compliance table ────────────────────

LEGEND_ROWS = [
    # (symbol, meaning, example, bg_color)
    ("✅ 已提供",   "Newrez 当前已交付，BPS 可直接读取使用",
     "字段存在、格式正确、值有效",              C_GREEN),
    ("⚠️ 部分提供", "已交付但存在覆盖率或质量问题",
     "部分贷款缺值 / 字段名需映射 / 格式需转换", C_YELLOW),
    ("❌ 未提供",   "Newrez 完全未交付，BPS 显示空白",
     "字段在 Newrez 数据文件中不存在",          C_ORANGE),
]

def insert_legend_above_table(wb):
    ws = None
    for sheet in wb.worksheets:
        if "Overview" in sheet.title or "总览" in sheet.title:
            ws = sheet
            break
    if ws is None:
        print("ERROR: Overview sheet not found")
        return

    # Find the row where col A == "字段组 (Section)" (the table header row)
    header_row = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        v = row[0].value
        if isinstance(v, str) and "字段组" in v and "Section" in v:
            header_row = row[0].row
            break
    if header_row is None:
        print("ERROR: Could not find '字段组 (Section)' header row in Overview sheet")
        return

    # Check if legend already inserted (idempotency guard)
    check_row = header_row - 1
    if check_row >= 1:
        check_val = ws.cell(check_row, 1).value
        if isinstance(check_val, str) and "图例" in check_val:
            print("  [Overview] Legend already present — skipping insert")
            return

    # Insert 5 rows above the table header (legend_header + 3 data + 1 spacer)
    INSERT_AT = header_row
    N_ROWS = 5
    ws.insert_rows(INSERT_AT, N_ROWS)

    r = INSERT_AT

    # Row 1: legend section header
    hdr = ws.cell(r, 1, "图例说明 — Newrez 合规状态定义")
    hdr.font = Font(name="微软雅黑", bold=True, size=11, color=C_WHITE)
    hdr.fill = PatternFill("solid", fgColor=C_DARK_BLUE)
    hdr.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 20
    r += 1

    # Row 2: column sub-headers for the legend table
    sub_labels = ["符号", "含义", "典型场景", "", "", ""]
    for c, lbl in enumerate(sub_labels[:3], 1):
        cell = ws.cell(r, c, lbl)
        cell.font = Font(name="微软雅黑", bold=True, size=10, color=C_WHITE)
        cell.fill = PatternFill("solid", fgColor=C_MID_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 18
    r += 1

    # Rows 3-5: legend data
    for symbol, meaning, example, bg in LEGEND_ROWS:
        c1 = ws.cell(r, 1, symbol)
        c1.font = Font(name="微软雅黑", bold=True, size=10)
        c1.fill = PatternFill("solid", fgColor=bg)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = thin_border()

        c2 = ws.cell(r, 2, meaning)
        c2.font = Font(name="微软雅黑", size=10)
        c2.fill = PatternFill("solid", fgColor=bg)
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c2.border = thin_border()

        c3 = ws.cell(r, 3, example)
        c3.font = Font(name="微软雅黑", size=10, italic=True, color="595959")
        c3.fill = PatternFill("solid", fgColor=bg)
        c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c3.border = thin_border()
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)

        ws.row_dimensions[r].height = 20
        r += 1

    # Row N_ROWS: blank spacer (already empty from insert)
    print(f"  [Overview] Legend inserted at rows {INSERT_AT}–{r-1} (before table header)")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not XLSX.exists():
        raise FileNotFoundError(f"Not found: {XLSX}")

    wb = load_workbook(XLSX)

    print("Patch 1: Fixing 阅读指南 numbers...")
    fix_guide_numbers(wb)

    print("Patch 2: Adding legend to Overview sheet...")
    insert_legend_above_table(wb)

    wb.save(XLSX)
    print(f"Saved: {XLSX}")


if __name__ == "__main__":
    main()
