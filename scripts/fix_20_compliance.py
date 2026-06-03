"""
Fix Section 2.0 compliance numbers based on Field Spec ground truth:
  Correct: ✅=2, ⚠️=4, ❌=5  (was: ✅=1, ⚠️=5, ❌=6)
  Rate: ~17%  (was: ~8%)
  Totals: ✅=63, ⚠️=12, ❌=13  (was: 62, 13, 14)

Run: python scripts/fix_20_compliance.py
"""

import sys, io
from pathlib import Path
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

XLSX = Path(__file__).parent.parent / "docs" / "14_servicer_fcl_field_spec.xlsx"


def main():
    wb = load_workbook(XLSX)

    # ── 阅读指南 (index 0) ──────────────────────────────────────────────────
    guide = wb.worksheets[0]
    # Row 20: 四维状态基础标志 (2.0) | col C=✅, D=⚠️, E=❌
    guide.cell(20, 3).value = 2
    guide.cell(20, 4).value = 4
    guide.cell(20, 5).value = 5
    # Row 29: 合计
    guide.cell(29, 3).value = 63
    guide.cell(29, 4).value = 12
    guide.cell(29, 5).value = 13
    print("[阅读指南] Row 20 (2.0): ✅=2, ⚠️=4, ❌=5")
    print("[阅读指南] Row 29 (合计): ✅=63, ⚠️=12, ❌=13")

    # ── Overview (index 2) ──────────────────────────────────────────────────
    ov = wb.worksheets[2]
    for r in range(1, 60):
        a = ov.cell(r, 1).value
        if isinstance(a, str) and "2.0" in a and "四维" in a:
            ov.cell(r, 3).value = 2
            ov.cell(r, 4).value = 4
            ov.cell(r, 5).value = 5
            ov.cell(r, 6).value = "~17%"
            print(f"[Overview]  Row {r} (2.0): ✅=2, ⚠️=4, ❌=5, ~17%")
        if isinstance(a, str) and "合计" in a:
            ov.cell(r, 3).value = 63
            ov.cell(r, 4).value = 12
            ov.cell(r, 5).value = 13
            print(f"[Overview]  Row {r} (合计): ✅=63, ⚠️=12, ❌=13")

    wb.save(XLSX)
    print(f"\nSaved: {XLSX}")


if __name__ == "__main__":
    main()
